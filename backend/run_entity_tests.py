#!/usr/bin/env python3
"""
🤖 AUTOMATED ENTITY EXTRACTION TEST RUNNER
==========================================
Reads questions from Excel, sends to API, captures metrics, updates Excel.

Usage:
    python run_entity_tests.py                    # Run all tests
    python run_entity_tests.py --start 1 --end 10 # Run tests 1-10
    python run_entity_tests.py --question "What is DSCA?"  # Single test

Requirements:
    - Backend running on localhost:3000
    - pip install requests openpyxl pandas
"""

import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import argparse
import time
import sys
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================
# Change this to your backend IP if different
BACKEND_URL = "http://172.16.200.12:3000"
API_ENDPOINT = f"{BACKEND_URL}/api/test/query"

# Input/Output files
INPUT_EXCEL = "Entity_Testing_Top20_Important.xlsx"
OUTPUT_EXCEL = "Entity_Testing_Results_Auto.xlsx"

# Styling for Excel
GREEN_FILL = PatternFill('solid', fgColor='90EE90')
RED_FILL = PatternFill('solid', fgColor='FFB6C1')
YELLOW_FILL = PatternFill('solid', fgColor='FFFACD')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============================================================================
# API CLIENT
# ============================================================================
class SAMMApiClient:
    """Client for SAMM RAG API"""
    
    def __init__(self, base_url=BACKEND_URL):
        self.base_url = base_url
        self.session = requests.Session()
    
    def query(self, question, timeout=180):
        """
        Send query to TEST API endpoint (no auth required)
        Returns dict with all metrics
        """
        result = {
            'question': question,
            'status': 'ERROR',
            'entities': [],
            'entity_metrics': {},
            'entity_metrics_passed': {},
            'confidence': 0,
            'intent': None,
            'answer': '',
            'error': None,
            'time': 0
        }
        
        try:
            payload = {
                "question": question
            }
            
            start_time = time.time()
            
            # Use TEST endpoint - no auth required!
            response = self.session.post(
                f"{self.base_url}/api/test/query",
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=timeout
            )
            
            result['time'] = round(time.time() - start_time, 2)
            
            if response.status_code != 200:
                result['error'] = f"HTTP {response.status_code}"
                return result
            
            # Parse JSON response (not streaming)
            data = response.json()
            
            if data.get('status') == 'success':
                result['status'] = 'SUCCESS'
                result['intent'] = data.get('intent')
                result['intent_confidence'] = data.get('intent_confidence', 0)
                result['entities'] = data.get('entities', [])
                result['confidence'] = data.get('entity_confidence', 0)
                result['entity_metrics'] = data.get('entity_metrics', {})
                result['entity_metrics_passed'] = data.get('entity_metrics_passed', {})
            else:
                result['error'] = data.get('error', 'Unknown error')
            
        except requests.exceptions.Timeout:
            result['error'] = 'Request timed out'
            result['status'] = 'TIMEOUT'
        except requests.exceptions.ConnectionError:
            result['error'] = 'Connection failed - is backend running?'
            result['status'] = 'CONNECTION_ERROR'
        except Exception as e:
            result['error'] = str(e)
            result['status'] = 'ERROR'
        
        return result


# ============================================================================
# TEST RUNNER
# ============================================================================
class EntityTestRunner:
    """Automated test runner for entity extraction"""
    
    def __init__(self, input_file=INPUT_EXCEL, output_file=OUTPUT_EXCEL):
        self.input_file = input_file
        self.output_file = output_file
        self.client = SAMMApiClient()
        self.results = []
        self.test_questions = []
    
    def load_questions(self):
        """Load test questions from Excel"""
        print("\n📂 Loading test questions from Excel...")
        
        df = pd.read_excel(self.input_file, sheet_name='Test Data', header=1)
        
        questions = []
        for idx, row in df.iterrows():
            test_id = row.iloc[0]
            question = row.iloc[1]
            expected = row.iloc[2]
            
            if pd.notna(test_id) and str(test_id).startswith('ET'):
                questions.append({
                    'test_id': test_id,
                    'question': question,
                    'expected_entities': expected,
                    'row_index': idx + 3  # Excel row (header=2, 0-indexed)
                })
        
        self.test_questions = questions
        print(f"✅ Loaded {len(questions)} test questions")
        return questions
    
    def run_single_test(self, question_data):
        """Run a single test"""
        test_id = question_data['test_id']
        question = question_data['question']
        
        print(f"\n🧪 [{test_id}] Testing: {question[:50]}...")
        
        result = self.client.query(question)
        result['test_id'] = test_id
        result['row_index'] = question_data['row_index']
        
        # Determine pass/fail based on metrics
        metrics = result.get('entity_metrics', {})
        passed = result.get('entity_metrics_passed', {})
        
        if result['status'] == 'SUCCESS' and metrics:
            precision = metrics.get('precision', 0)
            recall = metrics.get('recall', 0)
            f1 = metrics.get('f1_score', 0)
            halluc = metrics.get('hallucination_rate', 1)
            
            all_passed = all([
                passed.get('precision', False),
                passed.get('recall', False),
                passed.get('f1_score', False),
                passed.get('hallucination_rate', False)
            ])
            
            result['test_status'] = 'PASS' if all_passed else 'PARTIAL'
            
            status_icon = '✅' if all_passed else '⚠️'
            print(f"   {status_icon} P:{precision:.0%} R:{recall:.0%} F1:{f1:.0%} H:{halluc:.0%} [{result['time']}s]")
        else:
            result['test_status'] = 'ERROR'
            print(f"   ❌ {result.get('error', 'Unknown error')}")
        
        return result
    
    def run_tests(self, start=1, end=None, delay=1.0):
        """Run multiple tests"""
        if not self.test_questions:
            self.load_questions()
        
        if end is None:
            end = len(self.test_questions)
        
        # Adjust for 1-based indexing
        start_idx = max(0, start - 1)
        end_idx = min(len(self.test_questions), end)
        
        questions_to_test = self.test_questions[start_idx:end_idx]
        
        print(f"\n{'='*70}")
        print(f"🚀 RUNNING ENTITY EXTRACTION TESTS")
        print(f"   Tests: {start} to {end} ({len(questions_to_test)} questions)")
        print(f"   Backend: {self.client.base_url}")
        print(f"{'='*70}")
        
        results = []
        passed = 0
        partial = 0
        failed = 0
        
        for i, q in enumerate(questions_to_test, 1):
            result = self.run_single_test(q)
            results.append(result)
            
            if result['test_status'] == 'PASS':
                passed += 1
            elif result['test_status'] == 'PARTIAL':
                partial += 1
            else:
                failed += 1
            
            # Progress
            print(f"   Progress: {i}/{len(questions_to_test)} | ✅{passed} ⚠️{partial} ❌{failed}")
            
            # Delay between requests
            if i < len(questions_to_test):
                time.sleep(delay)
        
        self.results = results
        
        # Summary
        print(f"\n{'='*70}")
        print(f"📊 TEST SUMMARY")
        print(f"{'='*70}")
        print(f"   Total:   {len(results)}")
        print(f"   Passed:  {passed} ({passed/len(results)*100:.0f}%)")
        print(f"   Partial: {partial} ({partial/len(results)*100:.0f}%)")
        print(f"   Failed:  {failed} ({failed/len(results)*100:.0f}%)")
        print(f"{'='*70}")
        
        return results
    
    def update_excel(self):
        """Update Excel with test results"""
        if not self.results:
            print("❌ No results to save")
            return
        
        print(f"\n📝 Updating Excel: {self.output_file}")
        
        # Load workbook
        try:
            wb = load_workbook(self.output_file)
        except:
            wb = load_workbook(self.input_file)
        
        sheet = wb['Test Data']
        
        # Column mapping (based on Excel structure)
        # A=Test_ID, B=Question, C=Expected, D=Count, E=Types, F=Importance
        # G=Actual_Entities, H=True_Positives, I=False_Positives, J=False_Negatives
        # K=Precision, L=Recall, M=F1_Score, N=Hallucination, O=Confidence, P=Status, Q=Notes
        
        updated = 0
        for result in self.results:
            row = result.get('row_index')
            if not row:
                continue
            
            metrics = result.get('entity_metrics', {})
            
            # G = Actual Entities
            entities = result.get('entities', [])
            sheet[f'G{row}'] = ', '.join(entities) if entities else ''
            sheet[f'G{row}'].border = THIN_BORDER
            
            # K = Precision
            if 'precision' in metrics:
                sheet[f'K{row}'] = metrics['precision']
                sheet[f'K{row}'].number_format = '0%'
                sheet[f'K{row}'].border = THIN_BORDER
            
            # L = Recall
            if 'recall' in metrics:
                sheet[f'L{row}'] = metrics['recall']
                sheet[f'L{row}'].number_format = '0%'
                sheet[f'L{row}'].border = THIN_BORDER
            
            # M = F1 Score
            if 'f1_score' in metrics:
                sheet[f'M{row}'] = metrics['f1_score']
                sheet[f'M{row}'].number_format = '0%'
                sheet[f'M{row}'].border = THIN_BORDER
            
            # N = Hallucination
            if 'hallucination_rate' in metrics:
                sheet[f'N{row}'] = metrics['hallucination_rate']
                sheet[f'N{row}'].number_format = '0%'
                sheet[f'N{row}'].border = THIN_BORDER
            
            # O = Confidence
            conf = result.get('confidence', 0)
            sheet[f'O{row}'] = conf
            sheet[f'O{row}'].number_format = '0%'
            sheet[f'O{row}'].border = THIN_BORDER
            
            # P = Status
            status = result.get('test_status', 'ERROR')
            sheet[f'P{row}'] = status
            sheet[f'P{row}'].border = THIN_BORDER
            sheet[f'P{row}'].font = Font(bold=True)
            
            if status == 'PASS':
                sheet[f'P{row}'].fill = GREEN_FILL
            elif status == 'PARTIAL':
                sheet[f'P{row}'].fill = YELLOW_FILL
            else:
                sheet[f'P{row}'].fill = RED_FILL
            
            # Q = Notes
            notes = f"Time: {result.get('time', 0)}s"
            if result.get('error'):
                notes = f"Error: {result['error']}"
            sheet[f'Q{row}'] = notes
            sheet[f'Q{row}'].border = THIN_BORDER
            
            updated += 1
        
        # Save
        wb.save(self.output_file)
        print(f"✅ Updated {updated} rows in {self.output_file}")
        
        return self.output_file


# ============================================================================
# MAIN
# ============================================================================
def main():
    parser = argparse.ArgumentParser(description='Automated Entity Extraction Test Runner')
    parser.add_argument('--start', type=int, default=1, help='Start test number (1-based)')
    parser.add_argument('--end', type=int, default=None, help='End test number')
    parser.add_argument('--input', type=str, default=INPUT_EXCEL, help='Input Excel file')
    parser.add_argument('--output', type=str, default=OUTPUT_EXCEL, help='Output Excel file')
    parser.add_argument('--delay', type=float, default=1.0, help='Delay between tests (seconds)')
    parser.add_argument('--question', type=str, default=None, help='Test a single question')
    parser.add_argument('--no-save', action='store_true', help='Do not save to Excel')
    parser.add_argument('--url', type=str, default=BACKEND_URL, help='Backend URL (default: http://172.16.200.12:3000)')
    
    args = parser.parse_args()
    
    # Use provided URL
    backend_url = args.url
    
    # Check backend
    print("\n🔍 Checking backend connection...")
    try:
        response = requests.get(f"{backend_url}/api/health", timeout=5)
        if response.status_code == 200:
            print(f"✅ Backend is running at {backend_url}")
        else:
            print(f"⚠️ Backend returned status {response.status_code}")
    except Exception as e:
        print(f"⚠️ Could not connect to {backend_url}")
        print(f"   Error: {e}")
        print("   Make sure the backend is running!")
        print("   Run: python app_5_0_E1_3.py")
        sys.exit(1)
    
    # Single question mode
    if args.question:
        print(f"\n🧪 Testing single question: {args.question}")
        client = SAMMApiClient(backend_url)
        result = client.query(args.question)
        
        print(f"\n{'='*70}")
        print("RESULT:")
        print(f"{'='*70}")
        print(f"Status: {result['status']}")
        print(f"Entities: {result['entities']}")
        print(f"Metrics: {result['entity_metrics']}")
        print(f"Passed: {result['entity_metrics_passed']}")
        print(f"Time: {result['time']}s")
        
        if result.get('error'):
            print(f"Error: {result['error']}")
        
        return
    
    # Initialize runner with correct URL
    runner = EntityTestRunner(args.input, args.output)
    runner.client = SAMMApiClient(backend_url)
    
    # Load questions
    runner.load_questions()
    
    # Run tests
    runner.run_tests(start=args.start, end=args.end, delay=args.delay)
    
    # Save results
    if not args.no_save:
        runner.update_excel()
        print(f"\n📁 Results saved to: {args.output}")


if __name__ == "__main__":
    main()
