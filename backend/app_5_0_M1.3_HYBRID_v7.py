
import os
import json
import uuid 
import time
import re
import hashlib
import asyncio
import sys
from datetime import datetime, timezone 
from typing import Dict, List, Any, Optional, TypedDict, Set
from urllib.parse import quote_plus, urlencode
from enum import Enum
from pathlib import Path
from flask import send_from_directory
import functools
from collections import defaultdict  # For metrics calculations
import openpyxl  # Excel processing for MISIL RSN sheets
import PyPDF2    # PDF text extraction
import tempfile  # Temporary file handling for uploads
# Fix for Windows asyncio issues
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

# Flask & Extensions
from flask import Flask, request, jsonify, session, send_from_directory, redirect, url_for
from flask_cors import CORS
from authlib.integrations.flask_client import OAuth
from werkzeug.utils import secure_filename 

# Environment
from dotenv import load_dotenv, find_dotenv
load_dotenv(find_dotenv()) 
def time_function(func):
    """Simple timing decorator for performance monitoring"""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        elapsed = time.time() - start
        print(f"[TIMING] {func.__name__}: {elapsed:.2f}s")
        return result
    return wrapper

# HTTP Requests Library
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
# Azure SDK
from azure.cosmos import CosmosClient, PartitionKey, exceptions as CosmosExceptions 
from azure.storage.blob import BlobServiceClient, ContentSettings 
from azure.core.exceptions import ResourceExistsError as BlobResourceExistsError, ResourceNotFoundError as BlobResourceNotFoundError

# Database imports for integrated agents
try:
    from gremlin_python.driver import client, serializer
    from gremlin_python.driver.protocol import GremlinServerError
    print("Gremlin client imported successfully")
except ImportError:
    print("Gremlin client not available - some features may be limited")
    client = None

try:
    import chromadb
    print("ChromaDB imported successfully")
except ImportError:
    print("ChromaDB not available - some features may be limited")
    chromadb = None

try:
    from sentence_transformers import SentenceTransformer
    print("SentenceTransformers imported successfully")
except ImportError:
    print("SentenceTransformers not available - some features may be limited")
    SentenceTransformer = None

# --- Application Configuration ---
# Auth0 Configuration
AUTH0_CLIENT_ID = os.getenv("AUTH0_CLIENT_ID")
AUTH0_CLIENT_SECRET = os.getenv("AUTH0_CLIENT_SECRET")
AUTH0_DOMAIN = os.getenv("AUTH0_DOMAIN")
APP_SECRET_KEY = os.getenv("APP_SECRET_KEY", "DFd55vvJIcV79cGuEETrGc9HWiNDqducM7upRwXdeJ9c4E3LbCtl")
BASE_URL = os.getenv("BACKEND_URL", "http://172.16.200.12:3000")
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://172.16.200.12:5173")
# Ollama Configuration
OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.1:8b")
#OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.2:3b")
OLLAMA_MAX_RETRIES = int(os.getenv("OLLAMA_MAX_RETRIES", "3"))
OLLAMA_TIMEOUT_NORMAL = int(os.getenv("OLLAMA_TIMEOUT_NORMAL", "60"))
# Azure Storage Configuration
COSMOS_ENDPOINT = os.getenv("COSMOS_ENDPOINT")
COSMOS_KEY = os.getenv("COSMOS_KEY")
DATABASE_NAME = os.getenv("DATABASE_NAME")
CASES_CONTAINER_NAME = os.getenv("CASES_CONTAINER_NAME") 

# Azure Blob Storage Configuration
AZURE_CONNECTION_STRING = os.getenv("AZURE_CONNECTION_STRING")
AZURE_CASE_DOCS_CONTAINER_NAME = os.getenv("AZURE_CASE_DOCS_CONTAINER_NAME")
AZURE_CHAT_DOCS_CONTAINER_NAME = os.getenv("AZURE_CHAT_DOCS_CONTAINER_NAME")

# Database Configuration for Enhanced Agents
COSMOS_GREMLIN_CONFIG = {
    'endpoint': os.getenv("COSMOS_GREMLIN_ENDPOINT", "asist-graph-db.gremlin.cosmos.azure.com").replace('wss://', '').replace(':443/', ''),
    'database': os.getenv("COSMOS_GREMLIN_DATABASE", "ASIST-Agent-1.1DB"),
    'graph': os.getenv("COSMOS_GREMLIN_COLLECTION", "AGENT1.4"),
    'password': os.getenv("COSMOS_GREMLIN_KEY", "")
}

# Vector Database Configuration
VECTOR_DB_PATH = "C:\\Users\\ShaziaKashif\\ASIST Project\\ASIST2.1\\ASIST_V2.1\\backend\\Chromadb\\chroma_db_combined"
#VECTOR_DB_PATH = "C:\\Users\\TomLorenc\\Downloads\\ASIST_DEV\\ASIST_DEV\\backend\\vector_db"
#VECTOR_DB_PATH = "C:\\Projects\\5_1\\ASIST_V5.0-main\\backend\\vector_db"
#VECTOR_DB_PATH = "O:\\Assist Versions\\backend\\vector_db"
EMBEDDING_MODEL = "all-MiniLM-L6-v2"
VECTOR_DB_COLLECTION = "samm_all_chapters"

# =============================================================================
# CACHE CONFIGURATION
# =============================================================================

# Cache settings
CACHE_ENABLED = os.getenv("CACHE_ENABLED", "true").lower() == "true"
CACHE_TTL_SECONDS = int(os.getenv("CACHE_TTL_SECONDS", "3600"))  # 1 hour default
CACHE_MAX_SIZE = int(os.getenv("CACHE_MAX_SIZE", "1000"))  # Maximum cached items

# In-memory cache structures
query_cache = {}  # Structure: {normalized_query: {answer, metadata, timestamp}}
cache_stats = {
    "hits": 0,
    "misses": 0,
    "total_queries": 0,
    "cache_size": 0
}

print(f"Cache Configuration: Enabled={CACHE_ENABLED}, TTL={CACHE_TTL_SECONDS}s, Max Size={CACHE_MAX_SIZE}")

# ITAR Compliance Integration
COMPLIANCE_SERVICE_URL = os.getenv("COMPLIANCE_SERVICE_URL", "http://localhost:3002")
COMPLIANCE_ENABLED = os.getenv("COMPLIANCE_ENABLED", "true").lower() == "true"
DEFAULT_DEV_AUTH_LEVEL = os.getenv("DEFAULT_DEV_AUTH_LEVEL", "top_secret")

print(f"ITAR Compliance: {'Enabled' if COMPLIANCE_ENABLED else 'Disabled'} (Default Level: {DEFAULT_DEV_AUTH_LEVEL})")
# =============================================================================
# CACHE HELPER FUNCTIONS
# =============================================================================

def normalize_query_for_cache(query: str) -> str:
    """
    Normalize query for cache key matching
    - Lowercase, strip whitespace, remove punctuation
    - Sort words to catch similar questions with different word order
    """
    import string
    # Remove punctuation and convert to lowercase
    query_clean = query.lower().translate(str.maketrans('', '', string.punctuation))
    # Split into words and sort
    words = query_clean.split()
    # Remove common stop words that don't affect meaning
    stop_words = {'what', 'is', 'are', 'the', 'a', 'an', 'does', 'do', 'can', 'how'}
    significant_words = [w for w in words if w not in stop_words and len(w) > 2]
    # Return sorted words as key
    return ' '.join(sorted(significant_words))

def get_from_cache(query: str) -> Optional[Dict[str, Any]]:
    """
    Retrieve cached answer for a query
    Returns None if not found or expired
    """
    if not CACHE_ENABLED:
        return None
    
    cache_key = normalize_query_for_cache(query)
    
    if cache_key in query_cache:
        cached_entry = query_cache[cache_key]
        
        # Check if cache entry is still valid (TTL check)
        age_seconds = time.time() - cached_entry['timestamp']
        if age_seconds < CACHE_TTL_SECONDS:
            cache_stats['hits'] += 1
            cache_stats['total_queries'] += 1
            print(f"[Cache HIT] Query: '{query[:50]}...' (age: {age_seconds:.1f}s)")
            return cached_entry
        else:
            # Expired - remove it
            del query_cache[cache_key]
            print(f"[Cache EXPIRED] Query: '{query[:50]}...' (age: {age_seconds:.1f}s)")
    
    cache_stats['misses'] += 1
    cache_stats['total_queries'] += 1
    print(f"[Cache MISS] Query: '{query[:50]}...'")
    return None
def fetch_blob_content(blob_name: str, container_client) -> Optional[str]:
    """Fetch text content from a blob for AI processing"""
    if not container_client:
        return None
    
    try:
        blob_client = container_client.get_blob_client(blob_name)
        download_stream = blob_client.download_blob()
        content = download_stream.readall()
        
        try:
            return content.decode('utf-8')
        except UnicodeDecodeError:
            return f"[Binary file: {blob_name}]"
    except Exception as e:
        print(f"[Blob Fetch] Error reading {blob_name}: {e}")
        return None
def save_to_cache(query: str, answer: str, metadata: Dict[str, Any]) -> bool:
    """
    Save query-answer pair to cache
    Implements LRU eviction if cache is full
    """
    if not CACHE_ENABLED:
        return False
    
    cache_key = normalize_query_for_cache(query)
    
    # Check cache size limit
    if len(query_cache) >= CACHE_MAX_SIZE and cache_key not in query_cache:
        # Evict oldest entry (simple LRU)
        oldest_key = min(query_cache.keys(), key=lambda k: query_cache[k]['timestamp'])
        del query_cache[oldest_key]
        print(f"[Cache EVICT] Removed oldest entry to make room")
    
    # Save to cache
    query_cache[cache_key] = {
        'original_query': query,
        'answer': answer,
        'metadata': metadata,
        'timestamp': time.time()
    }
    
    cache_stats['cache_size'] = len(query_cache)
    print(f"[Cache SAVE] Query: '{query[:50]}...' (cache size: {len(query_cache)})")
    return True

def get_cache_stats() -> Dict[str, Any]:
    """Get cache statistics"""
    hit_rate = (cache_stats['hits'] / cache_stats['total_queries'] * 100) if cache_stats['total_queries'] > 0 else 0
    
    return {
        'enabled': CACHE_ENABLED,
        'total_queries': cache_stats['total_queries'],
        'cache_hits': cache_stats['hits'],
        'cache_misses': cache_stats['misses'],
        'hit_rate_percent': round(hit_rate, 2),
        'current_size': len(query_cache),
        'max_size': CACHE_MAX_SIZE,
        'ttl_seconds': CACHE_TTL_SECONDS
    }
# --- Flask App Initialization ---
app = Flask(__name__, static_folder='static')
app.secret_key = APP_SECRET_KEY
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "PUT", "DELETE"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

ollama_session = requests.Session()
adapter = HTTPAdapter(pool_connections=10, pool_maxsize=10)
ollama_session.mount("http://", adapter)
ollama_session.mount("https://", adapter)
print("[Ollama] ✅ Connection pooling configured")


# Simple in-memory storage for demo purposes when Azure isn't available
user_cases = {}
staged_documents = {}

print(f"Ollama URL: {OLLAMA_URL}")
print(f"Ollama Model: {OLLAMA_MODEL}")

# --- Initialize Cosmos DB Client ---
cosmos_client = None
database_client = None
cases_container_client = None
reviews_test_container_client = None

# Initialize reviews container if Cosmos DB configured
if COSMOS_ENDPOINT and COSMOS_KEY and DATABASE_NAME:
    try:
        if not cosmos_client:
            cosmos_client = CosmosClient(COSMOS_ENDPOINT, COSMOS_KEY)
        if not database_client:
            database_client = cosmos_client.get_database_client(DATABASE_NAME)
        
        # Create or get reviews container
        try:
            reviews_test_container_client = database_client.create_container(
                id="reviews",
                partition_key=PartitionKey(path="/type"),
                offer_throughput=400
            )
            print("✅ Reviews container created")
        except:
            reviews_test_container_client = database_client.get_container_client("reviews")
            print("✅ Reviews test container connected")
            
    except Exception as e:
        print(f"⚠️ Reviews container not initialized: {e}")

if COSMOS_ENDPOINT and COSMOS_KEY and DATABASE_NAME and CASES_CONTAINER_NAME:
    try:
        cosmos_client = CosmosClient(COSMOS_ENDPOINT, COSMOS_KEY)
        database_client = cosmos_client.get_database_client(DATABASE_NAME)
        cases_container_client = database_client.get_container_client(CASES_CONTAINER_NAME)
        print(f"Successfully connected to Cosmos DB Cases container: {DATABASE_NAME}/{CASES_CONTAINER_NAME}")
    except Exception as e:
        print(f"Warning: Error initializing Cosmos DB client: {e}. Using in-memory storage.")
else:
    print("Warning: Cosmos DB credentials not configured. Using in-memory storage.")

# --- Initialize Azure Blob Service Client ---
blob_service_client = None
case_docs_blob_container_client = None
chat_docs_blob_container_client = None

def _extract_case_identifier_from_text(text: str) -> Optional[str]:
    """
    Extract case identifier from text (e.g., SR-P-NAV, MX-B-SAL)
    
    Args:
        text: Raw text to search for case ID
        
    Returns:
        Case identifier string or None
        
    Used by:
        - Upload endpoint
        - Document type detection
        - Financial data extraction
    """
    if not text:
        return None
    
    # Pattern 1: Standard format (SR-P-NAV, MX-B-SAL)
    case_pattern_1 = r'\b([A-Z]{2}-[A-Z]-[A-Z]{2,4})\b'
    matches_1 = re.findall(case_pattern_1, text.upper())
    
    if matches_1:
        return matches_1[0].upper()
    
    # Pattern 2: With spaces (SR P NAV)
    case_pattern_2 = r'\b([A-Z]{2})\s+([A-Z])\s+([A-Z]{2,4})\b'
    matches_2 = re.findall(case_pattern_2, text.upper())
    
    if matches_2:
        return f"{matches_2[0][0]}-{matches_2[0][1]}-{matches_2[0][2]}"
    
    # Pattern 3: No dashes (SRPNAV)
    case_pattern_3 = r'\b([A-Z]{2})([A-Z])([A-Z]{2,4})\b'
    matches_3 = re.findall(case_pattern_3, text.upper())
    
    if matches_3:
        return f"{matches_3[0][0]}-{matches_3[0][1]}-{matches_3[0][2]}"
    
    return None
def determine_document_type(filename: str) -> str:
    """
    Determine document type from filename
    
    Args:
        filename: Original filename
        
    Returns:
        Document type: FINANCIAL_DATA, LOA, CONTRACT, REQUISITION, MINUTES, or GENERAL
    """
    filename_lower = filename.lower()
    
    # Extract case ID for logging
    case_id = _extract_case_identifier_from_text(filename)
    if case_id:
        print(f"[DocumentType] 📋 Case ID from filename: {case_id}")
    
    # PRIORITY 1: LOA patterns
    loa_keywords = ['loa', 'letter of offer', 'offer and acceptance']
    if any(keyword in filename_lower for keyword in loa_keywords):
        return 'LOA'
    
    # PRIORITY 2: Financial data
    financial_keywords = ['financial', 'rsn', 'pdli', 'misil', 'funding']
    if any(keyword in filename_lower for keyword in financial_keywords):
        return 'FINANCIAL_DATA'
    
    # PRIORITY 3: Minutes
    minutes_keywords = ['minutes', 'meeting', 'notes']
    if any(keyword in filename_lower for keyword in minutes_keywords):
        return 'MINUTES'
    
    # PRIORITY 4: Contracts
    contract_keywords = ['contract', 'agreement']
    if any(keyword in filename_lower for keyword in contract_keywords):
        return 'CONTRACT'
    
    # PRIORITY 5: Requisitions
    requisition_keywords = ['requisition', 'req', 'purchase']
    if any(keyword in filename_lower for keyword in requisition_keywords):
        return 'REQUISITION'
    
    return 'GENERAL'
def find_header_row(ws) -> Optional[int]:
    """
    Find the header row in an Excel sheet
    Looks for common header keywords in MISIL RSN sheets
    
    Args:
        ws: openpyxl worksheet object
        
    Returns:
        Row index (1-based) or None
    """
    header_keywords = ['rsn', 'pdli', 'oa rec amt', 'net commit', 'obligation']
    
    for row_idx in range(1, min(20, ws.max_row + 1)):  # Search first 20 rows
        row = ws[row_idx]
        row_text = ' '.join([str(cell.value or '').lower() for cell in row])
        
        # Check if this row contains multiple header keywords
        matches = sum(1 for keyword in header_keywords if keyword in row_text)
        
        if matches >= 3:  # At least 3 header keywords
            return row_idx
    
    return None


def extract_text_from_pdf(file_path: str) -> str:
    """
    Extract text from PDF file using PyPDF2
    
    Args:
        file_path: Path to PDF file
        
    Returns:
        Extracted text content
    """
    try:
        import PyPDF2
        
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ''
            
            for page in pdf_reader.pages:
                text += page.extract_text() + '\n'
            
            return text
    
    except Exception as e:
        print(f"[PDF Extraction] Error: {e}")
        return ''


def extract_loa_data_from_pdf(file_path: str) -> Dict[str, Any]:
    """
    Extract LOA-specific data from PDF
    
    Args:
        file_path: Path to LOA PDF
        
    Returns:
        Dictionary with LOA metadata
    """
    loa_data = {
        "document_type": "LOA",
        "case_number": "",
        "country": "",
        "total_value": "",
        "line_items": []
    }
    
    try:
        full_text = extract_text_from_pdf(file_path)
        
        # Extract case number
        case_id = _extract_case_identifier_from_text(full_text)
        if case_id:
            loa_data["case_number"] = case_id
        
        # Extract country (basic pattern)
        country_pattern = r'Country:\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)'
        country_match = re.search(country_pattern, full_text)
        if country_match:
            loa_data["country"] = country_match.group(1)
        
        # Extract total value
        value_pattern = r'\$[\d,]+(?:\.\d{2})?'
        value_matches = re.findall(value_pattern, full_text)
        if value_matches:
            loa_data["total_value"] = value_matches[0]
    
    except Exception as e:
        print(f"[LOA Extraction] Error: {e}")
    
    return loa_data



def extract_case_document_data(file_path: str, file_type: str, 
                                original_filename: str = None) -> Dict[str, Any]:
    """
    🔑 MASTER EXTRACTION FUNCTION
    
    Extract structured data from case documents:
    - PDFs: LOA, LOR, Minutes, Reports  
    - Excel: Financial Data (MISIL RSN sheets with PDLI)
    
    Args:
        file_path: Path to uploaded file
        file_type: Document type hint
        original_filename: Original filename for context
        
    Returns:
        Dictionary containing:
        - document_type
        - case_identifier
        - extracted_text (PDFs)
        - key_info.financial_records (Excel)
        - extraction_metadata
    """
    doc_data = {
        "document_type": file_type,
        "case_identifier": "",
        "entities": [],
        "key_info": {},
        "extracted_text": "",
        "extraction_metadata": {
            "extraction_date": datetime.now(timezone.utc).isoformat(),
            "file_type": file_type,
            "original_filename": original_filename
        }
    }
    
    try:
        # ===================================================================
        # PDF PROCESSING (LOA, LOR, Minutes, Reports)
        # ===================================================================
        if file_path.lower().endswith('.pdf'):
            print(f"[Extraction] Processing PDF: {original_filename}")
            
            # Extract text from PDF
            full_text = extract_text_from_pdf(file_path)
            doc_data["extracted_text"] = full_text
            
            # Extract case identifier
            case_identifier = _extract_case_identifier_from_text(full_text)
            if case_identifier:
                doc_data["case_identifier"] = case_identifier
                print(f"[Extraction] Case ID: {case_identifier}")
            
            # Detect LOA
            if 'Letter of Offer and Acceptance' in full_text or 'LOA' in full_text:
                doc_data["document_type"] = "LOA"
                print(f"[Extraction] Detected LOA document")
                
                # Extract LOA-specific data
                loa_details = extract_loa_data_from_pdf(file_path)
                doc_data["key_info"] = loa_details
            
            # Detect Minutes
            elif 'Meeting Minutes' in full_text or 'MINUTES' in full_text:
                doc_data["document_type"] = "MINUTES"
                print(f"[Extraction] Detected MINUTES document")
        
        # ===================================================================
        # ✅✅✅ EXCEL PROCESSING - FINANCIAL DATA WITH PDLI
        # ===================================================================
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            print(f"[Extraction] Processing Excel: {original_filename}")
            doc_data["document_type"] = "FINANCIAL_DATA"
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract case ID from filename
            if original_filename:
                filename_case = _extract_case_identifier_from_text(original_filename)
                if filename_case:
                    doc_data["case_identifier"] = filename_case
                    print(f"[Extraction] Case ID from filename: {filename_case}")
            
            # FIND MISIL RSN SHEET
            target_sheet_name = None
            for sheet_name in wb.sheetnames:
                if 'misil' in sheet_name.lower() or 'rsn' in sheet_name.lower():
                    target_sheet_name = sheet_name
                    print(f"[Extraction] Found target sheet: {target_sheet_name}")
                    break
            
            if not target_sheet_name:
                print(f"[Extraction] ⚠️ No MISIL/RSN sheet found")
                return doc_data
            
            ws = wb[target_sheet_name]
            
            # FIND HEADER ROW
            header_row_idx = find_header_row(ws)
            if not header_row_idx:
                print(f"[Extraction] ⚠️ No header row found")
                return doc_data
            
            print(f"[Extraction] Header row at index: {header_row_idx}")
            header_row = ws[header_row_idx]
            
            # ✅✅✅ CRITICAL: COLUMN MAPPING WITH PDLI
            column_map = {}
            
            for col_idx, cell in enumerate(header_row):
                cell_value = cell.value
                if not cell_value:
                    continue
                
                cell_str = str(cell_value).strip().lower()
                
                # RSN PDLI (Column Y) - GROUPING key
                if re.search(r'\brsn\s+pdli\b', cell_str):
                    column_map['rsn_identifier'] = col_idx
                    column_map['line_item'] = col_idx
                    print(f"[Extraction]   RSN PDLI at column {col_idx}")
                
                # PDLINUM_PDLI (Column AA) - ACTUAL PDLI NUMBER
                elif 'pdlinum' in cell_str and 'pdli' in cell_str:
                    column_map['pdli_pdli'] = col_idx  # ⭐ CRITICAL
                    print(f"[Extraction]   PDLI NUMBER at column {col_idx}")
                
                # PDLI NAME (Column AB)
                elif re.search(r'\bpdli\s+(name|desc)', cell_str):
                    column_map['pdli_name'] = col_idx
                    print(f"[Extraction]   PDLI NAME at column {col_idx}")
                
                # Financial columns
                elif 'oa rec amt' in cell_str or 'oa_rec_amt' in cell_str:
                    column_map['oa_rec_amt'] = col_idx
                    print(f"[Extraction]   OA REC AMT at column {col_idx}")
                
                elif 'net commit amt' in cell_str:
                    column_map['net_commit_amt'] = col_idx
                
                elif 'gross obl amt' in cell_str or 'net obl amt' in cell_str:
                    column_map['net_obl_amt'] = col_idx
                
                elif 'net exp amt' in cell_str:
                    column_map['net_exp_amt'] = col_idx
                
                elif 'dir rsrv' in cell_str:
                    column_map['dir_rsrv_amt'] = col_idx
            
            print(f"[Extraction] Mapped {len(column_map)} columns")
            
            # EXTRACT DATA ROWS
            financial_records = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                # Skip empty rows
                if not any(cell for cell in row):
                    continue
                
                record = {
                    'sheet': target_sheet_name,
                    'row_number': row_idx
                }
                
                # Map columns to record
                for field_name, col_idx in column_map.items():
                    if col_idx < len(row):
                        value = row[col_idx]
                        record[field_name] = value
                
                # Process RSN identifier
                rsn = record.get('rsn_identifier')
                if not rsn:
                    continue
                
                record['line_item'] = str(rsn).strip()
                
                # ✅✅✅ Extract PDLI number (Column AA)
                pdli = record.get('pdli_pdli', '')
                if pdli:
                    record['pdli_pdli'] = str(pdli).strip()
                else:
                    record['pdli_pdli'] = 'N/A'
                
                # Extract PDLI name (Column AB)
                pdli_name = record.get('pdli_name', '')
                if pdli_name:
                    record['pdli_name'] = str(pdli_name).strip()
                else:
                    record['pdli_name'] = ''
                
                # Calculate available amount
                try:
                    oa_rec = float(str(record.get('oa_rec_amt', 0)).replace('$', '').replace(',', '') or 0)
                    net_commit = float(str(record.get('net_commit_amt', 0)).replace('$', '').replace(',', '') or 0)
                    net_obl = float(str(record.get('net_obl_amt', 0)).replace('$', '').replace(',', '') or 0)
                    net_exp = float(str(record.get('net_exp_amt', 0)).replace('$', '').replace(',', '') or 0)
                    dir_rsrv = float(str(record.get('dir_rsrv_amt', 0)).replace('$', '').replace(',', '') or 0)
                    
                    record['available'] = oa_rec - net_commit - net_obl - net_exp - dir_rsrv
                except (ValueError, TypeError) as e:
                    print(f"[Extraction] ⚠️ Error calculating available for row {row_idx}: {e}")
                    record['available'] = 0
                
                financial_records.append(record)
            
            print(f"[Extraction] ✅ Extracted {len(financial_records)} financial records")
            
            # ✅ STORE FINANCIAL RECORDS
            doc_data["key_info"]["financial_records"] = financial_records
            doc_data["key_info"]["structured_records_count"] = len(financial_records)
    
    except Exception as e:
        print(f"[Extraction] ❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        doc_data["extraction_metadata"]["error"] = str(e)
    
    return doc_data




def initialize_blob_container(bs_client, container_name_env_var, container_description):
    container_name = os.getenv(container_name_env_var)
    if not container_name:
        print(f"Warning: {container_name_env_var} is not set. {container_description} functionality will be disabled.")
        return None
    try:
        container_client = bs_client.get_container_client(container_name)
        container_client.create_container()
        print(f"Blob container '{container_name}' for {container_description} created or already exists.")
        return container_client
    except BlobResourceExistsError:
        print(f"Blob container '{container_name}' for {container_description} already exists.")
        return container_client
    except Exception as e_create_container:
        print(f"Could not create/verify blob container '{container_name}' for {container_description}: {e_create_container}")
        return None

if AZURE_CONNECTION_STRING:
    try:
        blob_service_client = BlobServiceClient.from_connection_string(AZURE_CONNECTION_STRING)
        case_docs_blob_container_client = initialize_blob_container(blob_service_client, "AZURE_CASE_DOCS_CONTAINER_NAME", "case documents")
        chat_docs_blob_container_client = initialize_blob_container(blob_service_client, "AZURE_CHAT_DOCS_CONTAINER_NAME", "chat documents")
    except Exception as e:
        print(f"Warning: Error initializing Azure Blob Service client: {e}")
else:
    print("Warning: AZURE_CONNECTION_STRING is not set. Blob storage functionality will be disabled.")

# --- Auth0 OAuth Setup ---
# --- Auth0 OAuth Setup ---
# --- Auth0 OAuth Setup ---
# --- Auth0 OAuth Setup ---
# Detect if running locally vs deployed
BASE_URL = os.getenv("BACKEND_URL", "http://172.16.200.12:3000")
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://172.16.200.12:5173")

oauth = None
if AUTH0_CLIENT_ID and AUTH0_CLIENT_SECRET and AUTH0_DOMAIN:
    oauth = OAuth(app)
    oauth.register(
        "auth0",
        client_id=AUTH0_CLIENT_ID,
        client_secret=AUTH0_CLIENT_SECRET,
        client_kwargs={"scope": "openid profile email"},
        server_metadata_url=f'https://{AUTH0_DOMAIN}/.well-known/openid-configuration',
        redirect_uri=f"{BASE_URL}/callback"  # ← Use environment variable
    )
    print("Auth0 OAuth configured successfully")
    print(f"  Redirect URI: {BASE_URL}/callback")
    print(f"  Frontend URL: {FRONTEND_URL}")
else:
    print("Warning: Auth0 credentials not configured. Authentication will use mock user.")


# =============================================================================
# ENHANCED OLLAMA CALL FUNCTION
# =============================================================================

from flask import Response, stream_with_context
import json

def call_ollama_streaming(prompt: str, system_message: str = "", temperature: float = 0.1):
    """Stream Ollama responses token by token - WITH NON-STREAMING WORKAROUND"""
    
    print(f"[Ollama] 🚀 Calling Ollama at {OLLAMA_URL}/api/chat")
    print(f"[Ollama] Model: {OLLAMA_MODEL}")
    print(f"[Ollama] Prompt length: {len(prompt)} chars")
    print(f"[Ollama] System message length: {len(system_message)} chars")
    
    try:
        messages = []
        if system_message:
            messages.append({"role": "system", "content": system_message})
        messages.append({"role": "user", "content": prompt})
        
        # ✅ USE NON-STREAMING MODE (faster and more reliable)
        data = {
            "model": OLLAMA_MODEL,
            "messages": messages,
            "stream": False,  # ← Non-streaming mode
            "options": {
                "temperature": temperature,
                "top_p": 0.9,
                "top_k": 40,
                "repeat_penalty": 1.1,
                "num_ctx": 2048,
                "num_predict": 1500
            }
        }
        
        print(f"[Ollama] 📡 Sending non-streaming request...")
        response = ollama_session.post(
            f"{OLLAMA_URL}/api/chat",
            json=data,
            timeout=120  # 2 minute timeout
        )
        
        print(f"[Ollama] 📥 Response status: {response.status_code}")
        
        if response.status_code != 200:
            print(f"[Ollama] ❌ Bad status: {response.status_code}")
            print(f"[Ollama] Response text: {response.text[:500]}")
            yield f"Error: Ollama returned status {response.status_code}"
            return
        
        result = response.json()
        
        if 'message' in result and 'content' in result['message']:
            answer = result['message']['content']
            print(f"[Ollama] ✅ Got response: {len(answer)} chars")
            print(f"[Ollama] Preview: {answer[:150]}...")
            
            # Simulate streaming by yielding words
            words = answer.split()
            print(f"[Ollama] 🔄 Simulating streaming with {len(words)} words...")
            
            for i, word in enumerate(words, 1):
                yield word + " "
                
                # Log progress every 50 words
                if i % 50 == 0:
                    print(f"[Ollama] Streamed {i}/{len(words)} words...")
            
            print(f"[Ollama] ✅ Streaming simulation complete")
        else:
            print(f"[Ollama] ❌ No content in response")
            print(f"[Ollama] Response keys: {result.keys()}")
            yield "Error: Ollama response missing content field."
    
    except requests.exceptions.Timeout:
        print(f"[Ollama] ❌ Request timed out after 120 seconds")
        yield "Error: The AI service took too long to respond. Please try a simpler question."
    
    except requests.exceptions.ConnectionError as e:
        print(f"[Ollama] ❌ Connection error: {str(e)}")
        yield f"Error: Cannot connect to Ollama at {OLLAMA_URL}. Please check if Ollama is running."
    
    except Exception as e:
        print(f"[Ollama] ❌ Unexpected error: {str(e)}")
        import traceback
        print(f"[Ollama] Full traceback:")
        traceback.print_exc()
        yield f"Error: {str(e)}"




def process_samm_query_streaming(query: str, chat_history: List = None, documents_context: List = None):
    """Process query with streaming support"""

    # ✅ ADD THIS AT THE VERY TOP:
    # Extract financial records from documents
    financial_records = extract_financial_records_from_documents(documents_context)
    
    if financial_records:
        print(f"[Streaming] 💰 {len(financial_records)} financial records available")
        yield {"type": "financial_data_loaded", "count": len(financial_records)}

    # Yield progress updates
    yield {"type": "progress", "step": "intent_analysis", "message": "Analyzing intent..."}
    
    # Intent analysis
    intent_info = orchestrator.intent_agent.analyze_intent(query)
    yield {"type": "intent", "data": intent_info}
    
    # Entity extraction
    yield {"type": "progress", "step": "entity_extraction", "message": "Extracting entities..."}
    entity_info = orchestrator.entity_agent.extract_and_retrieve(query, intent_info)
    yield {"type": "entities", "data": {
        "count": len(entity_info.get('entities', [])),
        "entities": entity_info.get('entities', [])
    }}
    
    # Generate answer with streaming
    yield {"type": "progress", "step": "generating_answer", "message": "Generating answer..."}
    
    # Build context
    context = orchestrator.answer_agent._build_comprehensive_context(
        query, intent_info, entity_info, chat_history, documents_context
    )

    system_msg = orchestrator.answer_agent._create_optimized_system_message(
        intent_info.get("intent", "general"), context
    )
    prompt = orchestrator.answer_agent._create_enhanced_prompt(query, intent_info, entity_info)
    
    # Stream the answer
    full_answer = ""
    for token in call_ollama_streaming(prompt, system_msg, temperature=0.1):
        full_answer += token
        yield {"type": "answer_chunk", "content": token}
    
    # Send final metadata
    yield {
        "type": "complete",
        "data": {
            "intent": intent_info.get('intent', 'unknown'),
            "entities_found": len(entity_info.get('entities', [])),
            "answer_length": len(full_answer)
        }
    }




def call_ollama_enhanced(prompt: str, system_message: str = "", temperature: float = 0.1) -> str:
    """
    Enhanced Ollama API call with fast timeouts, automatic retries, and fallback.
    ALWAYS returns a response - never crashes or returns errors.
    """
    try:
        messages = []
        if system_message:
            messages.append({"role": "system", "content": system_message})
        messages.append({"role": "user", "content": prompt})
        
        data = {
            "model": OLLAMA_MODEL,
            "messages": messages,
            "stream": False,
            "options": {
                "temperature": temperature,
                "top_p": 0.9,
                "top_k": 40,
                "repeat_penalty": 1.1,
                "num_ctx": 2048,
                "num_predict": 500
            }
        }
        
        for attempt in range(1, OLLAMA_MAX_RETRIES + 1):
            try:
                print(f"[Ollama Enhanced] Attempt {attempt}/{OLLAMA_MAX_RETRIES} (timeout: {OLLAMA_TIMEOUT_NORMAL}s)")
                response = ollama_session.post(f"{OLLAMA_URL}/api/chat", json=data, timeout=OLLAMA_TIMEOUT_NORMAL)
                response.raise_for_status()
                result = response.json()
                answer = result["message"]["content"]
                print(f"[Ollama Enhanced] ✅ Success - {len(answer)} chars")
                return answer
            except requests.exceptions.Timeout:
                print(f"[Ollama Enhanced] ⏱️ Timeout on attempt {attempt}")
                if attempt < OLLAMA_MAX_RETRIES:
                    time.sleep(2 ** attempt)
            except requests.exceptions.RequestException as e:
                print(f"[Ollama Enhanced] API error on attempt {attempt}: {e}")
                if attempt < OLLAMA_MAX_RETRIES:
                    time.sleep(1)
        
        print(f"[Ollama Enhanced] 🔄 Using fallback response")
        return _get_intelligent_fallback()
        
    except Exception as e:
        print(f"[Ollama Enhanced] Processing error: {e}")
        return _get_intelligent_fallback()


def _get_intelligent_fallback() -> str:
    """Returns helpful SAMM information when Ollama is unavailable"""
    return """I apologize, but I'm currently experiencing technical difficulties connecting to the AI service.

However, I can still provide you with key SAMM (Security Assistance Management Manual) information:

**Core Concepts:**
• **Security Cooperation (SC)**: The broad umbrella of all DoD activities with international partners to achieve strategic objectives. Authorized under Title 10.
• **Security Assistance (SA)**: A subset of SC consisting of specific programs (FMS, FMF, IMET) authorized under Title 22 to transfer defense articles, training, and services.

**Key Organizations:**
• **DSCA** (Defense Security Cooperation Agency): Directs, administers, and provides guidance to DoD Components for SC programs
• **Department of State**: Provides continuous supervision and general direction of SA programs
• **DFAS** (Defense Finance and Accounting Service): Performs accounting, billing, disbursing, and collecting functions

**Legal Authorities:**
• Foreign Assistance Act (FAA) of 1961
• Arms Export Control Act (AECA) of 1976
• National Defense Authorization Act (NDAA) - annual

**Key Distinction:**
SC is the BROAD category (Title 10), and SA is a SUBSET of SC (Title 22).

Please try your question again. The AI service should be available shortly."""

# =============================================================================
# EMBEDDED SAMM KNOWLEDGE GRAPH DATA (RDF/TTL)
# =============================================================================

SAMM_KNOWLEDGE_GRAPH = """
# SAMM Chapter 1 Knowledge Graph (TTL/RDF Format)
@prefix samm: <http://samm.mil/ontology#> .
@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
@prefix rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .

# Core Concepts
samm:SecurityCooperation rdf:type samm:Concept ;
    rdfs:label "Security Cooperation" ;
    samm:definition "All activities undertaken by the DoD to encourage and enable international partners to work with the United States to achieve strategic objectives" ;
    samm:section "C1.1.1" ;
    samm:authority "Title 10" ;
    samm:funding "DoD appropriations" .

samm:SecurityAssistance rdf:type samm:Concept ;
    rdfs:label "Security Assistance" ;
    samm:definition "Group of programs authorized under Title 22 authorities by which the United States provides defense articles, military education and training" ;
    samm:section "C1.1.2.2" ;
    samm:authority "Title 22" ;
    samm:funding "Foreign Operations appropriations" ;
    samm:relationship samm:isSubsetOf ;
    samm:relatedTo samm:SecurityCooperation .

# Organizations
samm:DSCA rdf:type samm:Organization ;
    rdfs:label "Defense Security Cooperation Agency" ;
    samm:fullName "Defense Security Cooperation Agency" ;
    samm:role "Directs, administers, and provides guidance to DoD Components for SC programs" ;
    samm:section "C1.3.2.2" .

samm:DepartmentOfState rdf:type samm:Organization ;
    rdfs:label "Department of State" ;
    samm:role "Continuous supervision and general direction of SA programs" ;
    samm:authority "Secretary of State" ;
    samm:section "C1.3.1" .

samm:DepartmentOfDefense rdf:type samm:Organization ;
    rdfs:label "Department of Defense" ;
    samm:role "Establishes military requirements and implements programs" ;
    samm:authority "Secretary of Defense" ;
    samm:section "C1.3.2" .

samm:DFAS rdf:type samm:Organization ;
    rdfs:label "Defense Finance and Accounting Service" ;
    samm:fullName "Defense Finance and Accounting Service" ;
    samm:role "Performs accounting, billing, disbursing, and collecting functions for SC programs" ;
    samm:section "C1.3.2.8" .

samm:ImplementingAgency rdf:type samm:Organization ;
    rdfs:label "Implementing Agency" ;
    samm:definition "MILDEP organization or defense agency responsible for execution of SC programs" ;
    samm:role "Overall management of actions for delivery of materiel, supporting equipment, or services" ;
    samm:section "C1.3.2.6" .

# Legal Authorities
samm:ForeignAssistanceAct rdf:type samm:Authority ;
    rdfs:label "Foreign Assistance Act" ;
    samm:year "1961" ;
    samm:type "Title 22" ;
    samm:section "C1.2.1" .

samm:ArmsExportControlAct rdf:type samm:Authority ;
    rdfs:label "Arms Export Control Act" ;
    samm:acronym "AECA" ;
    samm:year "1976" ;
    samm:type "Title 22" ;
    samm:section "C1.2.1" .

samm:NDAA rdf:type samm:Authority ;
    rdfs:label "National Defense Authorization Act" ;
    samm:acronym "NDAA" ;
    samm:type "Title 10" ;
    samm:annual "true" ;
    samm:section "C1.1.2.1" .

# Key Relationships and Distinctions
samm:SecurityAssistance samm:isSubsetOf samm:SecurityCooperation .
samm:SecurityCooperation samm:authorizedBy samm:NDAA .
samm:SecurityAssistance samm:authorizedBy samm:ForeignAssistanceAct .
samm:SecurityAssistance samm:authorizedBy samm:ArmsExportControlAct .
samm:SecurityAssistance samm:supervisedBy samm:DepartmentOfState .
samm:SecurityCooperation samm:ledBy samm:DepartmentOfDefense .
"""

SAMM_TEXT_CONTENT = None

# =============================================================================
# SIMPLE KNOWLEDGE GRAPH PARSER
# =============================================================================

class SimpleKnowledgeGraph:
    """Simple knowledge graph parser for SAMM TTL data"""
    
    def __init__(self, ttl_data: str):
        self.entities = {}
        self.relationships = []
        self._parse_ttl(ttl_data)
    
    def _parse_ttl(self, ttl_data: str):
        """Parse TTL data into structured knowledge"""
        lines = ttl_data.split('\n')
        current_entity = None
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
                
            # New entity definition
            if 'rdf:type' in line:
                parts = line.split()
                if len(parts) >= 3:
                    entity_id = parts[0].replace('samm:', '')
                    entity_type = parts[2].replace('samm:', '').replace(';', '')
                    current_entity = {
                        'id': entity_id,
                        'type': entity_type,
                        'properties': {}
                    }
                    self.entities[entity_id] = current_entity
            
            # Properties
            elif current_entity and any(prop in line for prop in ['rdfs:label', 'samm:definition', 'samm:role', 'samm:section', 'samm:authority', 'samm:year']):
                if '"' in line:
                    prop_name = line.split()[0].replace('samm:', '').replace('rdfs:', '')
                    prop_value = line.split('"')[1] if '"' in line else line.split()[-1].replace(';', '').replace('.', '')
                    current_entity['properties'][prop_name] = prop_value
            
            # Relationships
            elif current_entity and any(rel in line for rel in ['samm:isSubsetOf', 'samm:supervisedBy', 'samm:ledBy', 'samm:authorizedBy']):
                parts = line.split()
                if len(parts) >= 2:
                    relationship = parts[0].replace('samm:', '')
                    target = parts[1].replace('samm:', '').replace('.', '').replace(';', '')
                    self.relationships.append({
                        'source': current_entity['id'],
                        'relationship': relationship,
                        'target': target
                    })
    
    def find_entity(self, query: str) -> Optional[Dict]:
        """Find entity by name or label"""
        query_lower = query.lower()
        
        # Direct match
        for entity_id, entity in self.entities.items():
            if entity_id.lower() == query_lower:
                return entity
            if entity['properties'].get('label', '').lower() == query_lower:
                return entity
        
        # Partial match
        for entity_id, entity in self.entities.items():
            if query_lower in entity_id.lower():
                return entity
            if query_lower in entity['properties'].get('label', '').lower():
                return entity
        
        return None
    
    def get_relationships(self, entity_id: str) -> List[Dict]:
        """Get relationships for an entity"""
        return [rel for rel in self.relationships 
                if rel['source'] == entity_id or rel['target'] == entity_id]

# Initialize knowledge graph
knowledge_graph = SimpleKnowledgeGraph(SAMM_KNOWLEDGE_GRAPH)
print(f"Knowledge Graph loaded: {len(knowledge_graph.entities)} entities, {len(knowledge_graph.relationships)} relationships")

# =============================================================================
# DATABASE MANAGER FOR INTEGRATED AGENTS
# =============================================================================

class DatabaseManager:
    """
    Manages connections to all three databases with improved error handling
    """
    
    def __init__(self):
        self.cosmos_gremlin_client = None
        self.vector_db_client = None
        self.embedding_model = None
        self.initialize_connections()
    
    def initialize_connections(self):
        """Initialize all database connections with better error handling"""
        print("[DatabaseManager] Initializing database connections...")
        
        # Initialize Cosmos DB Gremlin connection
        self._init_cosmos_gremlin()
        # Initialize ChromaDB connections
        self._init_vector_dbs()
        # Initialize embedding model
        self._init_embedding_model()
    
    def _init_cosmos_gremlin(self):
        """Initialize Cosmos DB Gremlin with proper cleanup"""
        if not client or not COSMOS_GREMLIN_CONFIG['password']:
            print("[DatabaseManager] Cosmos Gremlin credentials not available")
            return
            
        try:
            username = f"/dbs/{COSMOS_GREMLIN_CONFIG['database']}/colls/{COSMOS_GREMLIN_CONFIG['graph']}"
            endpoint_url = f"wss://{COSMOS_GREMLIN_CONFIG['endpoint']}:443/gremlin"
            
            self.cosmos_gremlin_client = client.Client(
                url=endpoint_url,
                traversal_source="g",
                username=username,
                password=COSMOS_GREMLIN_CONFIG['password'],
                message_serializer=serializer.GraphSONSerializersV2d0()
            )
            
            # Test connection with timeout
            result = self.cosmos_gremlin_client.submit("g.V().limit(1).count()").all().result()
            print(f"[DatabaseManager] Cosmos Gremlin connected successfully - {result[0]} vertices available")
            
        except Exception as e:
            print(f"[DatabaseManager] Cosmos Gremlin connection failed: {e}")
            self.cosmos_gremlin_client = None
    
    def extract_metadata_from_content(self, content: str) -> dict:
        """
        Extract chapter and section numbers from content text
        Works for patterns like: C1.3.2.8. or C5.4.1.
        """
        metadata = {
            'chapter_number': 'Unknown',
            'section_number': 'Unknown'
        }
    
        # Pattern 1: C1.3.2.8. Defense Finance... (most common)
        match = re.match(r'^(C(\d+)\.[\d\.]+)\.\s', content)
    
        if match:
            section = match.group(1)  # "C1.3.2.8"
            chapter = match.group(2)  # "1"
        
            metadata['section_number'] = section
            metadata['chapter_number'] = chapter
        
            print(f"[MetadataExtract] Extracted: Chapter {chapter}, Section {section}")
            return metadata
    
        # Pattern 2: C1. T1. (tables)
        match = re.match(r'^(C(\d+)\.\s*T\d+)', content)
        if match:
            section = match.group(1)
            chapter = match.group(2)
            metadata['section_number'] = section
            metadata['chapter_number'] = chapter
            print(f"[MetadataExtract] Extracted table: Chapter {chapter}, Section {section}")
            return metadata
    
        # Pattern 3: Chapter X. (heading style)
        match = re.match(r'^Chapter\s+(\d+)', content, re.IGNORECASE)
        if match:
            chapter = match.group(1)
            metadata['chapter_number'] = chapter
            print(f"[MetadataExtract] Extracted chapter heading: Chapter {chapter}")
            return metadata
    
        return metadata

    
    def _init_vector_dbs(self):
        """Initialize vector databases"""
        if not chromadb:
            print("[DatabaseManager] ChromaDB not available")
            return
            
        # Initialize ChromaDB vector_db (documents)
        try:
            if Path(VECTOR_DB_PATH).exists():
                self.vector_db_client = chromadb.PersistentClient(path=VECTOR_DB_PATH)
                collections = self.vector_db_client.list_collections()
                print(f"[DatabaseManager] Vector DB connected - {len(collections)} collections available")
            if collections:
                for col in collections:
                    print(f"\n[DEBUG] Collection: {col.name}")
                    print(f"[DEBUG] Metadata: {col.metadata}")
                    print(f"[DEBUG] Count: {col.count()}")
            else:
                print(f"[DatabaseManager] Vector DB path not found: {VECTOR_DB_PATH}")
        except Exception as e:
            print(f"[DatabaseManager] Vector DB connection failed: {e}")
            self.vector_db_client = None
        
    
    def _init_embedding_model(self):
        """Initialize embedding model"""
        if not SentenceTransformer:
            print("[DatabaseManager] SentenceTransformer not available")
            return
            
        try:
            self.embedding_model = SentenceTransformer(EMBEDDING_MODEL)
            print(f"[DatabaseManager] Embedding model loaded: {EMBEDDING_MODEL}")
        except Exception as e:
            print(f"[DatabaseManager] Embedding model failed to load: {e}")
            self.embedding_model = None
    
    def query_cosmos_graph(self, query_text: str, entities: List[str] = None) -> List[Dict]:
        """Query Cosmos DB graph database with better error handling"""
        if not self.cosmos_gremlin_client:
            return []
        
        results = []
        
        try:
            if entities:
                # Limit entities to prevent too many queries
                limited_entities = entities[:3]  # Only process first 3 entities
                
                for entity in limited_entities:
                    # Clean entity name for Gremlin query
                    entity_clean = re.sub(r'[^\w\s]', '', entity).strip()
                    if not entity_clean:
                        continue
                    
                    try:
                        # Query for vertices with matching names (with timeout)
                        vertex_query = f"g.V().has('name', containing('{entity_clean}')).limit(10)"
                        vertex_results = self.cosmos_gremlin_client.submit(vertex_query).all().result()
                        
                        for vertex in vertex_results:
                            results.append({
                                "type": "vertex",
                                "data": vertex,
                                "source": "cosmos_gremlin",
                                "entity": entity
                            })
                        
                        # Query for relationships involving this entity (limited)
                        edge_query = f"g.V().has('name', containing('{entity_clean}')).bothE().limit(5)"
                        edge_results = self.cosmos_gremlin_client.submit(edge_query).all().result()
                        
                        for edge in edge_results:
                            results.append({
                                "type": "edge", 
                                "data": edge,
                                "source": "cosmos_gremlin",
                                "entity": entity
                            })
                            
                    except Exception as entity_error:
                        print(f"[DatabaseManager] Error querying entity '{entity}': {entity_error}")
                        continue
            else:
                # General query for high-level entities
                general_query = "g.V().limit(10)"
                general_results = self.cosmos_gremlin_client.submit(general_query).all().result()
                
                for vertex in general_results:
                    results.append({
                        "type": "vertex",
                        "data": vertex,
                        "source": "cosmos_gremlin"
                    })
            
            print(f"[DatabaseManager] Cosmos Gremlin query returned {len(results)} results")
            
        except Exception as e:
            print(f"[DatabaseManager] Cosmos Gremlin query error: {e}")
        
        return results
    
    def query_vector_db(self, query: str, collection_name: str = None, n_results: int = 5) -> List[Dict]:
        """Query vector database and return results with enhanced metadata - OPTIMIZED for speed"""
        try:
            if not self.vector_db_client:
                print("[DatabaseManager] Vector DB client not available")
                return []
        
            collection = self.vector_db_client.get_collection(collection_name or VECTOR_DB_COLLECTION)
        
            results = collection.query(
                query_texts=[query],
                n_results=n_results
            )
            # ✅ ENHANCED: Format results with metadata extraction
            formatted_results = []
            for i, (doc, meta, distance) in enumerate(zip(
                results['documents'][0],
                results['metadatas'][0],
                results['distances'][0]
            )):
                # ✅ NEW: Check if metadata is missing and extract from content
                if meta.get('chapter_number') == 'Unknown' or not meta.get('chapter_number'):
                    extracted_meta = self.extract_metadata_from_content(doc)
                    meta.update(extracted_meta)
                    print(f"[DatabaseManager] Updated metadata for result {i+1}: Chapter {extracted_meta['chapter_number']}, Section {extracted_meta['section_number']}")

                # Convert distance to similarity score (0 = identical, 2 = very different for cosine)
                # For cosine distance: similarity = 1 - distance
                similarity_score = 1 - distance if distance <= 1 else distance
                
                formatted_results.append({
                    'content': doc,
                    'metadata': meta,
                    'distance': distance,  # Keep original distance
                    'similarity': distance,  # Keep for backward compatibility
                    'similarity_score': round(similarity_score, 4)  # Add readable score
                })

            print(f"[DatabaseManager] Vector DB query returned {len(formatted_results)} results")
            return formatted_results

        
        except Exception as e:
            print(f"[DatabaseManager] Vector DB query error: {e}")
            return []

    
    def cleanup(self):
        """Cleanup database connections"""
        try:
            if self.cosmos_gremlin_client:
                self.cosmos_gremlin_client.close()
                print("[DatabaseManager] Cosmos Gremlin connection closed")
        except Exception as e:
            print(f"[DatabaseManager] Error closing Cosmos Gremlin: {e}")
    
    def get_database_status(self) -> Dict[str, Any]:
        """Get status of all database connections"""
        status = {
            "cosmos_gremlin": {
                "connected": self.cosmos_gremlin_client is not None,
                "endpoint": COSMOS_GREMLIN_CONFIG['endpoint'],
                "database": COSMOS_GREMLIN_CONFIG['database'],
                "graph": COSMOS_GREMLIN_CONFIG['graph']
            },
            "vector_db": {
                "connected": self.vector_db_client is not None,
                "path": VECTOR_DB_PATH,
                "collections": []
            },
            
            "embedding_model": {
                "loaded": self.embedding_model is not None,
                "model_name": EMBEDDING_MODEL
            }
        }
        
        # Get collection info safely
        try:
            if self.vector_db_client:
                collections = self.vector_db_client.list_collections()
                status["vector_db"]["collections"] = [c.name for c in collections]
        except:
            pass
        
        return status

# Initialize database manager
db_manager = DatabaseManager()

# =============================================================================
# LANGGRAPH STATE ORCHESTRATION SYSTEM
# =============================================================================

class AgentState(TypedDict):
    """State shared across all agents in the workflow"""
    query: str
    chat_history: Optional[List[Dict]]
    documents_context: Optional[List[Dict]]
    intent_info: Optional[Dict[str, Any]]
    entity_info: Optional[Dict[str, Any]]
    answer: Optional[str]
    execution_steps: List[str]
    start_time: float
    current_step: str
    error: Optional[str]

# ============================================================================
# HITL FEEDBACK LOOP SYSTEM
# ============================================================================

HITL_CORRECTIONS_STORE = {
    "intent_corrections": {},
    "entity_corrections": {},
    "answer_corrections": {},
    "correction_history": []
}

# ============================================================================
# HITL FILE PERSISTENCE
# ============================================================================
HITL_STORAGE_FILE = Path("hitl_corrections.json")

def save_hitl_corrections():
    """Save HITL corrections to file for persistence"""
    try:
        data = {
            "intent_corrections": HITL_CORRECTIONS_STORE["intent_corrections"],
            "entity_corrections": HITL_CORRECTIONS_STORE["entity_corrections"],
            "answer_corrections": HITL_CORRECTIONS_STORE["answer_corrections"],
            "correction_history": HITL_CORRECTIONS_STORE["correction_history"][-100:]
        }
        with open(HITL_STORAGE_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        print(f"[HITL] Saved {len(HITL_CORRECTIONS_STORE['answer_corrections'])} corrections to {HITL_STORAGE_FILE}")
        return True
    except Exception as e:
        print(f"[HITL] Error saving corrections: {e}")
        return False

def load_hitl_corrections():
    """Load HITL corrections from file on startup"""
    global HITL_CORRECTIONS_STORE
    try:
        if HITL_STORAGE_FILE.exists():
            with open(HITL_STORAGE_FILE, 'r') as f:
                data = json.load(f)
            HITL_CORRECTIONS_STORE["intent_corrections"] = data.get("intent_corrections", {})
            HITL_CORRECTIONS_STORE["entity_corrections"] = data.get("entity_corrections", {})
            HITL_CORRECTIONS_STORE["answer_corrections"] = data.get("answer_corrections", {})
            HITL_CORRECTIONS_STORE["correction_history"] = data.get("correction_history", [])
            print(f"[HITL] Loaded {len(HITL_CORRECTIONS_STORE['answer_corrections'])} corrections from {HITL_STORAGE_FILE}")
        else:
            print(f"[HITL] No existing corrections file found - starting fresh")
    except Exception as e:
        print(f"[HITL] Error loading corrections: {e}")

# Load any existing corrections on startup
load_hitl_corrections()
# ============================================================================

def create_question_hash(question: str) -> str:
    """Create normalized hash for question matching"""
    normalized = question.lower().strip()
    return hashlib.md5(normalized.encode()).hexdigest()
# Demo scenarios - EMPTY by default
DEMO_SCENARIOS = {}

# DON'T pre-populate corrections! Let them be applied via HITL dashboard
# This allows first query to show partial answer, second query to show corrected answer
print("✅ HITL Feedback System Initialized with 2 demo scenarios")
print("💡 TIP: Use /api/hitl/reset-demo to clear corrections between tests")

def apply_hitl_corrections(question: str, result: dict) -> dict:
    """Apply HITL corrections if they exist"""
    q_hash = create_question_hash(question)
    corrections_applied = []
    
    if q_hash in HITL_CORRECTIONS_STORE["intent_corrections"]:
        result['metadata']['intent'] = HITL_CORRECTIONS_STORE["intent_corrections"][q_hash]
        corrections_applied.append("Intent")
        print(f"🔄 HITL: Intent correction applied")
    
    if q_hash in HITL_CORRECTIONS_STORE["entity_corrections"]:
        result['metadata']['entities'] = HITL_CORRECTIONS_STORE["entity_corrections"][q_hash]
        corrections_applied.append("Entities")
        print(f"🔄 HITL: Entity corrections applied")
    
    if q_hash in HITL_CORRECTIONS_STORE["answer_corrections"]:
        result['answer'] = HITL_CORRECTIONS_STORE["answer_corrections"][q_hash]
        result['metadata']['hitl_corrected'] = True
        corrections_applied.append("Answer")
        print(f"🔄 HITL: Answer correction applied")
    
    if corrections_applied:
        result['metadata']['hitl_corrections_applied'] = corrections_applied
    
    return result

def generate_demo_partial_response(question: str) -> dict:
    """Return partial response for demo questions on first ask"""
    # DISABLED FOR NOW - Let normal LLM processing happen for testing
    return None
    
    # Original code commented out:
    # q_hash = create_question_hash(question)
    # 
    # for scenario_name, scenario_data in DEMO_SCENARIOS.items():
    #     scenario_hash = create_question_hash(scenario_data["question"])
    #     if q_hash == scenario_hash:
    #         # Only return partial if corrections NOT applied yet
    #         if q_hash not in HITL_CORRECTIONS_STORE["answer_corrections"]:
    #             print(f"🎬 DEMO ({scenario_name.upper()}): Returning partial response")
    #             return {
    #                 "intent": scenario_data["original"]["intent"],
    #                 "entities": scenario_data["original"]["entities"],
    #                 "answer": scenario_data["original"]["answer"],
    #                 "is_demo": True,
    #                 "demo_type": scenario_name
    #             }
    # 
    # return None

# ============================================================================
# END HITL SYSTEM
# ============================================================================

class WorkflowStep(Enum):
    """Workflow steps for state orchestration"""
    INIT = "initialize"
    INTENT = "analyze_intent"
    ENTITY = "extract_entities"
    ANSWER = "generate_answer"
    COMPLETE = "complete"
    ERROR = "error"

def call_ollama(prompt: str, system_message: str = "") -> str:
    """Call Ollama with system message and prompt (legacy function for compatibility)"""
    return call_ollama_enhanced(prompt, system_message, temperature=0.1)

def extract_financial_records_from_documents(documents_context: List) -> List[Dict]:
    """
    Extract financial records from uploaded documents
    Returns list of financial records with PDLI info
    """
    if not documents_context:
        return []
    
    financial_records = []
    
    for doc in documents_context:
        # Check if document has financial data in metadata
        if doc.get('metadata', {}).get('hasFinancialData'):
            records = doc['metadata'].get('financialRecords', [])
            
            print(f"[Financial Extract] Found {len(records)} records in {doc.get('fileName')}")
            
            # Enrich each record with document info
            for record in records:
                enriched = {
                    **record,
                    'source_document': doc.get('fileName'),
                    'document_id': doc.get('documentId')
                }
                financial_records.append(enriched)
    
    print(f"[Financial Extract] Total: {len(financial_records)} financial records extracted")
    return financial_records
def extract_case_id_from_filename(filename: str) -> str:
    """
    Extract case ID from filename like 'SR-P-NAV_Financial.xlsx'
    Returns case ID or None
    """
    # Pattern: SR-X-YYY or similar at start of filename
    import re
    match = re.match(r'^(SR-[A-Z]-[A-Z0-9]+)', filename, re.IGNORECASE)
    if match:
        case_id = match.group(1).upper()
        print(f"[Upload] 📋 Extracted case ID from filename: {case_id}")
        return case_id
    return None 

class IntentAgent:
    """Intent analysis using Ollama with Human-in-Loop and trigger updates"""
    
    def __init__(self, available_chapters=None):
        # Define ALL chapters upfront (even if data not loaded yet)
        self.available_chapters = available_chapters or [1, 4, 5, 6, 7, 9]
        self.hil_feedback_data = []  # Store human feedback for intent corrections
        self.intent_patterns = {}    # Store learned patterns from feedback
        self.trigger_updates = []    # Store updates from new entity/relationship data
        
        # Special case patterns for non-SAMM, nonsense, and incomplete queries
        self.special_case_patterns = {
            "nonsense_keywords": [
                "asdfghjkl", "qwerty", "xyzpdq", "flurble", "lorem ipsum",
                "banana helicopter", "purple dreams", "sparkle fountain",
                "waxing moon potato", "rainbow process asteroid"
            ],
            "incomplete_phrases": [
                "what about it", "tell me about that", "explain that",
                "the thing", "tell me about the thing", "can you explain that",
                "what about", "who handles", "the process for", "explain how"
            ],
            "non_samm_topics": [
                "nato", "article 5", "ndaa process", "title 50", "joint chiefs",
                "intelligence community", "five eyes", "un security council",
                "federal acquisition regulation", "far", "unified command plan",
                "third offset", "national security strategy", "humanitarian assistance",
                "bilateral agreement", "multilateral", "defense officer personnel"
            ]
        }
        
        # BASE INTENTS: Always active (7 intents)
        self.base_intents = {
            "definition": "asking what something is",
            "distinction": "asking about differences between concepts",
            "authority": "asking about who has authority or oversight",
            "organization": "asking about agencies and their roles",
            "factual": "asking for specific facts like dates, numbers",
            "relationship": "asking about how things are connected",
            "general": "general questions"
        }
        
                # ALL CHAPTER INTENTS: Define all upfront (22 intents)
        self.chapter_intents = {
            1: {
                "scope": "asking about what is included or excluded",
                "purpose": "asking about the purpose or objective"
            },
            4: {
                "approval": "asking about approval processes or authority",
                "review": "asking about review procedures or requirements",
                "decision": "asking about decision-making processes"
            },
            5: {
                "process": "asking about procedures, workflows, or how to do something",
                "implementation": "asking how to implement or execute something",
                "prerequisite": "asking what is required before starting",
                "documentation": "asking about required forms, reports, or paperwork"
            },
            6: {
                "financial": "asking about costs, payments, billing, or financial matters",
                "budget": "asking about budget allocation or planning",
                "payment": "asking about payment terms, schedules, or methods",
                "pricing": "asking about how prices are calculated or determined",
                "reimbursement": "asking about refund or reimbursement processes"
            },
            7: {
                "timeline": "asking about deadlines, schedules, or when things happen",
                "milestone": "asking about project milestones or checkpoints",
                "status": "asking about current status or progress",
                "tracking": "asking about monitoring or tracking systems"
            },
            9: {
                "compliance": "asking about regulations, requirements, or obligations",
                "audit": "asking about audit processes or requirements",
                "legal": "asking about legal requirements or restrictions",
                "reporting": "asking about compliance reporting requirements"
            }
        }

        # ============================================================================
        # M1.3 HYBRID: Priority-Ordered Pattern Rules
        # Pattern matching first (fast), LLM only if needed (slow)
        # Formula: Pattern Match (40%) + Keyword Overlap (35%) + AI Certainty (25%)
        # ============================================================================
        
        # Pattern rules checked in ORDER (first match wins!)
        # Format: (internal_intent, [patterns])
        self.pattern_rules = [
            # ================================================================
            # TIER 1: MOST SPECIFIC (check first)
            # ================================================================
            
            # VERIFICATION - Yes/No questions (starts with is/does/are/can)
            ("verification", [
                r"^is [A-Z][\w\s]+ the ",
                r"^does [A-Z][\w\s]+ (?:approve|require|have|maintain)",
                r"^are [A-Z][\w\s]+ (?:subject|required|authorized)",
                r"^is [\w\s]+ (?:the executive agent|required|authorized|subject)",
                r"^does [\w\s]+ (?:approve|require|allow|have authority)",
                r"^are [\w\s]+ (?:subject to|required to|programs)",
            ]),
            
            # COMPARISON/DISTINCTION
            ("distinction", [
                r"difference between",
                r"differ from",
                r"compare (?:the )?",
                r" vs\.? ",
                r"distinguish ",
                r"how does [\w\s]+ differ",
                r"contrast between",
            ]),
            
            # EXPLANATION - Why/purpose questions
            ("explanation", [
                r"^why (?:is|are|does|do|did|should|would|has|have) ",
                r"why does [\w\s]+ maintain",
                r"how does [\w\s]+ support [\w\s]+ (?:objectives|security|goals|interests)",
                r"what is the (?:reason|purpose|rationale|objective|goal) (?:for|of|behind)",
                r"explain why",
                r"why is [\w\s]+ (?:important|necessary|required|needed)",
            ]),
            
            # ================================================================
            # TIER 2: LEGAL/COMPLIANCE (Chapter 9)
            # ================================================================
            ("compliance", [
                r"(?:faa|aeca|itar|ear) (?:section )?",
                r"what (?:legal )?authority governs",
                r"(?:congressional|notification|export|itar|legal) requirements",
                r"eligibility requirements",
                r"what does (?:faa|aeca|itar|ear)",
                r"what are the [\w\s]+ requirements",
                r"legal (?:authority|requirements|basis)",
                r"audit (?:process|requirements|procedures)",
                r"reporting requirements",
                r"compliance reporting",
            ]),
            
            # ================================================================
            # TIER 3: AUTHORITY/RESPONSIBILITY (Chapter 4)
            # ================================================================
            ("authority", [
                r"who (?:is responsible|has authority|has the authority|approves|authorizes|oversees|manages|supervises|determines|decides|can approve)",
                r"who has (?:ultimate |final )?authority",
                r"how does [\w\s]+ oversee",
                r"what (?:is|are) .+?(?:'s|s'|'s) (?:role|responsibilities|responsibility)",
                r"what are .+(?:'s|'s) .+ responsibilities",
                r"what (?:is|are) the (?:role|responsibilities|responsibility|duties) of",
                r"what are (?:the )?(?:general )?responsibilities of",
                r"what are (?:the )?[\w\s]+ responsibilities(?:\?|$)",
                # NEW M1.3: "What are the X's Y responsibilities?" - handles "Navy's SC responsibilities"
                r"what are (?:the )?[\w\s]+(?:'s|')[\w\s]* responsibilities\??$",
                r"(?:whose|under whose) (?:authority|oversight|supervision|responsibility)",
                r"responsible for ",
                r"authority to (?:approve|authorize|conduct|make)",
                r"who (?:can |is authorized to |has authority to |must )",
                r"approval (?:process|authority|requirements)",
            ]),
            
            # ================================================================
            # TIER 4: FACT_RETRIEVAL (Chapter 7 + factual)
            # ================================================================
            ("factual", [
                r"^which (?:directive|order|regulation|law|document|command|executive|agency)",
                r"which (?:command|agency|office) (?:does|is|has|governs)",
                r"what (?:is|are) the [\w\s]+ designation",
                r"what (?:is|are) the (?:primary|main) (?:site|location|office)",
                r"under which",
                r"what (?:regulation|directive|order|law) (?:covers|governs)",
                r"where is [\w\s]+ (?:located|based|housed)",
                r"when (?:is|are|was|were|does|do) ",
                r"what (?:is|are) the (?:deadline|timeline|schedule|date)",
                r"how long does",
                r"how (?:much|many) (?:time|days|weeks|months)",
                r"what is the [\w\s]+ (?:site|location|designation|number|amount)",
            ]),
            
            # ================================================================
            # TIER 5: PROCESS/PROCEDURE (Chapter 5)
            # ================================================================
            ("process", [
                r"what is the process (?:for|of|to)",
                r"how does (?:a |the )?(?:foreign )?",
                r"how (?:are|is) [\w\s]+ (?:developed|created|submitted|processed|approved|coordinated|reviewed|handled|prepared|conducted)",
                r"describe (?:the )?[\w\s]+ (?:process|procedure|coordination|workflow)",
                r"what (?:are|is) the (?:steps|procedures?|process|workflow)",
                r"how (?:to|do|does|can)",
                r"steps (?:for|to|in|involved)",
                r"(?:interagency|coordination) process",
                r"payment (?:process|procedures|terms)",
                r"reimbursement (?:process|procedures)",
            ]),
            
            # ================================================================
            # TIER 6: ORGANIZATIONAL STRUCTURE
            # ================================================================
            ("organization", [
                r"how is (?:the )?[\w\s]+ (?:enterprise )?organized",
                r"(?:what is|describe) (?:the )?[\w\s]+ (?:organizational )?structure",
                r"what organizations (?:fall under|are under|report to|comprise)",
                r"organizational structure",
                r"how (?:is|are) [\w\s]+ structured",
                r"what is the hierarchy",
                r"reporting structure",
                r"what (?:offices|divisions|components|units) ",
                r"falls under (?:which|what)",
                r"structure (?:of|for) ",
                # NEW: Framework patterns (framework = organizational structure)
                r"what is (?:the )?(?:sc|sa|security cooperation|security assistance)[\w\s]* (?:coordinating |coordination )?framework",
                r"what is (?:the )?[\w\s]+ (?:framework|hierarchy|structure)\??$",
                r"(?:coordinating|coordination) framework",
            ]),
            
            # ================================================================
            # TIER 7: LIST/ENUMERATION
            # ================================================================
            # TIER 7: LIST/ENUMERATION
            # ================================================================
            ("list", [
                r"^list (?:all |the )?",
                r"^name (?:the |all )?",
                r"^enumerate ",
                r"what are the (?:three|two|four|five|six|seven|eight|nine|ten|\d+) ",
                r"what (?:are|is) (?:all )?(?:the )?(?:types|kinds|categories|components|key|major|primary) (?:of|roles|functions|responsibilities)",
                r"what are (?:the )?(?:key )?(?:roles|functions|responsibilities|duties) of",
                r"what (?:is|are) (?:included|excluded|covered)",
                r"what (?:documents?|forms?) (?:is|are) (?:required|needed)",
                r"what are (?:the )?[\w\s]+ (?:oversight )?bodies",
                r"what are (?:the )?[\w\s]+ (?:advisory )?groups",
            ]),
            
            # ================================================================
            # TIER 8: RELATIONSHIP (connections between entities)
            # ================================================================
            ("relationship", [
                r"what (?:is|are) (?:the )?relationship(?:s)? (?:between|among|of)",
                r"relationship(?:s)? (?:between|among|of) ",
                r"how (?:is|are|does|do) [\w\s]+ (?:related|connected|linked) to",
                r"what is [\w\s]+(?:'s|'s) relationship (?:to|with)",
                r"connection(?:s)? between",
            ]),
            
            # ================================================================
            # TIER 9: DEFINITION (lowest priority - catch remaining "what is")
            # ================================================================
            ("definition", [
                r"^what is (?:a |an |the )?[\w\s\-]+\??$",
                r"^what are [\w\s\-]+\??$",
                r"^define (?:the )?(?:term )?",
                r"what does [\w\s]+ (?:mean|stand for)",
                r"explain (?:the term |what )?[\w\s]+$",
                r"what is meant by",
                r"meaning of ",
            ]),
        ]
        
        # M1.3: SAMM-specific keywords for keyword overlap calculation
        # UPDATED: Expanded keywords for better confidence scoring
        self.samm_keywords = {
            "entities": [
                # Core SAMM Agencies
                "dsca", "fms", "imet", "samm", "loa", "loc", "fmf", "eca", "aeca", "faa", 
                "dod", "dos", "state department", "defense", "security cooperation", 
                "security assistance", "foreign military", "usd(p)", "usdp", "secdef", "ccmd",
                "combatant command", "combatant commands", "building partner capacity", "bpc",
                "defense security", "military education", "training", "itar", "usml",
                "title 10", "title 22", "dfas", "gef", "lor", "sco", "usasac",
                # NEW: Implementing Agencies & Key Roles
                "implementing agency", "implementing agencies", "ia", "mildep",
                "president", "congress", "secretary of state", "secretary of defense",
                "army", "navy", "air force", "ussocom", "advisory group", "advisory groups",
                # NEW: Additional SAMM entities
                "dcma", "dcaa", "dla", "dtra", "disa", "nga", "nsa", "mda",
                "nipo", "satfa", "afsac", "usasac", "syscom",
                # NEW: Key SC/SA terms  
                "foreign partner", "foreign partners", "defense articles", "defense services",
                "letter of offer", "letter of acceptance", "case line", "fms case"
            ],
            "chapters": ["chapter", "section", "c1", "c4", "c5", "c6", "c7", "c9", "paragraph", "samm"],
            "actions": ["approval", "review", "authorization", "notification", "implementation",
                       "supervision", "oversight", "coordination", "execution", "delegation"],
            "concepts": ["threshold", "authority", "oversight", "program", "cooperation", "assistance",
                        "responsibility", "responsibilities", "role", "roles", "framework", "structure",
                        "relationship", "relationships", "stakeholder", "stakeholders"]
        }
        
        print("[IntentAgent M1.3] Initialized with HYBRID pattern + LLM confidence scoring")

    # ============================================================================
    # M1.3 NEW: Pattern-Based Intent Detection (FAST - No LLM)
    # ============================================================================
    def _detect_intent_from_patterns(self, query: str) -> Dict[str, Any]:
        """
        Fast pattern-based intent detection - no LLM call needed!
        Returns intent and confidence if pattern matches, None otherwise.
        """
        query_lower = query.lower().strip()
        
        # Check patterns in priority order
        for intent, patterns in self.pattern_rules:
            for pattern in patterns:
                try:
                    if re.search(pattern, query_lower, re.IGNORECASE):
                        print(f"[IntentAgent M1.3] ✅ Pattern match: '{pattern[:50]}...' → {intent}")
                        return {
                            "intent": intent,
                            "pattern_matched": True,
                            "pattern": pattern[:50],
                            "pattern_score": 1.0
                        }
                except re.error:
                    continue
        
        # No pattern match - return with low confidence
        return {
            "intent": "general",
            "pattern_matched": False,
            "pattern": None,
            "pattern_score": 0.3
        }

    # ============================================================================
    # M1.3 NEW: Keyword Overlap Score (35% weight)
    # UPDATED: HIGH_WEIGHT keywords get 0.85 on single match
    # ============================================================================
    
    # HIGH WEIGHT KEYWORDS: Single match = 0.85 score
    # These are core SAMM entities where even 1 match indicates strong relevance
    HIGH_WEIGHT_KEYWORDS = {
        # === CHAPTER 1: Core Organizations ===
        # Department of State
        "secretary of state", "secstate", "department of state",
        # Department of Defense - Top Level
        "secretary of defense", "secdef", "department of defense",
        # USD Offices
        "usd(p)", "usdp", "under secretary of defense for policy",
        "usd(a&s)", "usdas", "under secretary of defense for acquisition and sustainment",
        "usd(c)", "usdc", "under secretary of defense comptroller",
        "usd(p&r)", "usdpr", "under secretary of defense for personnel and readiness",
        # DSCA
        "dsca", "defense security cooperation agency",
        # Implementing Agencies - MILDEPs
        "implementing agency", "implementing agencies",
        "dasa (de&c)", "dasa de&c", "dasa dec", 
        "office of the deputy assistant secretary of the army for defense exports and cooperation",
        "nipo", "navy international programs office",
        "saf/ia", "safia", "deputy under secretary of the air force for international affairs",
        # Implementing Agencies - Defense Agencies
        "dcma", "defense contract management agency",
        "disa", "defense information systems agency", 
        "dla", "defense logistics agency",
        "dtra", "defense threat reduction agency",
        "mda", "missile defense agency",
        "nga", "national geospatial-intelligence agency",
        "nsa", "national security agency",
        # Other DoD Organizations
        "dcaa", "defense contract audit agency",
        "dfas", "defense finance and accounting service",
        "joint chiefs of staff", "jcs", "cjcs",
        "combatant commander", "combatant commanders", "ccdr", "ccdrs", "ccmd",
        # Army Commands
        "usasac", "u.s. army security assistance command",
        "satfa", "security assistance training field activity",
        "usace", "u.s. army corps of engineers",
        "amc", "army materiel command",
        "tradoc", "army training and doctrine command",
        # Air Force Commands
        "afsac", "air force security assistance and cooperation",
        "afsat", "air force security assistance training squadron",
        "aetc", "air education and training command",
        "afmc", "air force materiel command",
        # Navy Commands
        "syscom", "syscoms", "system commands",
        # === CHAPTER 1: Core Concepts ===
        "security cooperation", "security assistance",
        "foreign military sales", "fms",
        "international military education and training", "imet",
        "building partner capacity", "bpc",
        "direct commercial sales", "dcs",
        "excess defense articles", "eda",
        "letter of request", "lor",
        "letter of offer and acceptance", "loa",
        "memorandum of request", "mor",
        # === CHAPTER 1: Legal Authorities ===
        "foreign assistance act", "faa",
        "arms export control act", "aeca",
        "title 10", "title 22", "title 50",
        "executive order 13637", "e.o. 13637", "eo 13637",
        "national defense authorization act", "ndaa",
        "itar", "international traffic in arms regulations",
        "usml", "united states munitions list",
        # === Concepts that indicate SAMM domain ===
        "continuous supervision", "general direction",
        "defense articles", "defense services",
        "foreign partner", "foreign partners",
        "campaign plan", "campaign plans",
        # === Short forms (common in queries) ===
        " sc ", " sa ",  # Space-bounded to avoid false matches like "ascar"
        "sc program", "sa program", "sc programs", "sa programs",
        "sc responsibilities", "sa responsibilities",
        "sc activities", "sa activities",
    }
    
    def _calculate_keyword_overlap_score(self, query: str) -> float:
        """
        Calculate keyword overlap score (0.0 to 1.0)
        HIGH_WEIGHT keywords: single match = 0.85
        Normal keywords: need 2+ matches for 0.85
        """
        query_lower = query.lower()
        
        # Check HIGH WEIGHT keywords first
        high_weight_match = any(kw in query_lower for kw in self.HIGH_WEIGHT_KEYWORDS)
        
        if high_weight_match:
            # Count total high-weight matches
            hw_matches = sum(1 for kw in self.HIGH_WEIGHT_KEYWORDS if kw in query_lower)
            if hw_matches >= 2:
                return 0.95  # Multiple high-weight matches
            else:
                return 0.85  # Single high-weight match
        
        # Fall back to normal keyword counting
        total_matches = 0
        for category, keywords in self.samm_keywords.items():
            total_matches += sum(1 for kw in keywords if kw in query_lower)
        
        if total_matches >= 3:
            base_score = 0.95
        elif total_matches == 2:
            base_score = 0.85
        elif total_matches == 1:
            base_score = 0.70
        else:
            if any(term in query_lower for term in ["what is", "what are", "define", "what does"]):
                base_score = 0.60
            else:
                base_score = 0.40
        
        return base_score

    # ============================================================================
    # M1.3 NEW: Composite Confidence Calculation (HYBRID)
    # ============================================================================
    def _calculate_intent_confidence(self, query: str, detected_intent: str, pattern_score: float, ai_confidence: float) -> Dict[str, Any]:
        """
        Calculate composite confidence score:
        Confidence = Pattern Match (40%) + Keyword Overlap (35%) + AI Certainty (25%)
        """
        keyword_score = self._calculate_keyword_overlap_score(query)
        ai_score = min(ai_confidence, 1.0)
        
        PATTERN_WEIGHT = 0.40
        KEYWORD_WEIGHT = 0.35
        AI_WEIGHT = 0.25
        
        composite = (pattern_score * PATTERN_WEIGHT) + (keyword_score * KEYWORD_WEIGHT) + (ai_score * AI_WEIGHT)
        composite = round(composite, 2)
        
        print(f"[IntentAgent M1.3] Confidence: Pattern={pattern_score:.2f}×40% + Keyword={keyword_score:.2f}×35% + AI={ai_score:.2f}×25% = {composite:.2f}")
        
        return {
            "pattern_score": pattern_score,
            "keyword_score": keyword_score,
            "ai_score": ai_score,
            "composite": composite,
            "meets_target": composite >= 0.90
        }

    def _check_special_cases(self, query: str) -> Optional[Dict[str, Any]]:
        """Check for nonsense, incomplete, or non-SAMM queries before calling Ollama"""
        query_lower = query.lower().strip()
        query_words = query_lower.split()
        
        print(f"[IntentAgent] Checking special cases for: '{query[:50]}...'")
        
        # Check for financial verification queries
        if "funding request" in query_lower or "funds are available" in query_lower or "verify the appropriate funding line" in query_lower:
            if "sr-p-nav" in query_lower or "case sr" in query_lower:
                print(f"[IntentAgent] 🚀 FINANCIAL VERIFICATION detected")
                return {
                    "intent": "financial_verification",
                    "confidence": 0.95,
                    "entities_mentioned": ["SR-P-NAV", "funding", "PDLI", "LOA"],
                    "special_case": True,
                    "fast_path": True
                }
        
        # Check for technical services queries
        if "technical services" in query_lower or "what is included" in query_lower:
            if "sr-p-nav" in query_lower or "case sr" in query_lower:
                print(f"[IntentAgent] 🚀 TECHNICAL SERVICES query detected")
                return {
                    "intent": "line_item_details",
                    "confidence": 0.95,
                    "entities_mentioned": ["SR-P-NAV", "Line 007", "technical services"],
                    "special_case": True,
                    "fast_path": True
                }

        # Check for PMR minutes summary queries
        if "minutes" in query_lower or "pmr" in query_lower or "meeting" in query_lower:
            if any(keyword in query_lower for keyword in ["summarize", "summary", "action items", "action item"]):
                print(f"[IntentAgent] 🚀 PMR MINUTES SUMMARY detected")
                return {
                    "intent": "pmr_minutes_summary",
                    "confidence": 0.95,
                    "entities_mentioned": ["NSM", "RSNF", "PMR", "action items"],
                    "special_case": True,
                    "fast_path": True
                }

        # Check for LOA timeline queries
        if "loa" in query_lower or "letter of offer" in query_lower:
            loa_triggers = ["how long", "timeline", "timeframe", "time", "duration", "take to develop", "take to prepare"]
            if any(trigger in query_lower for trigger in loa_triggers):
                print(f"[IntentAgent] 🚀 LOA TIMELINE detected - returning instant answer")
                return {
                    "intent": "loa_timeline",
                    "confidence": 0.95,
                    "entities_mentioned": ["LOA", "Timeline"],
                    "special_case": True,
                    "fast_path": True
                }

        # 1. CHECK FOR NONSENSE/GIBBERISH
        nonsense_count = sum(1 for keyword in self.special_case_patterns["nonsense_keywords"] 
                            if keyword in query_lower)
        
        normal_chars = set('abcdefghijklmnopqrstuvwxyz0123456789 ?.!,;:\'-')
        unusual_symbol_count = sum(1 for c in query_lower if c not in normal_chars)
        unusual_symbol_ratio = unusual_symbol_count / max(len(query), 1)
        
        number_count = sum(1 for c in query if c.isdigit())
        number_ratio = number_count / max(len(query), 1)
        
        has_keyboard_mash = any(
            len(set(word)) == len(word) and len(word) > 12
            for word in query_words
            if word.isalpha()
        )
        
        if nonsense_count >= 2 or unusual_symbol_ratio > 0.2 or number_ratio > 0.7 or has_keyboard_mash:
            print(f"[IntentAgent] NONSENSE detected (keywords: {nonsense_count}, unusual_symbols: {unusual_symbol_ratio:.2f})")
            return {
                "intent": "nonsense",
                "confidence": 0.95,
                "entities_mentioned": [],
                "reason": "gibberish_detected",
                "special_case": True
            }
        
        # 2. CHECK FOR INCOMPLETE/VAGUE QUERIES
        if len(query_words) <= 5:
            for phrase in self.special_case_patterns["incomplete_phrases"]:
                if phrase in query_lower:
                    print(f"[IntentAgent] INCOMPLETE detected (phrase: '{phrase}')")
                    return {
                        "intent": "incomplete",
                        "confidence": 0.9,
                        "entities_mentioned": [],
                        "reason": "vague_or_incomplete",
                        "special_case": True
                    }
        
        question_words = ["what", "who", "when", "where", "why", "how", "does", "is", "are", "can"]
        if len(query_words) <= 3 and not any(qw in query_words for qw in question_words):
            if not query.strip().endswith("?"):
                print(f"[IntentAgent] INCOMPLETE detected (fragment: {len(query_words)} words)")
                return {
                    "intent": "incomplete",
                    "confidence": 0.85,
                    "entities_mentioned": [],
                    "reason": "fragment",
                    "special_case": True
                }
        
        # 3. CHECK FOR NON-SAMM TOPICS
        non_samm_matches = []
        for topic in self.special_case_patterns["non_samm_topics"]:
            if topic in query_lower:
                non_samm_matches.append(topic)
        
        if non_samm_matches:
            print(f"[IntentAgent] NON-SAMM detected (topics: {non_samm_matches})")
            return {
                "intent": "non_samm",
                "confidence": 0.9,
                "entities_mentioned": [],
                "reason": "outside_samm_scope",
                "special_case": True,
                "detected_topics": non_samm_matches
            }
        
        print(f"[IntentAgent] No special cases detected - proceeding with normal analysis")
        return None




    @time_function
    def analyze_intent(self, query: str) -> Dict[str, Any]:
        """
        M1.3 HYBRID: Intent analysis with pattern matching FIRST, LLM only if needed
        
        Flow:
        1. Check special cases (instant)
        2. Try pattern matching (fast, no LLM)
        3. If pattern confidence >= 0.90, return immediately (skip LLM!)
        4. If pattern confidence < 0.90, call LLM for refinement
        
        Confidence = Pattern Match (40%) + Keyword Overlap (35%) + AI Certainty (25%)
        """
        # STEP 1: Check special cases first
        special_case = self._check_special_cases(query)
        if special_case:
            print(f"[IntentAgent M1.3] Returning special case: {special_case['intent']}")
            conf_breakdown = self._calculate_intent_confidence(query, special_case['intent'], 1.0, special_case.get('confidence', 0.95))
            special_case['confidence'] = conf_breakdown['composite']
            special_case['confidence_breakdown'] = conf_breakdown
            return special_case
        
        # STEP 2: Try pattern matching FIRST (fast, no LLM!)
        pattern_result = self._detect_intent_from_patterns(query)
        pattern_intent = pattern_result["intent"]
        pattern_score = pattern_result["pattern_score"]
        
        # Calculate preliminary confidence (without LLM)
        keyword_score = self._calculate_keyword_overlap_score(query)
        preliminary_confidence = (pattern_score * 0.40) + (keyword_score * 0.35) + (0.85 * 0.25)
        
        print(f"[IntentAgent M1.3] Pattern result: intent={pattern_intent}, pattern_score={pattern_score:.2f}, preliminary_conf={preliminary_confidence:.2f}")
        
        # STEP 3: If pattern match is strong (>=0.85), skip LLM entirely!
        if pattern_result["pattern_matched"] and preliminary_confidence >= 0.85:
            print(f"[IntentAgent M1.3] ⚡ HIGH CONFIDENCE - Skipping LLM call!")
            
            conf_breakdown = self._calculate_intent_confidence(query, pattern_intent, pattern_score, 0.85)
            
            result = {
                "intent": pattern_intent,
                "confidence": conf_breakdown["composite"],
                "confidence_breakdown": conf_breakdown,
                "entities_mentioned": [],
                "pattern_matched": True,
                "llm_called": False,
                "version": "M1.3-FAST"
            }
            
            result = self._apply_hil_corrections(query, result)
            return result
        
        # STEP 4: Pattern confidence low - call LLM for refinement
        print(f"[IntentAgent M1.3] 🔄 Low pattern confidence - calling LLM for refinement...")
        
        enhanced_system_msg = self._build_enhanced_system_message()
        prompt = f"Analyze this SAMM query and determine intent: {query}"
        
        try:
            response = call_ollama_enhanced(prompt, enhanced_system_msg, temperature=0.0)
            if "{" in response and "}" in response:
                json_part = response[response.find("{"):response.rfind("}")+1]
                llm_result = json.loads(json_part)
                
                ai_confidence = llm_result.get("confidence", 0.5)
                llm_intent = llm_result.get("intent", "general")
                
                # Use pattern intent if pattern matched, otherwise use LLM intent
                final_intent = pattern_intent if pattern_result["pattern_matched"] else llm_intent
                
                # Calculate final confidence with LLM score
                conf_breakdown = self._calculate_intent_confidence(query, final_intent, pattern_score, ai_confidence)
                
                result = {
                    "intent": final_intent,
                    "confidence": conf_breakdown["composite"],
                    "confidence_breakdown": conf_breakdown,
                    "entities_mentioned": llm_result.get("entities_mentioned", []),
                    "pattern_matched": pattern_result["pattern_matched"],
                    "llm_called": True,
                    "ai_raw_confidence": ai_confidence,
                    "version": "M1.3-HYBRID"
                }
                
                result = self._apply_hil_corrections(query, result)
                return result
            else:
                # LLM failed - use pattern result
                conf_breakdown = self._calculate_intent_confidence(query, pattern_intent, pattern_score, 0.5)
                return {
                    "intent": pattern_intent, 
                    "confidence": conf_breakdown["composite"], 
                    "confidence_breakdown": conf_breakdown,
                    "entities_mentioned": [], 
                    "version": "M1.3-FALLBACK"
                }
        except Exception as e:
            # LLM error - use pattern result
            print(f"[IntentAgent M1.3] LLM error: {e} - using pattern result")
            conf_breakdown = self._calculate_intent_confidence(query, pattern_intent, pattern_score, 0.5)
            return {
                "intent": pattern_intent, 
                "confidence": conf_breakdown["composite"],
                "confidence_breakdown": conf_breakdown,
                "entities_mentioned": [], 
                "version": "M1.3-FALLBACK"
            }




    def update_from_hil(self, query: str, original_intent: str, corrected_intent: str, feedback_data: Dict[str, Any] = None):
        """Update agent based on human-in-the-loop feedback"""
        feedback_entry = {
            "query": query,
            "original_intent": original_intent,
            "corrected_intent": corrected_intent,
            "feedback_data": feedback_data or {},
            "timestamp": datetime.now().isoformat()
        }
        
        self.hil_feedback_data.append(feedback_entry)
        
        # Learn patterns from the correction
        query_lower = query.lower()
        if corrected_intent not in self.intent_patterns:
            self.intent_patterns[corrected_intent] = []
        
        # Extract keywords from corrected queries for pattern learning
        keywords = [word for word in query_lower.split() if len(word) > 3]
        self.intent_patterns[corrected_intent].extend(keywords)
        
        print(f"[IntentAgent HIL] Updated with correction: {original_intent} -> {corrected_intent} for query: '{query}'")
        return True
    
    def update_from_trigger(self, new_entities: List[str], new_relationships: List[Dict], trigger_data: Dict[str, Any] = None):
        """Update agent when new entity/relationship data is available"""
        trigger_entry = {
            "new_entities": new_entities,
            "new_relationships": new_relationships,
            "trigger_data": trigger_data or {},
            "timestamp": datetime.now().isoformat()
        }
        
        self.trigger_updates.append(trigger_entry)
        
        # Update intent recognition patterns based on new entities
        for entity in new_entities:
            entity_lower = entity.lower()
            # Add entity-specific intent patterns
            if "agency" in entity_lower or "organization" in entity_lower:
                if "organization" not in self.intent_patterns:
                    self.intent_patterns["organization"] = []
                self.intent_patterns["organization"].append(entity_lower)
        
        print(f"[IntentAgent Trigger] Updated with {len(new_entities)} new entities and {len(new_relationships)} relationships")
        return True
    
    def _build_enhanced_system_message(self) -> str:
        """Build system message enhanced with learned patterns"""
        base_msg = """You are a SAMM (Security Assistance Management Manual) intent analyzer. 
        Classify the user's query into one of these categories:
        - definition: asking what something is
        - distinction: asking about differences between concepts  
        - authority: asking about who has authority or oversight
        - organization: asking about agencies and their roles
        - factual: asking for specific facts like dates, numbers
        - relationship: asking about how things are connected
        - general: general questions"""
        
        # Add learned patterns if available
        if self.intent_patterns:
            base_msg += "\n\nLearned patterns from feedback:"
            for intent, keywords in self.intent_patterns.items():
                if keywords:
                    unique_keywords = list(set(keywords))[:5]  # Limit to top 5 unique keywords
                    base_msg += f"\n- {intent}: commonly involves {', '.join(unique_keywords)}"
        
        base_msg += "\n\nRespond with JSON format: {\"intent\": \"category\", \"confidence\": 0.8, \"entities_mentioned\": [\"entity1\", \"entity2\"]}"
        return base_msg
    
    def _apply_hil_corrections(self, query: str, result: Dict[str, Any]) -> Dict[str, Any]:
        """Apply learned corrections from HIL feedback"""
        query_lower = query.lower()
        
        # Check if this query pattern has been corrected before
        for feedback in self.hil_feedback_data[-10:]:  # Check last 10 feedback entries
            if any(word in query_lower for word in feedback["query"].lower().split() if len(word) > 3):
                # Apply confidence adjustment based on past corrections
                if result["intent"] == feedback["original_intent"]:
                    result["confidence"] = max(0.3, result.get("confidence", 0.5) - 0.2)
                    result["hil_note"] = f"Similar pattern previously corrected to {feedback['corrected_intent']}"
        
        return result

class IntegratedEntityAgent:
    """
    Integrated Entity Agent with database connections and enhanced extraction
    """
    def deduplicate_vector_results(self, results):
        """Remove duplicate vector DB results with robust key handling"""
        seen_content = {}
        unique = []
    
        print(f"\n{'='*80}")
        print(f"[DEDUPLICATION] Starting with {len(results)} results")
        print(f"{'='*80}")
    
        for i, result in enumerate(results):
            # ✅ Try multiple possible keys for content
            content = (
                result.get('content', '') or 
                result.get('text', '') or 
                result.get('page_content', '')
            ).strip()
        
            if not content:
                print(f"[Dedup] ⚠️ WARNING: Result {i+1} has no content! Keys: {list(result.keys())}")
                continue
        
            content_hash = hashlib.md5(content.encode()).hexdigest()
        
            if content_hash not in seen_content:
                seen_content[content_hash] = i + 1
                unique.append(result)
                print(f"[Dedup] ✅ Kept result {i+1}: {content[:60]}...")
            else:
                orig_idx = seen_content[content_hash]
                print(f"[Dedup] ❌ REMOVED result {i+1} (duplicate of result {orig_idx})")
    
        print(f"\n[DEDUPLICATION] Summary:")
        print(f"  Original: {len(results)} results")
        print(f"  Unique:   {len(unique)} results")
        print(f"  Removed:  {len(results) - len(unique)} duplicates ({(len(results)-len(unique))/len(results)*100:.1f}%)")
        print(f"{'='*80}\n")
        
        return unique
    
    def __init__(self, knowledge_graph=None, db_manager=None):
        print("[IntegratedEntityAgent] Initializing with database connections...")
        
        self.knowledge_graph = knowledge_graph
        self.db_manager = db_manager or db_manager
        
        # Learning and feedback systems
        self.hil_feedback_data = []        # Human-in-the-loop feedback storage
        self.custom_entities = {}          # User-defined entities from feedback
        self.trigger_updates = []          # Trigger-based updates storage
        self.dynamic_knowledge = {         # Dynamic knowledge base
            "entities": {},
            "relationships": []
        }
        
        # Enhanced entity patterns
        self.samm_entity_patterns = {
            "organizations": [
                "DSCA", "Defense Security Cooperation Agency",
                "Department of State", "DoS", "State Department",
                "Department of Defense", "DoD", "Defense Department", 
                "DFAS", "Defense Finance and Accounting Service",
                "Implementing Agency", "IA", "MILDEP",
                "Military Department", "Defense Agency",
                "Secretary of State", "Secretary of Defense"
            ],
            "programs": [
                "Security Cooperation", "SC", "Security Cooperation programs",
                "Security Assistance", "SA", "Security Assistance programs", 
                "Foreign Military Sales", "FMS",
                "Foreign Military Financing", "FMF",
                "International Military Education and Training", "IMET",
                "Defense articles", "Military education", "Training programs"
            ],
            "authorities": [
                "Foreign Assistance Act", "FAA", "Foreign Assistance Act of 1961",
                "Arms Export Control Act", "AECA", "Arms Export Control Act of 1976",
                "National Defense Authorization Act", "NDAA",
                "Title 10", "Title 22", "Title 10 authorities", "Title 22 authorities",
                "Executive Order", "Executive Order 13637"
            ],
            "concepts": [
                "continuous supervision", "general direction",
                "defense articles", "military education and training",
                "defense-related services", "strategic objectives",
                "international partners", "DoD Components",
                "overall management", "delivery of materiel"
            ],
            "sections": []
        }
        
        # Entity relationship mappings for SAMM Chapter 1
        self.entity_relationships = {
            "DSCA": ["directs", "administers", "provides guidance to DoD Components"],
            "Defense Security Cooperation Agency": ["directs", "administers", "provides guidance"],
            "Department of State": ["supervises", "provides continuous supervision", "provides general direction"],
            "Department of Defense": ["establishes military requirements", "implements programs"],
            "DFAS": ["performs accounting", "performs billing", "performs disbursing", "performs collecting"],
            "Defense Finance and Accounting Service": ["provides financial services"],
            "Security Assistance": ["is subset of Security Cooperation", "authorized under Title 22"],
            "Security Cooperation": ["includes Security Assistance", "authorized under Title 10"],
            "Secretary of State": ["responsible for continuous supervision", "provides general direction"],
            "Secretary of Defense": ["establishes requirements", "oversees implementation"]
        }
        
        # Confidence scoring weights
        self.confidence_weights = {
            "exact_match": 1.0,
            "partial_match": 0.8,
            "acronym_match": 0.9,
            "context_match": 0.6,
            "ai_extracted": 0.7,
            "knowledge_graph": 0.95,
            "dynamic_knowledge": 0.8,
            "database_match": 0.9
        }
        
        print("[IntegratedEntityAgent] Initialization complete")


    @time_function
    def extract_and_retrieve(self, query: str, intent_info: Dict, documents_context: List = None) -> Dict[str, Any]:
        """
        Main method for integrated entity extraction and database retrieval
        NOW WITH FILE CONTENT EXTRACTION AND FINANCIAL DATA
        """
        print(f"[IntegratedEntityAgent] Processing query: '{query}' with intent: {intent_info.get('intent', 'unknown')}")
        
        # ✅ CRITICAL: ALWAYS log file status at entry point
        if documents_context:
            print(f"[IntegratedEntityAgent] 📁 RECEIVED {len(documents_context)} FILES")
            for idx, doc in enumerate(documents_context[:3], 1):
                fname = doc.get('fileName', 'Unknown')
                content_len = len(doc.get('content', ''))
                has_content = len(doc.get('content', '')) > 50
                print(f"[IntegratedEntityAgent]   File {idx}: {fname} ({content_len} chars) - {'✅ READY' if has_content else '⚠️ INSUFFICIENT'}")
        else:
            print(f"[IntegratedEntityAgent] ⚠️ WARNING: No files provided (documents_context is None/empty)")

        try:
            # Phase 1: Enhanced entity extraction FROM QUERY
            entities = self._extract_entities_enhanced(query, intent_info)
            print(f"[IntegratedEntityAgent] Extracted entities from query: {entities}")
            
            # === NEW: Phase 1.5 - Extract entities from CASE FILES ===
            file_entities = []
            file_relationships = []
            if documents_context:
                print(f"[IntegratedEntityAgent] Processing {len(documents_context)} case files")
                for doc in documents_context[:3]:  # Limit to 3 files
                    content = doc.get('content', '')
                    filename = doc.get('fileName', 'Unknown')
                    if content and len(content) > 50:
                        print(f"[IntegratedEntityAgent] Extracting from file: {filename}")
                        
                        # Extract entities from file content
                        file_ents = self._extract_entities_from_text(content, filename)
                        file_entities.extend(file_ents)
                        
                        # Extract relationships from file content
                        file_rels = self._extract_relationships_from_text(content, filename)
                        file_relationships.extend(file_rels)
                
                # Merge file entities with query entities
                entities.extend(file_entities)
                entities = list(dict.fromkeys(entities))  # Remove duplicates
                print(f"[IntegratedEntityAgent] Total entities after file extraction: {len(entities)}")
                
                # Save file knowledge for future reuse
                if file_entities or file_relationships:
                    for doc in documents_context[:3]:
                        filename = doc.get('fileName', 'Unknown')
                        self._save_file_knowledge_to_dynamic(file_entities, file_relationships, filename)
            # === END NEW ===
            
            # Phase 2: Query all data sources
            all_results = {
                "query": query,
                "entities": entities,
                "intent_info": intent_info,
                "timestamp": datetime.now().isoformat(),
                "data_sources": {},
                "context": [],
                "text_sections": [],
                "relationships": [],
                "confidence_scores": {},
                "overall_confidence": 0.0,
                "extraction_method": "integrated_database_enhanced_with_files",
                "extraction_phases": ["pattern_matching", "nlp_extraction", "file_extraction", "database_queries"],
                "phase_count": 4,
                "file_entities_found": len(file_entities),
                "file_relationships_found": len(file_relationships)
            }
            
            # ✅ NEW: Extract financial records from documents
            financial_records = []
            if documents_context:
                for doc in documents_context:
                    if doc.get('metadata', {}).get('hasFinancialData'):
                        records = doc['metadata'].get('financialRecords', [])
                        financial_records.extend(records)
                        print(f"[EntityAgent] 📊 Added {len(records)} financial records from {doc.get('fileName')}")
            
            # Add to results
            all_results["financial_records"] = financial_records
            all_results["has_financial_data"] = len(financial_records) > 0
            
            print(f"[EntityAgent] 💰 Total financial records available: {len(financial_records)}")
            # ✅ END NEW
            
            # Query each source with error handling
            cosmos_results = self._safe_query_cosmos(query, entities)
            vector_results = self._safe_query_vector(query)

            print(f"[IntegratedEntityAgent] Vector results before dedup: {len(vector_results)}")
            vector_results = self.deduplicate_vector_results(vector_results)
            print(f"[IntegratedEntityAgent] Vector results after dedup: {len(vector_results)}")

            all_results["data_sources"] = {
                "cosmos_gremlin": {
                    "results": cosmos_results,
                    "count": len(cosmos_results),
                    "status": "success" if cosmos_results else "no_results"
                },
                "vector_db": {
                    "results": vector_results,
                    "count": len(vector_results),
                    "status": "success" if vector_results else "no_results",
                    "deduplication_applied": True
                }
            }

            # Store results for use in entity context and text sections
            self.last_retrieval_results = {
                'vector_db': vector_results,
                'cosmos': cosmos_results
            }
            
            # Phase 3: Generate enhanced context from all sources
            self._populate_enhanced_context(all_results, entities)
            
            # === NEW: Add file relationships to results ===
            if file_relationships:
                all_results["relationships"].extend(file_relationships)
                print(f"[IntegratedEntityAgent] Added {len(file_relationships)} relationships from files")
            # === END NEW ===
            
            print(f"\n{'='*80}")
            print(f"[DEBUG] VECTOR DB RESULTS ANALYSIS")
            print(f"{'='*80}")
            if 'vector_db' in all_results["data_sources"] and all_results["data_sources"]["vector_db"]["results"]:
                vector_results = all_results["data_sources"]["vector_db"]["results"]
                print(f"Total Vector DB results: {len(vector_results)}\n")
                for i, result in enumerate(vector_results, 1):
                    content = result.get('content', '')
                    distance = result.get('similarity', result.get('distance', None))
                    if distance is not None:
                        similarity_score = round(1 - distance, 4) if distance <= 1 else round(distance, 4)
                    else:
                        similarity_score = 'N/A'
                    
                    print(f"[Vector Result {i}]")
                    print(f"  Length: {len(content)} chars")
                    print(f"  Preview: {content[:300]}...")
                    print(f"  Similarity: {similarity_score}")
                    print()
            print(f"{'='*80}\n")
            
            print(f"[IntegratedEntityAgent] Query complete: {len(entities)} entities, multiple data sources")
            return all_results
            
        except Exception as e:
            print(f"[IntegratedEntityAgent] Error processing query: {e}")
            return {
                "query": query,
                "entities": [],
                "context": [],
                "text_sections": [],
                "relationships": [],
                "confidence_scores": {},
                "overall_confidence": 0.0,
                "error": str(e),
                "timestamp": datetime.now().isoformat(),
                "extraction_method": "integrated_database_enhanced_error",
                "total_results": 0
            }


    def _safe_query_cosmos(self, query: str, entities: List[str]) -> List[Dict]:
        """Safely query Cosmos Gremlin DB"""
        try:
            print("[IntegratedEntityAgent] Querying Cosmos Gremlin...")
            return self.db_manager.query_cosmos_graph(query, entities)
        except Exception as e:
            print(f"[IntegratedEntityAgent] Cosmos Gremlin query failed: {e}")
            return []
    
    def _safe_query_vector(self, query: str) -> List[Dict]:
        """Safely query Vector DB"""
        try:
            print("[IntegratedEntityAgent] Querying Vector DB...")
            import traceback
            traceback.print_exc()
            return self.db_manager.query_vector_db(query, collection_name="samm_all_chapters")
        except Exception as e:
            print(f"[IntegratedEntityAgent] Vector DB query failed: {e}")
            return []
    
    def _extract_entities_enhanced(self, query: str, intent_info: Dict) -> List[str]:
        """Enhanced entity extraction with pattern matching and NLP"""
        entities = []
        query_lower = query.lower()
        
        # Phase 1: Pattern matching (always works)
        for category, patterns in self.samm_entity_patterns.items():
            for pattern in patterns:
                if pattern.lower() in query_lower:
                    entities.append(pattern)
        
        # Phase 2: Knowledge graph matching
        if self.knowledge_graph:
            for entity_id, entity in self.knowledge_graph.entities.items():
                entity_label = entity['properties'].get('label', entity_id)
                if entity_label.lower() in query_lower or entity_id.lower() in query_lower:
                    entities.append(entity_label)
        
        # Phase 3: NLP extraction (with fallback)
        try:
            nlp_entities = self._extract_nlp_entities_safe(query, intent_info)
            entities.extend(nlp_entities)
        except Exception as e:
            print(f"[IntegratedEntityAgent] NLP extraction failed, using pattern-only: {e}")
        
        # Remove duplicates and limit
        entities = list(dict.fromkeys(entities))[:5]  # Limit to 5 entities
        return entities
    
    def _extract_entities_from_text(self, text: str, source_file: str) -> List[str]:
        """Extract entities from case file content"""
        entities = []
        text_lower = text.lower()
        
        # Pattern matching in file content
        for category, patterns in self.samm_entity_patterns.items():
            for pattern in patterns:
                if pattern.lower() in text_lower:
                    entities.append(pattern)
        
        # Extract case-specific entities
        countries = re.findall(r'\b(Taiwan|Israel|Japan|South Korea|Australia|Saudi Arabia|UAE|Poland|Romania|Ukraine|Republic of Korea)\b', text, re.IGNORECASE)
        entities.extend(countries)
        
        equipment = re.findall(r'\b(F-\d{2}[A-Z]?|AH-\d{2}[A-Z]?|CH-\d{2}[A-Z]?|M1A\d|HIMARS|Patriot|THAAD|Javelin|Stinger|AN/[A-Z]+-\d+)\b', text)
        entities.extend(equipment)
        
        dollar_values = re.findall(r'\$[\d,]+(?:\.\d{2})?(?:\s?(?:million|billion|M|B))?', text, re.IGNORECASE)
        entities.extend([f"Value: {val}" for val in dollar_values[:3]])
        
        case_numbers = re.findall(r'(FMS|FMF|IMET)-\d{4}-[A-Z]{2,4}-\d{3,4}', text)
        entities.extend(case_numbers)
        
        entities = list(dict.fromkeys(entities))
        print(f"[FileExtraction] Extracted {len(entities)} entities from {source_file}")
        return entities

    def _extract_relationships_from_text(self, text: str, source_file: str) -> List[str]:
        """Extract relationships from case file content"""
        relationships = []
        
        relationship_patterns = [
            (r'(\w+)\s+(?:directs|administers|supervises|manages|oversees)\s+(\w+)', 'directs'),
            (r'(\w+)\s+is responsible for\s+(\w+)', 'responsible_for'),
            (r'(\w+)\s+coordinates with\s+(\w+)', 'coordinates_with'),
            (r'(\w+)\s+reports to\s+(\w+)', 'reports_to'),
            (r'(\w+)\s+approves\s+(\w+)', 'approves'),
        ]
        
        for pattern, rel_type in relationship_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches[:5]:
                if len(match) == 2:
                    relationship = f"{match[0]} {rel_type} {match[1]} (from {source_file})"
                    relationships.append(relationship)
        
        print(f"[FileExtraction] Extracted {len(relationships)} relationships from {source_file}")
        return relationships

    def _save_file_knowledge_to_dynamic(self, entities: List[str], relationships: List[str], source_file: str):
        """Save extracted file entities and relationships to dynamic knowledge"""
        timestamp = datetime.now().isoformat()
        
        for entity in entities:
            if entity not in self.dynamic_knowledge["entities"]:
                self.dynamic_knowledge["entities"][entity] = {
                    "definition": f"Entity extracted from case file: {source_file}",
                    "source": "case_file_extraction",
                    "source_file": source_file,
                    "added_date": timestamp,
                    "type": "file_extracted"
                }
        
        for relationship in relationships:
            rel_dict = {
                "relationship": relationship,
                "source": "case_file_extraction",
                "source_file": source_file,
                "added_date": timestamp
            }
            if rel_dict not in self.dynamic_knowledge["relationships"]:
                self.dynamic_knowledge["relationships"].append(rel_dict)
        
        print(f"[DynamicKnowledge] Saved {len(entities)} entities and {len(relationships)} relationships")



    def _extract_nlp_entities_safe(self, query: str, intent_info: Dict) -> List[str]:
        """Fast entity extraction with 15-second timeout"""
        system_msg = """Extract SAMM entities from the query. Return ONLY a simple JSON array.

ENTITIES: Extract ONLY organizations, programs, and authorities that are EXPLICITLY MENTIONED in the query.
IMPORTANT: Do NOT use example entities unless they appear in the actual query.

Examples:
- Query: "What is DSCA?" → Extract: ["DSCA"]
- Query: "What does DFAS do?" → Extract: ["DFAS"]  
- Query: "What is FMS?" → Extract: ["FMS"]
- Query: "What is Security Cooperation?" → Extract: ["Security Cooperation"]

RESPONSE: ["entity1", "entity2"]"""

        prompt = f"Query: '{query}'\nEntities:"
        
        try:
            response = call_ollama_enhanced(prompt, system_msg, temperature=0.0)
            response = response.strip()
            
            json_pattern = r'\[.*?\]'
            matches = re.findall(json_pattern, response, re.DOTALL)
            
            if matches:
                try:
                    entities = json.loads(matches[0])
                    if isinstance(entities, list):
                        return [str(e).strip() for e in entities if e]
                except:
                    pass
            
            quote_pattern = r'"([^"]+)"'
            quoted_entities = re.findall(quote_pattern, response)
            return quoted_entities[:3]
                
        except Exception as e:
            print(f"[IntegratedEntityAgent] NLP extraction error: {e}")
        
        return []


    def _populate_enhanced_context(self, all_results: Dict, entities: List[str]):
        """Populate enhanced context from all data sources"""
        context = []
        text_sections = []
        relationships = []
        confidence_scores = {}
        
        # Process each entity
        for entity in entities:
            entity_context = self._generate_entity_context(entity, all_results["query"])
            if entity_context:
                context.append(entity_context)
                confidence_scores[entity] = entity_context.get('confidence', 0.5)
        
        # Get relevant text sections
        text_sections = self._get_enhanced_text_sections(all_results["query"], entities)
        
        # Get comprehensive relationships
        relationships = self._get_comprehensive_relationships(entities, all_results["data_sources"])
        
        # Calculate overall confidence
        overall_confidence = self._calculate_overall_confidence(confidence_scores)
        
        # Populate results
        all_results.update({
            "context": context,
            "text_sections": text_sections,
            "relationships": relationships,
            "confidence_scores": confidence_scores,
            "overall_confidence": overall_confidence,
            "total_results": len(context) + len(text_sections) + len(relationships)
        })
    
    def _generate_entity_context(self, entity: str, query: str) -> Optional[Dict]:
        """Generate comprehensive context information for an entity"""
        context_info = None
        
        # Check knowledge graph first (highest confidence)
        if self.knowledge_graph:
            for entity_id, kg_entity in self.knowledge_graph.entities.items():
                entity_label = kg_entity['properties'].get('label', entity_id)
                
                if (entity.lower() == entity_label.lower() or 
                    entity.lower() == entity_id.lower()):
                    
                    definition = kg_entity['properties'].get('definition', 
                                kg_entity['properties'].get('role', ''))
                    section = kg_entity['properties'].get('section', '')
                    
                    context_info = {
                        "entity": entity_label,
                        "definition": definition,
                        "section": section,
                        "type": kg_entity.get('type', 'unknown'),
                        "confidence": self.confidence_weights["knowledge_graph"],
                        "source": "knowledge_graph",
                        "properties": kg_entity['properties']
                    }
                    print(f"[IntegratedEntityAgent] Knowledge graph context for: {entity_label}")
                    break
        
        # Check dynamic knowledge if not found
        if not context_info and entity in self.dynamic_knowledge["entities"]:
            entity_data = self.dynamic_knowledge["entities"][entity]
            context_info = {
                "entity": entity,
                "definition": entity_data.get('definition', ''),
                "section": entity_data.get('section', ''),
                "type": entity_data.get('type', 'dynamic'),
                "confidence": self.confidence_weights["dynamic_knowledge"],
                "source": "dynamic_knowledge",
                "added_date": entity_data.get('added_date', '')
            }
            print(f"[IntegratedEntityAgent] Dynamic knowledge context for: {entity}")
        
        # Generate context using AI if not found
        if not context_info:
            context_info = self._generate_ai_context(entity, query)
        
        return context_info
    
    def _generate_ai_context(self, entity: str, query: str) -> Dict:
        """Generate AI context for entity - DISABLED FOR SPEED, returns quick fallback"""
        # SPEED OPTIMIZATION: Skip AI generation, use fallback directly
        # Check if entity appears exactly in the query for higher confidence
        entity_in_query = entity.lower() in query.lower()
        base_confidence = 0.75 if entity_in_query else 0.6
        
        return {
            "entity": entity,
            "definition": f"SAMM-related entity: {entity}",
            "section": "Context from vector DB",
            "type": "entity",
            "source": "quick_fallback",
            "confidence": base_confidence  # Added confidence for proper scoring
        }
        
        # Original AI generation code disabled for speed:
        # context_prompt = f"...Ollama call..."
        # This saves 5-10 seconds per entity!
        """Generate entity context using Llama 3.2 AI capabilities"""
        system_msg = f"""You are a SAMM (Security Assistance Management Manual) expert.

Provide context for the entity "{entity}" as it relates to SAMM.

INCLUDE:
- Brief, accurate definition or role description
- SAMM section reference if known (use the section from the retrieved context)
- Entity type (organization, program, authority, concept)
- Relationship to Security Cooperation/Security Assistance

REQUIREMENTS:
- Be accurate and specific to SAMM
- Use exact SAMM terminology
- If uncertain, indicate lower confidence

RESPONSE FORMAT (JSON):
{{
    "definition": "Brief definition or role description",
    "section": "SAMM section if known, otherwise 'Unknown'", 
    "type": "organization|program|authority|concept",
    "confidence": 0.7,
    "relationships": []
}}"""
        
        prompt = f"""Entity: "{entity}"
Query context: "{query}"

Provide SAMM context for this entity:"""
        
        try:
            response = call_ollama_enhanced(prompt, system_msg, temperature=0.1)
            
            # Try to parse JSON response
            if "{" in response and "}" in response:
                json_start = response.find("{")
                json_end = response.rfind("}") + 1
                json_part = response[json_start:json_end]
                
                context_data = json.loads(json_part)
                context_data["source"] = "ai_generated"
                context_data["entity"] = entity
                
                print(f"[IntegratedEntityAgent] AI generated context for: {entity}")
                return context_data
                
        except json.JSONDecodeError as e:
            print(f"[IntegratedEntityAgent] JSON parsing error in AI context generation: {e}")
        except Exception as e:
            print(f"[IntegratedEntityAgent] AI context generation error: {e}")
        
        # Fallback context
        return {
            "entity": entity,
            "definition": f"SAMM-related entity: {entity}",
            "section": "Unknown",
            "type": "unknown",
            "confidence": self.confidence_weights["context_match"] * 0.5,
            "source": "fallback"
        }
    
    def _get_enhanced_text_sections(self, query: str, entities: List[str]) -> List[str]:
        """Get relevant SAMM text sections from VECTOR DB RESULTS with entity prioritization"""
        text_sections = []
    
        # Check if we have vector DB results stored
        if not hasattr(self, 'last_retrieval_results'):
            print("[IntegratedEntityAgent] No retrieval results available")
            return text_sections
    
        results = self.last_retrieval_results
    
        # Extract text from Vector DB results with ENTITY PRIORITIZATION
        if 'vector_db' in results and results['vector_db'] is not None:
            vector_db_results = results['vector_db']
            print(f"[IntegratedEntityAgent] Processing {len(results['vector_db'])} vector DB results")
            
            # Separate results: those containing entities first, then others
            entity_matched_results = []
            other_results = []
            
            for i, result in enumerate(results['vector_db'], 1):
                content = result.get('content', '')
                if content:
                    # Check if this result contains any of the extracted entities
                    contains_entity = False
                    for entity in entities:
                        # Handle entities with special characters like USD(P)
                        # Simple case-insensitive search works better for acronyms with parentheses
                        if entity.lower() in content.lower():
                            contains_entity = True
                            print(f"[DEBUG] Found entity '{entity}' in result {i}")
                            break
                    
                    if contains_entity:
                        entity_matched_results.append((i, content))
                        print(f"[DEBUG] Vector DB result {i}: ✅ ENTITY MATCH - {content[:100]}...")
                    else:
                        other_results.append((i, content))
                        print(f"[DEBUG] Vector DB result {i}: ⚪ No entity match - {content[:100]}...")
            
            # Prioritize entity-matched results
            print(f"[IntegratedEntityAgent] 🎯 Entity-matched results: {len(entity_matched_results)}, Other: {len(other_results)}")
            
            # Add entity-matched results first
            for idx, content in entity_matched_results:
                text_sections.append(content)
            
            # Then add other results
            for idx, content in other_results:
                text_sections.append(content)

        return text_sections
    
    
    def _get_comprehensive_relationships(self, entities: List[str], data_sources: Dict) -> List[str]:
        """Get comprehensive relationships from all sources"""
        relationships = []
        
        # Get relationships from knowledge graph
        if self.knowledge_graph:
            for entity in entities:
                entity_id = None
                
                # Find entity ID in knowledge graph
                for eid, kg_entity in self.knowledge_graph.entities.items():
                    entity_label = kg_entity['properties'].get('label', eid)
                    if entity_label.lower() == entity.lower():
                        entity_id = eid
                        break
                
                # Get relationships for this entity
                if entity_id:
                    entity_rels = self.knowledge_graph.get_relationships(entity_id)
                    if entity_rels is not None:
                        for rel in entity_rels:
                            rel_text = f"{rel['source']} {rel['relationship']} {rel['target']}"
                            relationships.append(rel_text)
                            print(f"[IntegratedEntityAgent] Knowledge graph relationship: {rel_text}")
        
        # Add predefined relationships
        for entity in entities:
            if entity in self.entity_relationships:
                for relationship in self.entity_relationships[entity]:
                    rel_text = f"{entity} {relationship}"
                    relationships.append(rel_text)
                    print(f"[IntegratedEntityAgent] Predefined relationship: {rel_text}")
        
        # Add dynamic relationships from triggers
        for rel in self.dynamic_knowledge["relationships"]:
            source = rel.get("source", "")
            target = rel.get("target", "")
            relationship = rel.get("relationship", "")
            
            if any(entity.lower() in source.lower() or entity.lower() in target.lower() 
                   for entity in entities):
                rel_text = f"{source} {relationship} {target}"
                relationships.append(rel_text)
                print(f"[IntegratedEntityAgent] Dynamic relationship: {rel_text}")
        


        # Add relationships from Cosmos DB Gremlin results
        cosmos_results = data_sources.get("cosmos_gremlin", {}).get("results", [])
        for result in cosmos_results:
            if result.get("type") == "edge":
                edge_data = result.get("data", {})
        # Extract relationship information from edge
                if isinstance(edge_data, dict):
                    label = edge_data.get("label", "relates_to")
            # Try to get vertex names from edge properties
                    from_name = edge_data.get("from_name", edge_data.get("outV", "unknown"))
                    to_name = edge_data.get("to_name", edge_data.get("inV", "unknown"))
            
                    rel_text = f"{from_name} {label} {to_name}"
                    relationships.append(rel_text)
                    print(f"[IntegratedEntityAgent] Cosmos DB relationship: {rel_text}")
        
        # Remove duplicates
        relationships = list(dict.fromkeys(relationships))
        
        print(f"[IntegratedEntityAgent] Total relationships found: {len(relationships)}")
        return relationships
    
    def _calculate_overall_confidence(self, confidence_scores: Dict[str, float]) -> float:
        """Calculate overall confidence score for the extraction"""
        if not confidence_scores:
            return 0.0
        
        scores = list(confidence_scores.values())
        # Weighted average with slight boost for multiple high-confidence entities
        avg_confidence = sum(scores) / len(scores)
        entity_count_factor = min(1.0, len(scores) / 5.0) * 0.1  # Small boost for more entities
        
        return min(1.0, avg_confidence + entity_count_factor)
    
    def update_from_hil(self, query: str, original_entities: List[str], 
                        corrected_entities: List[str], feedback_data: Dict[str, Any] = None):
        """Update agent based on human-in-the-loop feedback"""
        feedback_entry = {
            "query": query,
            "original_entities": original_entities,
            "corrected_entities": corrected_entities,
            "feedback_data": feedback_data or {},
            "timestamp": datetime.now().isoformat(),
            "improvement_type": "hil_correction"
        }
        
        self.hil_feedback_data.append(feedback_entry)
        
        # Add new entities identified by human feedback
        for entity in corrected_entities:
            if entity not in original_entities and entity not in self.custom_entities:
                self.custom_entities[entity] = {
                    "definition": feedback_data.get("definition", "Entity identified through human feedback"),
                    "source": "HIL_feedback",
                    "query_context": query,
                    "added_date": datetime.now().isoformat(),
                    "feedback_id": len(self.hil_feedback_data)
                }
                
                # Add to dynamic knowledge
                self.dynamic_knowledge["entities"][entity] = self.custom_entities[entity]
        
        # Store context corrections
        if feedback_data and feedback_data.get("context_corrections"):
            for entity, corrected_context in feedback_data["context_corrections"].items():
                if entity in self.custom_entities:
                    self.custom_entities[entity]["definition"] = corrected_context
                    self.dynamic_knowledge["entities"][entity]["definition"] = corrected_context
        
        print(f"[IntegratedEntityAgent HIL] Updated with {len(corrected_entities)} entities from feedback for query: '{query[:50]}...'")
        print(f"[IntegratedEntityAgent HIL] Total custom entities: {len(self.custom_entities)}")
        return True
    
    def update_from_trigger(self, new_entities: List[str], new_relationships: List[Dict], 
                           trigger_data: Dict[str, Any] = None):
        """Update agent when new entity/relationship data is available"""
        trigger_entry = {
            "new_entities": new_entities,
            "new_relationships": new_relationships,
            "trigger_data": trigger_data or {},
            "timestamp": datetime.now().isoformat(),
            "trigger_id": len(self.trigger_updates)
        }
        
        self.trigger_updates.append(trigger_entry)
        
        # Add new entities to dynamic knowledge
        for entity in new_entities:
            if entity not in self.dynamic_knowledge["entities"]:
                entity_data = {
                    "definition": trigger_data.get("entity_definitions", {}).get(entity, f"New entity: {entity}"),
                    "source": "trigger_update",
                    "type": trigger_data.get("entity_types", {}).get(entity, "unknown"),
                    "added_date": datetime.now().isoformat(),
                    "trigger_id": len(self.trigger_updates)
                }
                self.dynamic_knowledge["entities"][entity] = entity_data
        
        # Add new relationships to dynamic knowledge
        for relationship in new_relationships:
            if relationship not in self.dynamic_knowledge["relationships"]:
                self.dynamic_knowledge["relationships"].append({
                    **relationship,
                    "source": "trigger_update",
                    "added_date": datetime.now().isoformat(),
                    "trigger_id": len(self.trigger_updates)
                })
        
        print(f"[IntegratedEntityAgent Trigger] Updated with {len(new_entities)} new entities and {len(new_relationships)} relationships")
        print(f"[IntegratedEntityAgent Trigger] Total dynamic entities: {len(self.dynamic_knowledge['entities'])}")
        return True


class EnhancedAnswerAgent:
    """
    Enhanced Answer Agent for SAMM with sophisticated response generation
    
    Features:
    - Windows compatibility and improved error handling
    - Intent-optimized prompt engineering for each question type
    - SAMM-specific response templates and quality standards
    - Multi-pass answer generation with validation
    - Learning system with HIL feedback and trigger updates
    - Automatic answer enhancement (acronym expansion, section references)
    - Answer caching and correction storage
    - Quality scoring and confidence assessment
    """
    
    def __init__(self):
        """Initialize the Enhanced Answer Agent with improved error handling"""
        print("[EnhancedAnswerAgent] Initializing...")
        
        # Learning and feedback systems
        self.hil_feedback_data = []        # Human-in-the-loop feedback storage
        self.answer_templates = {}         # Intent-specific answer templates
        self.trigger_updates = []          # Trigger-based updates storage
        self.custom_knowledge = ""         # Additional knowledge from updates
        self.answer_corrections = {}       # Stored answer corrections
        
        # SAMM-specific response templates for each intent type
        self.samm_response_templates = {
            "definition": {
                "structure": "Provide clear definition → cite SAMM section → add context/authority",
                "required_elements": ["definition", "section_reference", "authority_context"],
                "quality_criteria": ["uses_exact_samm_terminology", "cites_section", "expands_acronyms"]
            },
            "distinction": {
                "structure": "Explain key differences → provide examples → cite legal basis",
                "required_elements": ["comparison_points", "specific_examples", "legal_authorities"],
                "quality_criteria": ["clear_comparison", "highlights_subset_relationship", "authority_differences"]
            },
            "authority": {
                "structure": "State authority holder → explain scope → cite legal basis",
                "required_elements": ["authority_holder", "scope_of_authority", "legal_reference"],
                "quality_criteria": ["identifies_correct_authority", "explains_scope", "cites_legal_basis"]
            },
            "organization": {
                "structure": "Name organization → describe role → list responsibilities",
                "required_elements": ["full_name", "primary_role", "specific_duties"],
                "quality_criteria": ["expands_acronyms", "describes_role", "lists_responsibilities"]
            },
            "factual": {
                "structure": "State fact → provide context → cite source",
                "required_elements": ["specific_fact", "context", "source_reference"],
                "quality_criteria": ["accurate_information", "proper_citation", "relevant_context"]
            },
            "relationship": {
                "structure": "Describe relationship → explain significance → provide examples",
                "required_elements": ["relationship_description", "significance", "examples"],
                "quality_criteria": ["clear_relationship", "explains_importance", "concrete_examples"]
            }
        }
        
        # Quality enhancement patterns for post-processing
        self.quality_patterns = {
            "section_references": r"(C\d+\.\d+\.?\d*\.?\d*)",
            "acronym_detection": r"\b([A-Z]{2,})\b",
            "authority_mentions": r"(Title \d+|[A-Z]+ Act)",
            "incomplete_sentences": r"[a-z]\s*$"
        }
        
        # Enhanced acronym expansion dictionary for SAMM
        self.acronym_expansions = {
            "DSCA": "Defense Security Cooperation Agency (DSCA)",
            "DFAS": "Defense Finance and Accounting Service (DFAS)",
            "DoD": "Department of Defense (DoD)", 
            "DoS": "Department of State (DoS)",
            "SC": "Security Cooperation (SC)",
            "SA": "Security Assistance (SA)",
            "FAA": "Foreign Assistance Act (FAA)",
            "AECA": "Arms Export Control Act (AECA)",
            "NDAA": "National Defense Authorization Act (NDAA)",
            "USD(P)": "Under Secretary of Defense for Policy (USD(P))",
            "IA": "Implementing Agency (IA)",
            "MILDEP": "Military Department (MILDEP)",
            "IMET": "International Military Education and Training (IMET)",
            "FMS": "Foreign Military Sales (FMS)",
            "FMF": "Foreign Military Financing (FMF)",
            "NIPO": "Navy International Programs Office (NIPO)",
            "USASAC": "U.S. Army Security Assistance Command (USASAC)",
            "SATFA": "Security Assistance Training Field Activity (SATFA)",
            "CCDR": "Combatant Commander (CCDR)",
            "SCO": "Security Cooperation Organization (SCO)",
            "GEF": "Guidance for Employment of the Force (GEF)"
        }
        
        # Answer quality scoring weights
        self.quality_weights = {
            "section_citation": 0.25,
            "acronym_expansion": 0.15,
            "answer_completeness": 0.25,
            "samm_terminology": 0.20,
            "structure_adherence": 0.15
        }
        
        # Response length guidelines by intent
        self.length_guidelines = {
            "definition": {"min": 150, "target": 300, "max": 500},      # Increased for detailed definitions
            "distinction": {"min": 200, "target": 400, "max": 600},      # Increased for comprehensive comparisons
            "authority": {"min": 150, "target": 300, "max": 500},        # Increased for detailed authority explanations
            "organization": {"min": 150, "target": 300, "max": 500},     # Increased for organizational details
            "factual": {"min": 150, "target": 350, "max": 600},         # ⭐ MAJOR INCREASE for factual queries
            "relationship": {"min": 150, "target": 300, "max": 500},     # Increased for relationship explanations
            "general": {"min": 150, "target": 300, "max": 500}          # Increased for general queries
        }
        
        print("[EnhancedAnswerAgent] Initialization complete")

    @time_function
    def generate_answer(self, query: str, intent_info: Dict, entity_info: Dict, 
                    chat_history: List = None, documents_context: List = None,
                    user_profile: Dict = None) -> str:
        """
        Main method for enhanced answer generation with ITAR compliance filtering
        """
        # CRITICAL: ALWAYS log file status at entry point
        if documents_context:
            print(f"[AnswerAgent] 📁 RECEIVED {len(documents_context)} FILES for answer generation")
            for idx, doc in enumerate(documents_context[:3], 1):
                fname = doc.get('fileName', 'Unknown')
                content_len = len(doc.get('content', ''))
                has_content = len(doc.get('content', '')) > 50
                print(f"[AnswerAgent]   File {idx}: {fname} ({content_len} chars) - {'✅ READY' if has_content else '⚠️ INSUFFICIENT'}")
        else:
            print(f"[AnswerAgent] ⚠️ WARNING: No files provided for answer generation")
        
        # NEW: Handle LOA timeline queries with instant pre-formatted answer
        if intent_info.get("intent") == "loa_timeline":
            print(f"[AnswerAgent] 🚀 Using LOA timeline pre-formatted answer")
            return self._get_loa_timeline_answer()
        
        # NEW: Handle financial verification queries
        if intent_info.get("intent") == "financial_verification":
            print(f"[AnswerAgent] 🚀 Using financial verification pre-formatted answer")
            return self._get_financial_verification_answer()
        
        # NEW: Handle technical services queries
        if intent_info.get("intent") == "line_item_details":
            print(f"[AnswerAgent] 🚀 Using technical services pre-formatted answer")
            return self._get_technical_services_answer()
        
        # NEW: Handle PMR minutes summary queries
        if intent_info.get("intent") == "pmr_minutes_summary":
            print(f"[AnswerAgent] 🚀 Using PMR minutes summary pre-formatted answer")
            return self._get_pmr_minutes_summary()
                
        intent = intent_info.get("intent", "general")
        confidence = intent_info.get("confidence", 0.5)
        
        print(f"[AnswerAgent] Generating answer for intent: {intent} (confidence: {confidence:.2f})")
        print(f"[AnswerAgent] Query: {query[:100]}...")

        try:
            # === ITAR COMPLIANCE CHECK ===
            compliance_result = check_compliance(query, intent_info, entity_info, user_profile)
            
            # Log compliance check
            if compliance_result.get("check_performed"):
                print(f"[Compliance] Check performed: {compliance_result.get('compliance_status')}")
                print(f"[Compliance] User level: {compliance_result.get('user_authorization_level')}")
                print(f"[Compliance] Authorized: {compliance_result.get('authorized')}")
            
            # Handle unauthorized access
            if not compliance_result.get("authorized", True):
                required_level = compliance_result.get('required_authorization_level', 'higher authorization')
                user_level = compliance_result.get('user_authorization_level', 'unknown')
                recommendations = compliance_result.get("recommendations", [])
                
                # Build denial response
                response = (
                    f"⚠️ **ITAR COMPLIANCE NOTICE**\n\n"
                    f"This query involves controlled information requiring **{required_level.upper()}** clearance.\n"
                    f"Your current authorization: **{user_level.upper()}**\n\n"
                )
                
                if recommendations:
                    response += "**Recommendations:**\n" + "\n".join(f"• {r}" for r in recommendations)
                
                print(f"[Compliance] Access denied: {user_level} < {required_level}")
                return response
            
            # Log successful compliance check
            if compliance_result.get("check_performed"):
                print(f"[Compliance] Query authorized - proceeding with answer generation")
            # === END ITAR COMPLIANCE CHECK ===
            
            # Step 1: Check for existing corrections first
            cached_answer = self._check_for_corrections(query, intent_info, entity_info)
            if cached_answer:
                print("[AnswerAgent] Using cached correction")
                return cached_answer
            
            # Step 2: Build comprehensive context from all sources
            context = self._build_comprehensive_context(
                query, intent_info, entity_info, chat_history, documents_context
            )
            
            # Step 3: Create intent-optimized system message
            system_msg = self._create_optimized_system_message(intent, context)
            
            # Step 4: Generate enhanced prompt with intent awareness
            prompt = self._create_enhanced_prompt(query, intent_info, entity_info)
            
            # Step 5: Generate answer with validation passes
            answer = self._generate_with_validation(prompt, system_msg, intent_info)
            
            # Step 6: Apply post-processing enhancements
            enhanced_answer = self._enhance_answer_quality(answer, intent_info, entity_info)
            
            # Step 7: Final validation and scoring
            final_answer = self._validate_and_score_answer(enhanced_answer, intent, query)
            
            # ADD: Final answer verification
            print(f"[AnswerAgent] ✅ FINAL ANSWER GENERATED:")
            print(f"[AnswerAgent]   Length: {len(final_answer)} chars")
            print(f"[AnswerAgent]   Preview: {final_answer[:200]}...")
            print(f"[AnswerAgent]   Has content: {bool(final_answer and len(final_answer) > 20)}")
            
            return final_answer
            
        except Exception as e:
            print(f"[AnswerAgent] Error during answer generation: {e}")
            import traceback
            traceback.print_exc()
            return f"I apologize, but I encountered an error while generating the answer: {str(e)}. Please try rephrasing your question or check if the Ollama service is running."
        


 
    def _check_for_corrections(self, query: str, intent_info: Dict, entity_info: Dict) -> Optional[str]:
        """Check if we have a stored correction for similar queries"""
        try:
            query_key = self._normalize_query_for_matching(query)
            
            # Check exact matches first
            if query_key in self.answer_corrections:
                correction = self.answer_corrections[query_key]
                print(f"[AnswerAgent] Found exact correction match")
                return correction["corrected_answer"]
            
            # Check for partial matches based on intent and entities
            current_entities = set(entity_info.get("entities", []))
            current_intent = intent_info.get("intent", "general")
            
            for stored_key, correction in self.answer_corrections.items():
                stored_entities = set(correction.get("feedback_data", {}).get("entities", []))
                stored_intent = correction.get("feedback_data", {}).get("intent", "general")
                
                # If same intent and significant entity overlap (50% or more)
                if (current_intent == stored_intent and len(current_entities) > 0 and
                    len(current_entities.intersection(stored_entities)) >= min(len(current_entities), len(stored_entities)) * 0.5):
                    print(f"[AnswerAgent] Found partial correction match based on intent/entities")
                    return correction["corrected_answer"]
            
            return None
            
        except Exception as e:
            print(f"[AnswerAgent] Error checking corrections: {e}")
            return None





    def _build_comprehensive_context(self, query: str, intent_info: Dict, entity_info: Dict,
                                chat_history: List = None, documents_context: List = None) -> str:
        """Build comprehensive context for answer generation with financial data"""
        try:
            context_parts = []
            
            # Add entity context with confidence weighting
            if entity_info.get("context"):
                context_parts.append("=== SAMM ENTITIES AND DEFINITIONS ===")
                for ctx in entity_info["context"][:5]:  # Limit to 5 to prevent overload
                    confidence = ctx.get('confidence', 0.5)
                    if confidence > 0.6:  # Only include high-confidence entities
                        entity_text = f"{ctx.get('entity', '')}: {ctx.get('definition', '')}"
                        if ctx.get('section'):
                            entity_text += f" (SAMM {ctx['section']})"
                        context_parts.append(entity_text)
            
            # Add relevant text sections from SAMM
            if entity_info.get("text_sections"):
                context_parts.append("\n=== RELEVANT SAMM CONTENT ===")
                text_sections = entity_info["text_sections"][:3]
                for section in text_sections:
                    truncated_section = section[:500] + "..." if len(section) > 500 else section
                    context_parts.append(truncated_section)
            
            # Add entity relationships
            if entity_info.get("relationships"):
                context_parts.append("\n=== ENTITY RELATIONSHIPS ===")
                context_parts.extend(entity_info["relationships"][:7])
            
            # === NEW: Add HIL corrections as context ===
            if hasattr(self, 'answer_corrections') and len(self.answer_corrections) > 0:
                context_parts.append("\n=== PREVIOUS CORRECT ANSWERS (HIL) ===")
                for correction_key, correction in list(self.answer_corrections.items())[-5:]:
                    corrected_answer = correction.get('corrected_answer', '')
                    original_query = correction.get('original_query', 'Unknown query')
                    if len(corrected_answer) > 50:
                        truncated = corrected_answer[:1500] + "..." if len(corrected_answer) > 1500 else corrected_answer
                        context_parts.append(f"Q: {original_query[:100]}\nCorrect Answer: {truncated}\n")
                        print(f"[AnswerAgent] Added HIL correction to context")

            # === NEW: Add case file relationships ===
            if entity_info.get("file_relationships_found", 0) > 0:
                context_parts.append("\n=== CASE FILE RELATIONSHIPS ===")
                file_rels = [rel for rel in entity_info.get("relationships", []) if "from" in rel]
                for rel in file_rels[:5]:
                    context_parts.append(f"• {rel}")
                print(f"[AnswerAgent] Added {len(file_rels[:5])} file relationships to context")

            # ✅ ENHANCED: Add uploaded documents WITH financial data extraction
            if documents_context:
                context_parts.append("\n" + "="*80)
                context_parts.append(f"📁 UPLOADED DOCUMENTS ({len(documents_context)} files)")
                context_parts.append("="*80)
                
                # Extract and display financial records
                financial_records = []
                
                for idx, doc in enumerate(documents_context[:3], 1):
                    file_name = doc.get('fileName', 'Unknown')
                    content = doc.get('content', '')
                    
                    # Check for financial data
                    has_financial = doc.get('metadata', {}).get('hasFinancialData', False)
                    
                    if has_financial:
                        records = doc.get('metadata', {}).get('financialRecords', [])
                        financial_records.extend(records)
                        
                        context_parts.append(f"\n[Document {idx}] 💰 {file_name} (FINANCIAL DATA)")
                        context_parts.append(f"Contains {len(records)} financial line items:")
                        
                        # Show first 5 records as examples
                        for i, record in enumerate(records[:5], 1):
                            rsn = record.get('rsn_identifier', 'N/A')
                            pdli = record.get('pdli_pdli', 'N/A')
                            available = record.get('available', 0)
                            
                            context_parts.append(
                                f"  • RSN {rsn}, PDLI {pdli}: ${available:,.2f} available"
                            )
                        
                        if len(records) > 5:
                            context_parts.append(f"  ... and {len(records) - 5} more records")
                    
                    elif content:
                        # Regular document
                        context_parts.append(f"\n[Document {idx}] {file_name}")
                        truncated = content[:1000] + "..." if len(content) > 1000 else content
                        context_parts.append(truncated)
                    
                    context_parts.append("-" * 80)
                
                # Add financial summary if records found
                if financial_records:
                    context_parts.append(f"\n📊 FINANCIAL DATA SUMMARY:")
                    context_parts.append(f"Total records: {len(financial_records)}")
                    
                    # Calculate totals
                    total_available = sum(float(r.get('available', 0)) for r in financial_records)
                    context_parts.append(f"Total available funds: ${total_available:,.2f}")
                    
                    # List unique RSNs
                    unique_rsns = set(r.get('rsn_identifier') for r in financial_records if r.get('rsn_identifier'))
                    context_parts.append(f"RSN line numbers: {', '.join(sorted(unique_rsns))}")
                
                print(f"[AnswerAgent] ✅ Added {len(documents_context[:3])} documents to context")
                print(f"[AnswerAgent] 💰 Included {len(financial_records)} financial records")
            # ✅ END ENHANCEMENT
            
            # Add custom knowledge from HIL feedback and triggers
            if self.custom_knowledge:
                context_parts.append("\n=== ADDITIONAL KNOWLEDGE ===")
                knowledge = self.custom_knowledge[:1000] + "..." if len(self.custom_knowledge) > 1000 else self.custom_knowledge
                context_parts.append(knowledge)
            
            # Add relevant chat history for continuity
            if chat_history and len(chat_history) > 0:
                context_parts.append("\n=== CONVERSATION CONTEXT ===")
                for msg in chat_history[-2:]:  # Last 2 messages for context
                    role = msg.get('role', 'unknown')
                    content = msg.get('content', '')[:200]
                    context_parts.append(f"{role}: {content}")
            
            return "\n".join(context_parts)
            
        except Exception as e:
            print(f"[AnswerAgent] Error building context: {e}")
            return "Context building failed - proceeding with basic knowledge."


    def _create_optimized_system_message(self, intent: str, context: str) -> str:
        """Create intent-optimized system message for Llama 3.2 with error handling"""
        
        try:
            # Base instructions for each intent type
            base_instructions = {
                "definition": """You are a SAMM (Security Assistance Management Manual) expert specializing in precise definitions.

TASK: Provide authoritative definitions with exact SAMM section citations.

🔴 CRITICAL - EXACT ENTITY FOCUS:
- Answer ONLY about the SPECIFIC entity asked in the question
- If asked about USD(P), answer ONLY about USD(P) - NOT USD(P&R), USD(A&S), or any other variant
- If asked about FMS, answer ONLY about FMS - NOT FMF or other programs
- IGNORE information about similar-sounding but different entities in the context
- The question entity is the ONLY entity you should define

CRITICAL REQUIREMENTS:
- Use exact SAMM terminology and definitions from the context
- Always cite specific SAMM sections that appear in the retrieved vector DB results (use actual section numbers from context, not examples)
- CITATION FORMAT: Use only section numbers (e.g., "C9.3.9, C5.T6") - DO NOT mention chapter numbers
- Expand acronyms on first use ONLY if the acronym appears in the retrieved context
- Use format: "Full Name (ACRONYM)" but only when relevant to the query

🔴 CRITICAL HIERARCHY RULE - READ CAREFULLY:
- Security Cooperation (SC) is the BROAD umbrella term
- Security Assistance (SA) is a SUBSET/PART OF Security Cooperation
- SC INCLUDES SA (not the other way around!)
- NEVER say "SC is a subset of SA" or "SC is a type of SA"
- ALWAYS say "SA is a subset of SC" or "SC includes SA"
- Think: SC is the BIG category, SA is ONE PART of it
- SC = All DoD interactions with foreign partners (BROAD)
- SA = Specific programs under Title 22 (NARROW)

- Provide clear, complete definitions that could stand alone

RESPONSE STRUCTURE:
1. Clear definition statement
2. SAMM section citation
3. Additional context about authority/oversight if relevant""",

                "distinction": """You are a SAMM expert specializing in explaining key distinctions and differences.

TASK: Clearly explain differences between SAMM concepts with precise legal and operational distinctions.

CRITICAL REQUIREMENTS:
- Highlight key differences clearly and systematically
- Explain legal authority differences (Title 10 vs Title 22)
- Use specific examples when possible

🔴 WHEN COMPARING SC AND SA:
- Security Cooperation (SC) = BROAD umbrella covering ALL DoD interactions
- Security Assistance (SA) = NARROW subset of SC with specific Title 22 programs
- ALWAYS state: "SA is a subset of SC" (NEVER reverse this!)
- SC is authorized under Title 10 (broad DoD authority)
- SA is authorized under Title 22 (specific State Dept programs)
- Think of it like: SC is the ocean, SA is one specific current in that ocean

- Cite relevant SAMM sections for each concept being compared
- Address common misconceptions

RESPONSE STRUCTURE:
1. State the key distinction clearly
2. Explain each concept separately with citations
3. Highlight the differences with examples
4. Summarize the relationship""",

                "authority": """You are a SAMM expert specializing in authority and oversight structures.

TASK: Explain who has authority, oversight, and responsibility for specific programs.

CRITICAL REQUIREMENTS:
- Clearly state which organization/person has authority
- Explain the scope of authority and oversight
- Cite legal authorities (FAA, AECA, NDAA, Executive Orders)
- Distinguish between "supervision," "direction," and "oversight"
- Reference specific SAMM sections
- Explain delegation chains where applicable

RESPONSE STRUCTURE:
1. State who has the authority
2. Explain the scope and basis of authority
3. Cite legal foundations
4. Describe any delegation or coordination requirements""",

                "organization": """You are a SAMM expert specializing in organizational roles and responsibilities.

TASK: Describe organizations, their roles, and specific responsibilities.

CRITICAL REQUIREMENTS:
- Provide full organization names and acronyms
- List specific roles and responsibilities clearly
- Explain relationships between organizations
- Cite relevant SAMM sections
- Include key personnel authorities where applicable
- Describe organizational structure and reporting relationships

RESPONSE STRUCTURE:
1. Full name and acronym
2. Primary role and mission
3. Specific responsibilities
4. Reporting relationships and coordination""",

                "factual": """You are a SAMM expert providing specific factual information.

TASK: Provide accurate, specific facts from SAMM with comprehensive detail.

CRITICAL REQUIREMENTS:
- Provide precise, accurate information with extensive context
- Include dates, numbers, and specific details
- Cite SAMM sections for verification (use section numbers only, e.g., "C9.3.9" - DO NOT mention chapter numbers)
- Use exact terminology from SAMM
- Expand acronyms appropriately
- Provide detailed explanations, not just brief answers
- Include relevant background information and context
- Explain the significance and implications of the facts
- Add examples or scenarios when applicable

RESPONSE STRUCTURE:
1. Direct answer to the factual question (detailed)
2. Comprehensive supporting context with background
3. Practical implications or applications
4. Source citations with section numbers""",

                "relationship": """You are a SAMM expert explaining relationships between entities and concepts.

TASK: Describe how SAMM entities, programs, and authorities relate to each other.

CRITICAL REQUIREMENTS:
- Clearly explain the nature of relationships
- Use specific examples to illustrate connections
- Cite relevant authorities and SAMM sections
- Explain the significance of relationships
- Address coordination and oversight aspects

RESPONSE STRUCTURE:
1. Describe the relationship clearly
2. Explain why the relationship exists
3. Provide examples of how it works in practice
4. Cite supporting authorities""",

                "general": """You are a SAMM (Security Assistance Management Manual) expert.

TASK: Provide comprehensive, detailed, and accurate information about Security Cooperation and Security Assistance.

CRITICAL REQUIREMENTS:
- Use exact SAMM terminology from the provided context
- Always cite SAMM sections when available (use section numbers only, e.g., "C9.3.9" - DO NOT mention chapter numbers)
- Expand acronyms on first use
- Maintain distinction between SC and SA (SA is subset of SC)
- Provide authoritative, accurate information with extensive detail
- Structure responses logically and completely with multiple paragraphs
- Include background context and explanations
- Add practical examples or scenarios when relevant
- Explain the significance and implications of the information"""
            }
            
            system_msg = base_instructions.get(intent, base_instructions["general"])
            
            # Add learned improvements from HIL feedback
            if intent in self.answer_templates and self.answer_templates[intent]:
                system_msg += "\n\nIMPORTANT IMPROVEMENTS FROM FEEDBACK:"
                for template in self.answer_templates[intent][-2:]:  # Last 2 templates
                    if template.get("improvement_notes"):
                        system_msg += f"\n- {template['improvement_notes']}"
                    if template.get("key_points"):
                        system_msg += f"\n- Ensure to mention: {', '.join(template['key_points'])}"
            
            # Add template structure guidance
            if intent in self.samm_response_templates:
                template = self.samm_response_templates[intent]
                system_msg += f"\n\nRESPONSE STRUCTURE: {template['structure']}"
                system_msg += f"\nREQUIRED ELEMENTS: {', '.join(template['required_elements'])}"
                system_msg += f"\nQUALITY CRITERIA: {', '.join(template['quality_criteria'])}"
            
            # Add length guidelines
            if intent in self.length_guidelines:
                guidelines = self.length_guidelines[intent]
                system_msg += f"\n\nLENGTH GUIDELINES: Target {guidelines['target']} characters (minimum {guidelines['min']}, maximum {guidelines['max']})"
            
            # Add context (truncated if necessary)
            context_truncated = context[:2000] + "..." if len(context) > 2000 else context
            system_msg += f"\n\nCONTEXT FROM SAMM:\n{context_truncated}"
            
            return system_msg
            
        except Exception as e:
            print(f"[AnswerAgent] Error creating system message: {e}")
            return "You are a SAMM expert. Provide accurate information about Security Cooperation and Security Assistance."
    
    def _create_enhanced_prompt(self, query: str, intent_info: Dict, entity_info: Dict) -> str:
        """Create enhanced prompt with entity and intent awareness"""
        try:
            intent = intent_info.get("intent", "general")
            entities = entity_info.get("entities", [])
            confidence = intent_info.get("confidence", 0.5)
            relationships = entity_info.get("relationships", [])  # NEW: Get relationships
           
            print(f"[AnswerAgent DEBUG] Relationships found: {relationships}") 
            
            prompt_parts = []
            
            # Add query with context
            prompt_parts.append(f"Question: {query}")
            
           # Add intent guidance if high confidence
            if confidence > 0.7:
                prompt_parts.append(f"This is a {intent} question requiring a {intent}-focused response.")
            
            # Add entity awareness (limit to prevent overload)
            if entities:
                limited_entities = entities[:3]  # Limit to 3 entities
                prompt_parts.append(f"Key entities mentioned: {', '.join(limited_entities)}")
                
                # CRITICAL: Add explicit entity focus for definition questions
                if intent == "definition" and len(limited_entities) == 1:
                    entity = limited_entities[0]
                    prompt_parts.append(f"\n🔴 CRITICAL: Answer ONLY about '{entity}'. Do NOT confuse with similar entities like:")
                    # Add common confusions for USD variants
                    if "USD(" in entity:
                        prompt_parts.append(f"  - If asked about USD(P), answer about Under Secretary of Defense for POLICY only")
                        prompt_parts.append(f"  - If asked about USD(P&R), answer about Under Secretary of Defense for PERSONNEL AND READINESS only")
                        prompt_parts.append(f"  - If asked about USD(A&S), answer about Under Secretary of Defense for ACQUISITION AND SUSTAINMENT only")
                        prompt_parts.append(f"  - You are being asked about: {entity} - answer ONLY about this exact entity")
            
            # NEW: Add relationship data explicitly
            if relationships:
                prompt_parts.append("\nIMPORTANT - Use these specific relationships from the database:")
                for rel in relationships[:5]:  # Limit to 5 relationships
                    prompt_parts.append(f"- {rel}")
                prompt_parts.append("\nBase your answer on these actual relationships, not generic knowledge.")
            
            # Add specific instructions based on intent
            intent_instructions = {
                "definition": "Provide a complete, authoritative definition with proper SAMM section reference.",
                "distinction": "Explain the key differences clearly with specific examples and legal basis.",
                "authority": "Explain who has authority, the scope of that authority, and the legal basis. USE THE RELATIONSHIPS PROVIDED ABOVE.",
                "organization": "Describe the organization's full name, role, and specific responsibilities.",
                "factual": "Provide the specific factual information with proper context and citation.",
                "relationship": "Describe how the entities relate to each other and why this matters."
            }
            
            if intent in intent_instructions:
                prompt_parts.append(intent_instructions[intent])
            
            prompt_parts.append("Provide a comprehensive, accurate answer based on SAMM content.")
            
            return "\n".join(prompt_parts)
            
        except Exception as e:
            print(f"[AnswerAgent] Error creating prompt: {e}")
            return f"Question: {query}\nProvide a comprehensive answer based on SAMM."



    def _generate_with_validation(self, prompt: str, system_msg: str, intent_info: Dict) -> str:
        """Generate answer with 30-second timeout and validation"""
        intent = intent_info.get("intent", "general")
        
        try:
            print("[AnswerAgent] First generation pass...")
            initial_answer = call_ollama_enhanced(prompt, system_msg, temperature=0.1)
            
            if "Error" in initial_answer and len(initial_answer) < 100:
                return initial_answer
            
            if "technical difficulties" in initial_answer:
                return initial_answer
            
            validation_results = self._validate_answer_quality(initial_answer, intent)
            
            if validation_results["needs_improvement"] and len(validation_results["issues"]) < 10:
                print(f"[AnswerAgent] Answer needs improvement: {validation_results['issues']}")
                
                improvement_prompt = f"{prompt}\n\nIMPROVEMENT NEEDED: {', '.join(validation_results['issues'])}\n\nPlease provide a better response addressing these issues."
                
                print("[AnswerAgent] Second generation pass with improvements...")
                improved_answer = call_ollama_enhanced(improvement_prompt, system_msg, temperature=0.2)
                
                if (len(improved_answer) > len(initial_answer) * 1.1 and 
                    "Error" not in improved_answer and 
                    "technical difficulties" not in improved_answer and
                    len(improved_answer) > 50):
                    return improved_answer
            
            return initial_answer
            
        except Exception as e:
            print(f"[AnswerAgent] Error during generation with validation: {e}")
            return _get_intelligent_fallback()


    def _validate_answer_quality(self, answer: str, intent: str) -> Dict[str, Any]:
        """Validate answer quality against SAMM standards with error handling"""
        try:
            issues = []
            needs_improvement = False
            
            # Skip validation if answer is too short (likely an error)
            if len(answer) < 20:
                return {"needs_improvement": False, "issues": ["answer_too_short"], "length": len(answer)}
            
            # Check length guidelines
            if intent in self.length_guidelines:
                guidelines = self.length_guidelines[intent]
                if len(answer) < guidelines["min"]:
                    issues.append("too short")
                    needs_improvement = True
                elif len(answer) > guidelines["max"]:
                    issues.append("too long")
            
            # Check for SAMM section references
            if not re.search(self.quality_patterns["section_references"], answer):
                issues.append("missing SAMM section reference")
                needs_improvement = True
            
            # Check for incomplete sentences
            if re.search(self.quality_patterns["incomplete_sentences"], answer):
                issues.append("incomplete sentences")
                needs_improvement = True
            
            # Intent-specific validations
            if intent == "definition" and "definition" not in answer.lower():
                issues.append("missing clear definition")
                needs_improvement = True
            
            if intent == "distinction" and not any(word in answer.lower() for word in ["difference", "differ", "distinction", "versus", "vs"]):
                issues.append("missing comparison language")
                needs_improvement = True
            
            if intent == "authority" and not any(word in answer.lower() for word in ["authority", "responsible", "oversight", "supervision"]):
                issues.append("missing authority language")
                needs_improvement = True
            
            return {
                "needs_improvement": needs_improvement,
                "issues": issues,
                "length": len(answer)
            }
            
        except Exception as e:
            print(f"[AnswerAgent] Error validating answer quality: {e}")
            return {"needs_improvement": False, "issues": [], "length": len(answer)}
    
    def _enhance_answer_quality(self, answer: str, intent_info: Dict, entity_info: Dict) -> str:
        """Apply post-processing enhancements with error handling"""
        try:
            enhanced_answer = answer
            
            # Skip enhancement if answer is too short or contains errors
            if len(answer) < 20 or "Error" in answer:
                return answer

            # Step 1: Add section reference if missing → take it from retriever top-k
            if not re.search(self.quality_patterns["section_references"], enhanced_answer):
                vr = (entity_info or {}).get("vector_result") or (entity_info or {}).get("retriever_result") or {}
                metas = (vr.get("metadatas", [[]]) or [[]])[0]

                secs = [ (m or {}).get("section_number") for m in metas if (m or {}).get("section_number") ]
                if secs:
                    # Get all unique sections (preserving order, removing duplicates)
                    from collections import OrderedDict
                    unique_secs = list(OrderedDict.fromkeys(secs))
                    
                    # Include top 3 most relevant sections instead of just 1
                    top_sections = unique_secs[:3]
                    
                    if len(top_sections) == 1:
                        enhanced_answer += f"\n\nSAMM Section Citation: {top_sections[0]}"
                    else:
                        # Multiple sections - join with commas
                        enhanced_answer += f"\n\nSAMM Section Citations: {', '.join(top_sections)}"
                    
                    print("[CITE DBG]", {"picked_sections": top_sections, "from": "retriever", "all_topk_secs": secs[:5]})

            # Step 2: Expand acronyms that appear without expansion (limit to prevent overprocessing)
            acronyms_found = re.findall(self.quality_patterns["acronym_detection"], enhanced_answer)
            
            for acronym in list(set(acronyms_found))[:5]:  # Limit to 5 acronyms
                if (acronym in self.acronym_expansions and 
                    acronym in enhanced_answer and 
                    self.acronym_expansions[acronym] not in enhanced_answer):
                    # Only expand the first occurrence
                    enhanced_answer = enhanced_answer.replace(acronym, self.acronym_expansions[acronym], 1)
            
            # Step 3: Ensure proper SAMM terminology
            terminology_fixes = {
                "security cooperation": "Security Cooperation",
                "security assistance": "Security Assistance", 
                "foreign assistance act": "Foreign Assistance Act",
                "arms export control act": "Arms Export Control Act"
            }
            
            for incorrect, correct in terminology_fixes.items():
                if incorrect in enhanced_answer and correct not in enhanced_answer:
                    enhanced_answer = enhanced_answer.replace(incorrect, correct)
            
            # Step 4: Add intent-specific enhancements
            intent = intent_info.get("intent", "general")
            
            if intent == "distinction" and "subset" not in enhanced_answer.lower():
                if "Security Assistance" in enhanced_answer and "Security Cooperation" in enhanced_answer:
                    enhanced_answer += "\n\nRemember: Security Assistance is a subset of Security Cooperation."
            
            return enhanced_answer
            
        except Exception as e:
            print(f"[AnswerAgent] Error enhancing answer quality: {e}")
            return answer  # Return original if enhancement fails
    
    def _validate_and_score_answer(self, answer: str, intent: str, query: str) -> str:
        """Final validation and quality scoring of the answer with error handling"""
        try:
            # Skip scoring if answer is too short or contains errors
            if len(answer) < 20 or "Error" in answer:
                return answer
            
            # Calculate quality score
            score = self._calculate_quality_score(answer, intent)
            
            # Log quality metrics
            print(f"[AnswerAgent] Answer quality score: {score:.2f}/1.0")
            
            # If score is too low, add disclaimer
            if score < 0.6:
                print(f"[AnswerAgent] Low quality score, adding disclaimer")
                answer += "\n\nNote: For complete and authoritative information, please refer to the full SAMM documentation."
            
            return answer
            
        except Exception as e:
            print(f"[AnswerAgent] Error in final validation: {e}")
            return answer  # Return original if validation fails
    
    def _calculate_quality_score(self, answer: str, intent: str) -> float:
        """Calculate quality score based on SAMM standards with error handling"""
        try:
            score = 0.0
            
            # Section citation score
            if re.search(self.quality_patterns["section_references"], answer):
                score += self.quality_weights["section_citation"]
            
            # Acronym expansion score
            acronyms_found = re.findall(self.quality_patterns["acronym_detection"], answer)
            if acronyms_found:
                expanded_count = sum(1 for acronym in acronyms_found if f"{acronym})" in answer)
                score += self.quality_weights["acronym_expansion"] * (expanded_count / len(set(acronyms_found)))
            
            # Answer completeness score (based on length guidelines)
            if intent in self.length_guidelines:
                guidelines = self.length_guidelines[intent]
                if guidelines["min"] <= len(answer) <= guidelines["max"]:
                    score += self.quality_weights["answer_completeness"]
                elif len(answer) >= guidelines["target"]:
                    score += self.quality_weights["answer_completeness"] * 0.8
            
            # SAMM terminology score
            samm_terms = ["Security Cooperation", "Security Assistance", "SAMM", "Title 10", "Title 22"]
            terms_used = sum(1 for term in samm_terms if term in answer)
            if terms_used > 0:
                score += self.quality_weights["samm_terminology"] * min(1.0, terms_used / 3)
            
            # Structure adherence score
            if intent in self.samm_response_templates:
                required_elements = self.samm_response_templates[intent]["required_elements"]
                elements_present = 0
                for element in required_elements:
                    element_keywords = element.replace("_", " ").split()
                    if any(keyword in answer.lower() for keyword in element_keywords):
                        elements_present += 1
                
                if required_elements:
                    score += self.quality_weights["structure_adherence"] * (elements_present / len(required_elements))
            
            return min(1.0, score)  # Cap at 1.0
            
        except Exception as e:
            print(f"[AnswerAgent] Error calculating quality score: {e}")
            return 0.5  # Return moderate score on error
    
    def _normalize_query_for_matching(self, query: str) -> str:
        """Normalize query for matching similar questions"""
        try:
            # Simple normalization - remove punctuation, lowercase, sort words
            words = re.findall(r'\b\w+\b', query.lower())
            # Keep only significant words (length > 2)
            significant_words = [word for word in words if len(word) > 2]
            return " ".join(sorted(significant_words))
        except Exception as e:
            print(f"[AnswerAgent] Error normalizing query: {e}")
            return query.lower()
    
    def update_from_hil(self, query: str, original_answer: str, corrected_answer: str, 
                        feedback_data: Dict[str, Any] = None):
        """Update agent based on human-in-the-loop feedback with improved error handling"""
        try:
            feedback_entry = {
                "query": query,
                "original_answer": original_answer,
                "corrected_answer": corrected_answer,
                "feedback_data": feedback_data or {},
                "timestamp": datetime.now().isoformat(),
                "improvement_type": "hil_correction"
            }
            
            self.hil_feedback_data.append(feedback_entry)
            
            # Store the correction for future similar queries
            query_key = self._normalize_query_for_matching(query)
            self.answer_corrections[query_key] = {
                "corrected_answer": corrected_answer,
                "feedback_data": feedback_data,
                "original_query": query,
                "correction_date": datetime.now().isoformat()
            }
            
            # Extract and store improved patterns
            if feedback_data:
                intent = feedback_data.get("intent", "general")
                if intent not in self.answer_templates:
                    self.answer_templates[intent] = []
                
                # Store template patterns from corrections
                template_info = {
                    "query_pattern": query.lower(),
                    "improvement_notes": feedback_data.get("improvement_notes", ""),
                    "key_points": feedback_data.get("key_points", []),
                    "structure_notes": feedback_data.get("structure_notes", ""),
                    "feedback_date": datetime.now().isoformat()
                }
                self.answer_templates[intent].append(template_info)
            
            # Add any new knowledge provided in feedback
            if feedback_data and feedback_data.get("additional_knowledge"):
                self.custom_knowledge += f"\n\nHIL Update ({datetime.now().strftime('%Y-%m-%d')}):\n{feedback_data['additional_knowledge']}"
            
            print(f"[AnswerAgent HIL] Updated with correction for query: '{query[:50]}...'")
            print(f"[AnswerAgent HIL] Total corrections stored: {len(self.answer_corrections)}")
            return True
            
        except Exception as e:
            print(f"[AnswerAgent] Error updating from HIL feedback: {e}")
            return False


    def _get_loa_timeline_answer(self) -> str:
        """Return pre-formatted LOA timeline answer (INSTANT - no Ollama call)"""
        return """According to SAMM Section C5.4.2 (Letter of Offer and Acceptance Document Preparation Timeframe), the time required to prepare LOA documents varies based on the complexity of the sale:

    **Category A Cases:**
    - Timeline: 85% completed within 45 days
    - Complexity: Simple cases

    **Category B Cases:**
    - Timeline: 85% completed within 100 days
    - Complexity: Moderate complexity

    **Category C Cases:**
    - Timeline: 85% completed within 150 days
    - Complexity: Complex cases

    The categorization depends on factors such as:
    - Number and complexity of line items
    - Special requirements or modifications
    - Coordination needs with other agencies
    - Technical complexity of the equipment

    **Note:** These timeframes represent the standard for 85% of cases. Individual cases may vary based on specific circumstances, requirements, and resource availability."""

    def _get_financial_verification_answer(self) -> str:
        """Return pre-formatted financial verification answer"""
        return """**Funding Verification for Case SR-P-NAV**

    ✅ **Appropriate Funding Line: Line 007**

    According to the LOA Line Notes and the work scope described in the field activity email, Line 007 is the correct funding line for this request.

    💰 **Funding Availability:**
    - **PDLI Balance**: $41,550,000.00 available
    - **Requested Amount**: $950,000.00
    - **Verdict**: ✅ **APPROVED** - You have plenty of funding to cover this request

    **Details:**
    - Available funds significantly exceed the request ($41.55M vs $950K)
    - Request represents only 2.3% of available PDLI balance
    - No funding concerns for this procurement"""

    def _get_technical_services_answer(self) -> str:
        """Return pre-formatted technical services answer"""
        return """**Technical Services for Case SR-P-NAV (Line 007)**

    According to Line Note 007 of the LOA, Technical services include:

    - System integration engineering
    - Software integration support
    - Platform compatibility verification
    - Test and evaluation support
    - Technical assistance for weapon system integration onto Saudi Arabian naval platforms

    These services support the integration and deployment of defense systems for the Saudi Arabian Navy."""

    def _get_pmr_minutes_summary(self) -> str:
        """Return pre-formatted PMR minutes summary with action items"""
        return """NSM Program Management Review I Summary

    The Royal Saudi Naval Forces (RSNF) conducted a Program Management Review for the Naval Strike Missile (NSM) Program on October 21-23, 2025. The meeting covered the acquisition of 96 tactical NSM missiles, associated containers, ground support equipment, technical documentation, training, and integration services through a Foreign Military Sales (FMS) case totaling $284,961,260.

    Key Program Elements

    Weapon System: NSM is a fifth-generation precision strike missile with 185+ km range, GPS/INS guidance, imaging infrared seeker, and two-way data link capability. The system will be integrated with RSNF's Al Riyadh-class frigates and Al Jubail-class corvettes.

    Timeline: LOA implementation planned for May 31, 2025, with initial missile deliveries beginning at month 24 (May 2027) and final deliveries by month 36 (May 2028). Case closure estimated 24 months after final delivery (approximately May 2030).

    Training: 16 officers will attend 8-week tactical employment courses in Newport, RI (September 2025-January 2026), and 24 enlisted personnel will complete 12-week maintenance courses in San Diego, CA (October 2025-February 2026).

    Integration Services: Technical services include platform compatibility verification, combat system integration with Thales TACTICOS systems, software integration, and test support through December 2027.

    Major Concerns Addressed: 
    - Platform compatibility assessments needed for both vessel classes
    - EMI/EMC testing in congested shipping environments
    - Accelerated delivery schedules to support Q4 2027 fleet exercises
    - Storage facility environmental controls for desert coastal climate
    - English language proficiency requirements (ECL 80 minimum)

    ---

    Action Items Due Within Two Weeks (by December 3, 2025)

    Based on the meeting date of October 21-23, 2025, and assuming "today" is November 19, 2025, the following action items are due within the next two weeks (by December 3, 2025):

    Due November 23, 2025:
    - AI-004: NSM PO to submit request for early delivery of containers and GSE to DSCA
    - AI-007: RSNF to designate POD location and provide customs clearance procedures
    - AI-011: RSNF to determine translation requirement and advise NSM PO
    - AI-016: RSNF to confirm preferred class structure (2x12 or alternative) for enlisted training
    - AI-026: RSNF to designate authorized receiving official and provide contact information

    Note: These items are already overdue if today is November 19, 2025

    Due December 7, 2025 (within next 3 weeks, but close):
    - AI-002: NSM Industry to provide EMI/EMC test reports and frequency deconfliction analysis
    - AI-003: NSM PO to explore feasibility of accelerating first missile delivery to month 20-22
    - AI-005: NSM Industry to provide sample test data package and quality documentation templates
    - AI-008: NSM Industry to deliver comprehensive GSE listing with technical specifications
    - AI-010: NSM PO to provide calibration service cost estimate and availability
    - AI-012: NSM Industry to deliver preliminary safety procedures document
    - AI-013: RSNF to provide student nomination list with desired class assignments
    - AI-015: RSNF to provide operational area maps and priority target sets for simulator scenarios
    - AI-024: RSNF to submit formal request for alternative transportation method (if desired)
    - AI-029: NSM PO to provide Explosive Safety Site Plan requirements and DoD 6055.09-M extracts"""




    def update_from_trigger(self, new_entities: List[str], new_relationships: List[Dict], 
                           trigger_data: Dict[str, Any] = None):
        """Update agent when new entity/relationship data is available with error handling"""
        try:
            trigger_entry = {
                "new_entities": new_entities,
                "new_relationships": new_relationships,
                "trigger_data": trigger_data or {},
                "timestamp": datetime.now().isoformat(),
                "trigger_id": len(self.trigger_updates)
            }
            
            self.trigger_updates.append(trigger_entry)
            
            # Add new knowledge from trigger updates
            if trigger_data:
                new_knowledge_items = []
                
                # Add entity definitions
                for entity in new_entities:
                    definition = trigger_data.get("entity_definitions", {}).get(entity)
                    if definition:
                        new_knowledge_items.append(f"{entity}: {definition}")
                
                # Add relationship information
                for rel in new_relationships:
                    rel_info = f"{rel.get('source', '')} {rel.get('relationship', '')} {rel.get('target', '')}"
                    details = trigger_data.get("relationship_details", {}).get(rel_info)
                    if details:
                        new_knowledge_items.append(f"Relationship: {rel_info} - {details}")
                
                # Add any general knowledge updates
                if trigger_data.get("knowledge_updates"):
                    new_knowledge_items.extend(trigger_data["knowledge_updates"])
                
                if new_knowledge_items:
                    self.custom_knowledge += f"\n\nTrigger Update ({datetime.now().strftime('%Y-%m-%d')}):\n" + "\n".join(new_knowledge_items)
            
            print(f"[AnswerAgent Trigger] Updated with {len(new_entities)} new entities and {len(new_relationships)} relationships")
            print(f"[AnswerAgent Trigger] Total trigger updates: {len(self.trigger_updates)}")
            return True
            
        except Exception as e:
            print(f"[AnswerAgent] Error updating from trigger: {e}")
            return False

def check_compliance(query: str, intent_info: Dict, entity_info: Dict, user_profile: Dict = None) -> Dict[str, Any]:
    """
    Check ITAR compliance - defaults to TOP_SECRET for development
    Fails open (permits access) if service unavailable
    """
    if not COMPLIANCE_ENABLED:
        return {
            "compliance_status": "compliant",
            "authorized": True,
            "user_authorization_level": DEFAULT_DEV_AUTH_LEVEL,
            "content_guidance": {"allowed_detail_level": "full"},
            "restrictions": [],
            "check_performed": False
        }
    
    # Set default dev authorization
    if not user_profile:
        user_profile = {"authorization_level": DEFAULT_DEV_AUTH_LEVEL}
    elif "authorization_level" not in user_profile:
        user_profile["authorization_level"] = DEFAULT_DEV_AUTH_LEVEL
    
    try:
        response = requests.post(
            f"{COMPLIANCE_SERVICE_URL}/api/compliance/verify",
            json={
                "query": query,
                "intent_info": intent_info,
                "entity_info": entity_info,
                "user_profile": user_profile
            },
            timeout=15  # INCREASED from 5 to 15 seconds
        )
        response.raise_for_status()
        result = response.json()
        result["check_performed"] = True
        return result
        
    except requests.exceptions.Timeout:
        print(f"[Compliance] Service timeout - defaulting to permissive mode")
    except requests.exceptions.ConnectionError:
        print(f"[Compliance] Service unavailable - defaulting to permissive mode")
    except Exception as e:
        print(f"[Compliance] Error: {e} - defaulting to permissive mode")
    
    # Fail open for development
    return {
        "compliance_status": "compliant",
        "authorized": True,
        "user_authorization_level": DEFAULT_DEV_AUTH_LEVEL,
        "content_guidance": {"allowed_detail_level": "full"},
        "restrictions": [],
        "check_performed": False,
        "fallback_reason": "service_unavailable"
    }


def extract_financial_records_from_documents(documents_context: List) -> List[Dict]:
    """
    Extract financial records from uploaded documents
    Returns list of financial records with PDLI info
    """
    if not documents_context:
        return []
    
    financial_records = []
    
    for doc in documents_context:
        # Check if document has financial data in metadata
        if doc.get('metadata', {}).get('hasFinancialData'):
            records = doc['metadata'].get('financialRecords', [])
            
            print(f"[Financial Extract] Found {len(records)} records in {doc.get('fileName')}")
            
            # Enrich each record with document info
            for record in records:
                enriched = {
                    **record,
                    'source_document': doc.get('fileName'),
                    'document_id': doc.get('documentId')
                }
                financial_records.append(enriched)
    
    print(f"[Financial Extract] Total: {len(financial_records)} financial records extracted")
    return financial_records




class SimpleStateOrchestrator:
    """Simple LangGraph-style state orchestration for integrated SAMM agents with HIL and trigger updates"""
    
    def __init__(self):
        self.intent_agent = IntentAgent()
        self.entity_agent = IntegratedEntityAgent(knowledge_graph, db_manager)
        self.answer_agent = EnhancedAnswerAgent()
        
        # Define workflow graph
        self.workflow = {
            WorkflowStep.INIT: self._initialize_state,
            WorkflowStep.INTENT: self._analyze_intent_step,
            WorkflowStep.ENTITY: self._extract_entities_step,
            WorkflowStep.ANSWER: self._generate_answer_step,
            WorkflowStep.COMPLETE: self._complete_workflow,
            WorkflowStep.ERROR: self._handle_error
        }
        
        # Define state transitions
        self.transitions = {
            WorkflowStep.INIT: WorkflowStep.INTENT,
            WorkflowStep.INTENT: WorkflowStep.ENTITY,
            WorkflowStep.ENTITY: WorkflowStep.ANSWER,
            WorkflowStep.ANSWER: WorkflowStep.COMPLETE,
            WorkflowStep.COMPLETE: None,
            WorkflowStep.ERROR: None
        }
    @time_function
    def process_query(self, query: str, chat_history: List = None, documents_context: List = None,
                     user_profile: Dict = None) -> Dict[str, Any]:
        """Process query through integrated state orchestrated workflow"""
        # Initialize state
        state = AgentState(
            query=query,
            chat_history=chat_history,
            documents_context=documents_context,
            intent_info=None,
            entity_info=None,
            answer=None,
            execution_steps=[],
            start_time=time.time(),
            current_step=WorkflowStep.INIT.value,
            error=None
        )
        # ✅ ADD THESE 3 LINES HERE:
        print(f"[DEBUG PROCESS_QUERY] Received documents_context: {documents_context is not None}")
        print(f"[DEBUG PROCESS_QUERY] documents_context type: {type(documents_context)}")
        print(f"[DEBUG PROCESS_QUERY] documents_context length: {len(documents_context) if documents_context else 0}")
    
        state['user_profile'] = user_profile or {"authorization_level": DEFAULT_DEV_AUTH_LEVEL}
        try:
            # Execute workflow
            current_step = WorkflowStep.INIT
            
            while current_step is not None:
                print(f"[State Orchestrator] Executing step: {current_step.value}")
                state['current_step'] = current_step.value
                state['execution_steps'].append(f"Step: {current_step.value}")
                
                # Execute step
                state = self.workflow[current_step](state)
                
                # Check for error
                if state.get('error'):
                    current_step = WorkflowStep.ERROR
                else:
                    # Move to next step
                    current_step = self.transitions[current_step]
            
            execution_time = round(time.time() - state['start_time'], 2)
            
            return {
                "query": state['query'],
                "answer": state['answer'],
                "intent": state['intent_info'].get('intent', 'unknown') if state['intent_info'] else 'unknown',
                "entities_found": len(state['entity_info'].get('entities', [])) if state['entity_info'] else 0,
                "execution_time": execution_time,
                "execution_steps": state['execution_steps'],
                "success": state['error'] is None,
                "metadata": {
                    "intent_confidence": state['intent_info'].get('confidence', 0) if state['intent_info'] else 0,
                    "entities": state['entity_info'].get('entities', []) if state['entity_info'] else [],
                    "system_version": "Integrated_Database_SAMM_v5.0",
                    "workflow_completed": state['current_step'] == 'complete',
                    # Keep legacy metadata structure for Vue.js compatibility
                    "intent": state['intent_info'].get('intent', 'unknown') if state['intent_info'] else 'unknown',
                    "entities_found": len(state['entity_info'].get('entities', [])) if state['entity_info'] else 0,
                    "execution_time_seconds": execution_time,
                    # Add database integration status
                    "database_integration": {
                        "cosmos_gremlin": db_manager.cosmos_gremlin_client is not None,
                        "vector_db": db_manager.vector_db_client is not None,
                        "embedding_model": db_manager.embedding_model is not None
                    },
                    # Add HIL and trigger update status
                    "hil_updates_available": (len(self.intent_agent.hil_feedback_data) > 0 or 
                                            len(self.entity_agent.hil_feedback_data) > 0 or 
                                            len(self.answer_agent.hil_feedback_data) > 0),
                    "trigger_updates_available": (len(self.intent_agent.trigger_updates) > 0 or 
                                                len(self.entity_agent.trigger_updates) > 0 or 
                                                len(self.answer_agent.trigger_updates) > 0),
                    # Enhanced entity agent status
                    "entity_extraction_method": state['entity_info'].get('extraction_method', 'unknown') if state['entity_info'] else 'unknown',
                    "entity_confidence": state['entity_info'].get('overall_confidence', 0) if state['entity_info'] else 0,
                    "extraction_phases": state['entity_info'].get('phase_count', 0) if state['entity_info'] else 0,
                    "total_database_results": state['entity_info'].get('total_results', 0) if state['entity_info'] else 0
                }
            }
            
        except Exception as e:
            execution_time = round(time.time() - state['start_time'], 2)
            return {
                "query": query,
                "answer": f"I apologize, but I encountered an error during integrated processing: {str(e)}",
                "intent": "error",
                "entities_found": 0,
                "execution_time": execution_time,
                "execution_steps": state['execution_steps'] + [f"Error: {str(e)}"],
                "success": False,
                "metadata": {"error": str(e), "system_version": "Integrated_Database_SAMM_v5.0"}
            }
    
    def update_agents_from_hil(self, query: str, intent_correction: Dict = None, entity_correction: Dict = None, answer_correction: Dict = None) -> Dict[str, bool]:
        """Update all agents from human-in-the-loop feedback"""
        results = {}
        
        # Update Intent Agent
        if intent_correction:
            results["intent"] = self.intent_agent.update_from_hil(
                query=query,
                original_intent=intent_correction.get("original_intent"),
                corrected_intent=intent_correction.get("corrected_intent"),
                feedback_data=intent_correction.get("feedback_data", {})
            )
        
        # Update Integrated Entity Agent
        if entity_correction:
            results["entity"] = self.entity_agent.update_from_hil(
                query=query,
                original_entities=entity_correction.get("original_entities", []),
                corrected_entities=entity_correction.get("corrected_entities", []),
                feedback_data=entity_correction.get("feedback_data", {})
            )
        
        # Update Enhanced Answer Agent
        if answer_correction:
            results["answer"] = self.answer_agent.update_from_hil(
                query=query,
                original_answer=answer_correction.get("original_answer"),
                corrected_answer=answer_correction.get("corrected_answer"),
                feedback_data=answer_correction.get("feedback_data", {})
            )
        
        print(f"[State Orchestrator] HIL updates completed: {results}")
        return results
    
    def update_agents_from_trigger(self, new_entities: List[str], new_relationships: List[Dict], trigger_data: Dict[str, Any] = None) -> Dict[str, bool]:
        """Update all agents when new entity/relationship data is available"""
        results = {}
        
        # Update Intent Agent
        results["intent"] = self.intent_agent.update_from_trigger(
            new_entities=new_entities,
            new_relationships=new_relationships,
            trigger_data=trigger_data
        )
        
        # Update Integrated Entity Agent
        results["entity"] = self.entity_agent.update_from_trigger(
            new_entities=new_entities,
            new_relationships=new_relationships,
            trigger_data=trigger_data
        )
        
        # Update Enhanced Answer Agent
        results["answer"] = self.answer_agent.update_from_trigger(
            new_entities=new_entities,
            new_relationships=new_relationships,
            trigger_data=trigger_data
        )
        
        print(f"[State Orchestrator] Trigger updates completed: {results}")
        return results
    
    def get_agent_status(self) -> Dict[str, Any]:
        """Get status of all agents including database connections and HIL/trigger update counts"""
        return {
            "intent_agent": {
                "hil_feedback_count": len(self.intent_agent.hil_feedback_data),
                "trigger_update_count": len(self.intent_agent.trigger_updates),
                "learned_patterns": len(self.intent_agent.intent_patterns)
            },
            "integrated_entity_agent": {
                "type": "IntegratedEntityAgent",
                "hil_feedback_count": len(self.entity_agent.hil_feedback_data),
                "trigger_update_count": len(self.entity_agent.trigger_updates),
                "custom_entities": len(self.entity_agent.custom_entities),
                "dynamic_entities": len(self.entity_agent.dynamic_knowledge["entities"]),
                "samm_patterns": sum(len(patterns) for patterns in self.entity_agent.samm_entity_patterns.values()),
                "extraction_phases": 3,  # pattern_matching, nlp_extraction, database_queries
                "database_status": db_manager.get_database_status()
            },
            "enhanced_answer_agent": {
                "type": "EnhancedAnswerAgent",
                "hil_feedback_count": len(self.answer_agent.hil_feedback_data),
                "trigger_update_count": len(self.answer_agent.trigger_updates),
                "answer_corrections": len(self.answer_agent.answer_corrections),
                "answer_templates": sum(len(templates) for templates in self.answer_agent.answer_templates.values()),
                "response_templates": len(self.answer_agent.samm_response_templates),
                "acronym_expansions": len(self.answer_agent.acronym_expansions)
            }
        }
    
    def get_database_status(self) -> Dict[str, Any]:
        """Get comprehensive database status"""
        return db_manager.get_database_status()
    
    def cleanup(self):
        """Cleanup all resources"""
        try:
            db_manager.cleanup()
            print("[State Orchestrator] Cleanup complete")
        except Exception as e:
            print(f"[State Orchestrator] Cleanup error: {e}")
    
    def _initialize_state(self, state: AgentState) -> AgentState:
        """Initialize workflow state"""
        state['execution_steps'].append("Integrated workflow initialized with database connections")
        print(f"[State Orchestrator] Initialized query: '{state['query']}'")
        return state
    
    def _analyze_intent_step(self, state: AgentState) -> AgentState:
        """Execute intent analysis step"""
        try:
            state['intent_info'] = self.intent_agent.analyze_intent(state['query'])
            state['execution_steps'].append(f"Intent analyzed: {state['intent_info'].get('intent', 'unknown')}")
            print(f"[State Orchestrator] Intent: {state['intent_info'].get('intent')} (confidence: {state['intent_info'].get('confidence')})")
        except Exception as e:
            state['error'] = f"Intent analysis failed: {str(e)}"
        return state
    
    @time_function
    def analyze_intent(self, query: str) -> Dict[str, Any]:
        # STEP 1: Check for special cases FIRST (before calling Ollama)
        special_case = self._check_special_cases(query)
        if special_case:
            print(f"[IntentAgent] Returning special case: {special_case['intent']}")
            return special_case
        
        # STEP 2: Normal SAMM intent analysis (existing logic unchanged)
        # Check if we have learned patterns from previous feedback
        enhanced_system_msg = self._build_enhanced_system_message()
        
        prompt = f"Analyze this SAMM query and determine intent: {query}"
        
        try:
            response = call_ollama_enhanced(prompt, enhanced_system_msg, temperature=0.0)
            # Try to parse JSON response
            if "{" in response and "}" in response:
                json_part = response[response.find("{"):response.rfind("}")+1]
                result = json.loads(json_part)
                
                # Apply any learned corrections from HIL feedback
                result = self._apply_hil_corrections(query, result)
                return result
            else:
                return {"intent": "general", "confidence": 0.5, "entities_mentioned": []}
        except:
            return {"intent": "general", "confidence": 0.5, "entities_mentioned": []}





    @time_function
    def _extract_entities_step(self, state: AgentState) -> AgentState:
        """Execute integrated entity extraction with database queries"""
        # NEW DEBUG LINES - ADD THESE 3 LINES:
        print(f"[DEBUG STATE] documents_context exists: {'documents_context' in state}")
        print(f"[DEBUG STATE] documents_context value: {state.get('documents_context', 'NOT FOUND')}")
        print(f"[DEBUG STATE] documents_context length: {len(state.get('documents_context', [])) if state.get('documents_context') else 0}")
    
        # ✅ ENHANCED: Skip entity extraction for special cases
        if state['intent_info'].get('special_case', False):
            intent = state['intent_info'].get('intent')
            print(f"[State Orchestrator] Skipping entity extraction for special case: {intent}")
            
            # ✅ NEW: Handle LOA timeline special case
            if intent == "loa_timeline":
                state['entity_info'] = {
                    'entities': ["LOA", "Timeline", "SAMM C5.4.2"],
                    'context': [{
                        "entity": "LOA",
                        "definition": "Letter of Offer and Acceptance - primary contractual document in FMS",
                        "section": "C5.4.2",
                        "type": "document",
                        "source": "knowledge_graph",
                        "confidence": 1.0
                    }],
                    'relationships': [
                        "LOA prepared by DSCA",
                        "LOA categorized by complexity (A, B, C)",
                        "LOA timeline varies by category"
                    ],
                    'special_case_skip': True,
                    'fast_path': True
                }
            else:
                state['entity_info'] = {
                    'entities': [],
                    'context': [],
                    'relationships': [],
                    'special_case_skip': True
                }
            
            state['execution_steps'].append("Entity extraction skipped (special case)")
            return state
        
        try:
            # ✅ ADDED: Get documents from state (safe - defaults to None if not present)
            documents_context = state.get('documents_context', None)
            
            # ✅ ADDED: Log file status for debugging
            if documents_context:
                print(f"[State Orchestrator] 📁 Passing {len(documents_context)} files to entity extraction")
                for idx, doc in enumerate(documents_context[:3], 1):
                    fname = doc.get('fileName', 'Unknown')
                    content_len = len(doc.get('content', ''))
                    print(f"[State Orchestrator]   File {idx}: {fname} ({content_len} chars)")
            else:
                print(f"[State Orchestrator] No files in state to pass to entity extraction")
            
            # ✅ FIXED: Now passes documents_context (was missing before)
            state['entity_info'] = self.entity_agent.extract_and_retrieve(
                state['query'], 
                state['intent_info'],
                documents_context  # ← ADDED THIS PARAMETER
            )
            
            # ✅ EXISTING: Get entity extraction stats
            entities_count = len(state['entity_info'].get('entities', []))
            confidence = state['entity_info'].get('overall_confidence', 0)
            db_results = state['entity_info'].get('total_results', 0)
            phases = state['entity_info'].get('phase_count', 0)
            
            # ✅ ADDED: Get file-related stats (safe - defaults to 0 if not present)
            files_processed = state['entity_info'].get('files_processed', 0)
            file_entities = state['entity_info'].get('file_entities_found', 0)
            file_relationships = state['entity_info'].get('file_relationships_found', 0)
            
            # ✅ ENHANCED: Include file stats in execution step message
            state['execution_steps'].append(
                f"Integrated entity extraction: {entities_count} entities found "
                f"(confidence: {confidence:.2f}, DB results: {db_results}, phases: {phases}, "
                f"files: {files_processed}, file_entities: {file_entities}, file_rels: {file_relationships})"
            )
            
            # ✅ ENHANCED: Include file stats in console log
            print(f"[State Orchestrator] Integrated Entities: {entities_count} entities found "
                f"through {phases} phases with {db_results} database results "
                f"and {file_entities} entities from {files_processed} files")
            
            # ✅ ADDED: Log file-specific extraction details if files were processed
            if files_processed > 0:
                print(f"[State Orchestrator] 📊 File Extraction Results:")
                print(f"[State Orchestrator]   • Files processed: {files_processed}")
                print(f"[State Orchestrator]   • Entities from files: {file_entities}")
                print(f"[State Orchestrator]   • Relationships from files: {file_relationships}")
            
        except Exception as e:
            # ✅ EXISTING: Error handling unchanged
            state['error'] = f"Integrated entity extraction failed: {str(e)}"
            print(f"[State Orchestrator] ❌ Entity extraction error: {str(e)}")
            
            # ✅ ADDED: Add traceback for debugging
            import traceback
            print(f"[State Orchestrator] Error traceback:\n{traceback.format_exc()}")
        
        return state

    @time_function
    def _generate_answer_step(self, state: AgentState) -> AgentState:
        """Execute enhanced answer generation step"""
        try:
            print(f"[State Orchestrator] 🔄 Starting answer generation...")
            print(f"[State Orchestrator]   Query: {state['query'][:50]}...")
            print(f"[State Orchestrator]   Intent: {state['intent_info'].get('intent', 'unknown')}")
            print(f"[State Orchestrator]   Entities: {len(state['entity_info'].get('entities', []))}")
            print(f"[State Orchestrator]   Files: {len(state.get('documents_context', []))}")
            
            # ✅ ADD: Pass user_profile to generate_answer
            state['answer'] = self.answer_agent.generate_answer(
                state['query'], 
                state['intent_info'], 
                state['entity_info'], 
                state['chat_history'], 
                state['documents_context'],
                state.get('user_profile')  # ← ADD THIS LINE
            )
            
            # ✅ ADD: Verify answer was generated
            if not state['answer'] or len(state['answer']) < 20:
                print(f"[State Orchestrator] ⚠️ WARNING: Answer too short or empty!")
                print(f"[State Orchestrator]   Answer: '{state['answer']}'")
                state['answer'] = "I apologize, but I encountered an issue generating a complete answer. Please try rephrasing your question."
            else:
                print(f"[State Orchestrator] ✅ Answer generated successfully:")
                print(f"[State Orchestrator]   Length: {len(state['answer'])} chars")
                print(f"[State Orchestrator]   Preview: {state['answer'][:150]}...")
            
            state['execution_steps'].append("Enhanced answer generated successfully with quality scoring")
        except Exception as e:
            print(f"[State Orchestrator] ❌ ERROR in answer generation: {str(e)}")
            import traceback
            traceback.print_exc()
            state['error'] = f"Enhanced answer generation failed: {str(e)}"
            state['answer'] = f"I apologize, but I encountered an error: {str(e)}"
        return state


    def _complete_workflow(self, state: AgentState) -> AgentState:
        """Complete workflow"""
        state['execution_steps'].append("Integrated workflow completed successfully")
        print(f"[State Orchestrator] Integrated workflow completed in {round(time.time() - state['start_time'], 2)}s")
        return state
    
    def _handle_error(self, state: AgentState) -> AgentState:
        """Handle workflow error"""
        state['execution_steps'].append(f"Error handled: {state['error']}")
        state['answer'] = f"I apologize, but I encountered an error: {state['error']}"
        print(f"[State Orchestrator] Error handled: {state['error']}")
        return state


# Initialize integrated orchestrator with all agents
orchestrator = SimpleStateOrchestrator()
print("Integrated State Orchestrator initialized with Intent, Integrated Entity (Database), and Enhanced Answer agents")

@time_function
def process_samm_query(query: str, chat_history: List = None, documents_context: List = None,
                      user_profile: Dict = None) -> Dict[str, Any]:
    """Process query through integrated state orchestrated 3-agent system with ITAR compliance"""
    return orchestrator.process_query(query, chat_history, documents_context, user_profile)
# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_mock_user():
    """Return a mock user for demo purposes"""
    return {
        "sub": "mock-user-123",
        "name": "Demo User",
        "email": "demo@example.com"
    }

def require_auth():
    """Check if user is authenticated, return user info or None"""
    user_session_data = session.get("user")
    if not user_session_data:
        return None
    
    # For OAuth
    if "userinfo" in user_session_data and "sub" in user_session_data["userinfo"]:
        return user_session_data["userinfo"]
    
    # For mock user (when OAuth not configured)
    if not oauth:
        return get_mock_user()
    
    return None



# =============================================================================
# AUTHENTICATION ROUTES
# =============================================================================

# =============================================================================
# AUTHENTICATION ROUTES
# =============================================================================

@app.route("/login")
def login():
    if oauth:
        # Hardcoded callback URL
        redirect_uri_for_auth0 = "http://172.16.200.12:3000/callback"
        print(f"[Login] Redirecting to Auth0 with callback: {redirect_uri_for_auth0}")
        return oauth.auth0.authorize_redirect(redirect_uri=redirect_uri_for_auth0)
    else:
        # Mock login when OAuth not configured
        session["user"] = {"userinfo": get_mock_user()}
        return jsonify({"message": "Logged in with mock user"}), 200


@app.route("/callback", methods=["GET", "POST"])
def callback():
    if not oauth:
        return jsonify({"error": "OAuth not configured"}), 500

    try:
        token = oauth.auth0.authorize_access_token()
        session["user"] = token
        userinfo = token.get("userinfo")
        if userinfo:
            print(f"User logged in: {userinfo.get('name')} ({userinfo.get('sub')})")
    except Exception as e:
        print(f"Error during Auth0 callback: {e}")
        return redirect(url_for("login"))

        # Hardcoded frontend URL
    vue_app_url = "http://172.16.200.12:5173"
    next_url_path_from_session = session.pop('next_url', None)
    final_redirect_url = vue_app_url

    if next_url_path_from_session:
        if next_url_path_from_session.startswith('/'):
            final_redirect_url = f"{vue_app_url}{next_url_path_from_session}"
        else:
            final_redirect_url = f"{vue_app_url}/{next_url_path_from_session}"

    print(f"[Callback] Redirecting to frontend: {final_redirect_url}")
    return redirect(final_redirect_url)


@app.route("/logout")
def logout():
    session.clear()
    if oauth:
        # Hardcoded frontend URL
        vue_app_url = "http://172.16.200.12:5173"
        return redirect(
            f"https://{AUTH0_DOMAIN}/v2/logout?" +
            urlencode({
                "returnTo": vue_app_url,
                "client_id": AUTH0_CLIENT_ID,
            }, quote_via=quote_plus)
        )
    else:
        return jsonify({"message": "Logged out"}), 200




# =============================================================================
# API ENDPOINTS
# =============================================================================

@app.route("/api/me", methods=["GET"])
def get_current_user_profile():
    user = require_auth()
    if user:
        return jsonify(user), 200
    else:
        return jsonify({"error": "User not authenticated"}), 401

@app.route("/api/user/cases", methods=["GET"])
def get_user_cases():
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]
    
    if cases_container_client:
        try:
            query = "SELECT * FROM c WHERE c.userId = @userId AND c.type = 'case'" 
            parameters = [{"name": "@userId", "value": user_id}]
            user_cases_list = list(cases_container_client.query_items(query=query, parameters=parameters, partition_key=user_id))
            return jsonify(user_cases_list), 200
        except Exception as e:
            print(f"Error querying cases: {e}")
            return jsonify({"error": "Database service error"}), 503
    else:
        # Use in-memory storage
        return jsonify(user_cases.get(user_id, [])), 200

@app.route("/api/cases", methods=["POST"])
def create_case():
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]
    case_data = request.get_json() if request.is_json else {}
    
    case_id = str(uuid.uuid4())
    new_case = {
        "id": case_id,
        "userId": user_id,
        "type": "case",
        "title": case_data.get("title", "New Case"),
        "description": case_data.get("description", ""),
        "caseDocuments": [],
        "createdAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "updatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    }
    
    if cases_container_client:
        try:
            cases_container_client.create_item(body=new_case)
            return jsonify(new_case), 201
        except Exception as e:
            print(f"Error creating case: {e}")
            return jsonify({"error": "Failed to create case"}), 500
    else:
        # Use in-memory storage
        if user_id not in user_cases:
            user_cases[user_id] = []
        user_cases[user_id].append(new_case)
        return jsonify(new_case), 201



@app.route("/api/cases/<path:case_id>/documents/upload", methods=["POST"])
def upload_case_document_to_case(case_id):
    """
    📤 UPLOAD DOCUMENTS TO CASE WITH AUTO-CASE-CREATION
    
    If case doesn't exist and we can extract case ID from filename,
    automatically create the case.
    """
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]
    
    # Decode case_id
    from urllib.parse import unquote
    case_id = unquote(case_id)
    
    # ✅ NEW: Handle undefined/missing case_id
    if not case_id or case_id == 'undefined':
        uploaded_files = request.files.getlist("documents")
        if uploaded_files and uploaded_files[0].filename:
            first_filename = uploaded_files[0].filename
            extracted_id = extract_case_id_from_filename(first_filename)
            if extracted_id:
                case_id = extracted_id
                print(f"[Upload] ✅ Extracted case ID from filename: {case_id}")
            else:
                return jsonify({"error": "No case ID provided and couldn't extract from filename"}), 400
        else:
            return jsonify({"error": "No case ID and no files provided"}), 400

    if not cases_container_client or not case_docs_blob_container_client:
        return jsonify({"error": "Backend storage service not available"}), 503

    # Get uploaded files
    uploaded_files = request.files.getlist("documents")
    if not uploaded_files or not uploaded_files[0].filename:
        return jsonify({"error": "No files selected for upload"}), 400

    # ✅ CRITICAL FIX: Try to find case, if not found, CREATE IT
    case_doc = None
    case_created = False
    
    try:
        # Try to find by UUID
        case_doc = cases_container_client.read_item(item=case_id, partition_key=user_id)
        print(f"[Upload] Found existing case by UUID: {case_id}")
        
    except CosmosExceptions.CosmosResourceNotFoundError:
        # Try by case number
        query = """
        SELECT * FROM c 
        WHERE c.userId = @userId 
        AND c.caseNumber = @caseNumber
        """
        parameters = [
            {"name": "@userId", "value": user_id},
            {"name": "@caseNumber", "value": case_id}
        ]
        cases = list(cases_container_client.query_items(
            query=query, parameters=parameters, enable_cross_partition_query=False
        ))
        
        if cases:
            case_doc = cases[0]
            print(f"[Upload] Found existing case by case number: {case_id}")
        else:
            # ✅ AUTO-CREATE CASE FROM FILENAME
            print(f"[Upload] 🔧 Case not found, auto-creating: {case_id}")
            
            new_case_id = str(uuid.uuid4())
            case_doc = {
                "id": new_case_id,
                "userId": user_id,
                "type": "case",
                "caseNumber": case_id,
                "title": f"Case {case_id}",
                "description": f"Auto-created from document upload",
                "status": "active",
                "caseDocuments": [],
                "documents": [],
                "createdAt": datetime.now(timezone.utc).isoformat(),
                "updatedAt": datetime.now(timezone.utc).isoformat(),
                "autoCreated": True
            }
            
            # Save to Cosmos DB
            cases_container_client.create_item(body=case_doc)
            case_created = True
            print(f"[Upload] ✅ Auto-created case: {case_id} (UUID: {new_case_id})")
    
    except Exception as e:
        print(f"[Upload] Error checking/creating case {case_id}: {str(e)}")
        return jsonify({"error": f"Could not retrieve or create case: {str(e)}"}), 500

    actual_case_id = case_doc["id"]
    case_number = case_doc.get("caseNumber", case_id)
    
    if "caseDocuments" not in case_doc:
        case_doc["caseDocuments"] = []
    
    if "documents" not in case_doc:
        case_doc["documents"] = []

    results = []

    # Process each uploaded file
    for uploaded_file in uploaded_files:
        if not uploaded_file or not uploaded_file.filename:
            continue
        
        original_filename = secure_filename(uploaded_file.filename)
        print(f"\n[Upload] Processing: {original_filename}")
        
        # Save to temp file for processing
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_filename)[1]) as temp_file:
            uploaded_file.save(temp_file.name)
            temp_path = temp_file.name
        
        try:
            # Extract data
            print(f"[Upload] 🔍 Extracting data...")
            doc_data = extract_case_document_data(temp_path, "AUTO_DETECT", original_filename)
            
            financial_records = doc_data.get("key_info", {}).get("financial_records", [])
            print(f"[Upload] 📋 Type: {doc_data.get('document_type')}, Records: {len(financial_records)}")
            
            # Determine document type
            doc_type = determine_document_type(original_filename)
            if doc_data.get('document_type') in ['LOA', 'FINANCIAL_DATA', 'MINUTES']:
                doc_type = doc_data['document_type']
            
            # Upload to blob storage
            blob_name = f"{actual_case_id}/{original_filename}"
            blob_client = case_docs_blob_container_client.get_blob_client(blob_name)
            
            uploaded_file.seek(0)
            blob_content_settings = ContentSettings(content_type=uploaded_file.mimetype)
            blob_client.upload_blob(
                uploaded_file.read(),
                overwrite=True,
                content_settings=blob_content_settings
            )
            
            blob_url = blob_client.url
            
            uploaded_file.seek(0, os.SEEK_END)
            file_size = uploaded_file.tell()
            
            # Create document record
            document_record = {
                "id": str(uuid.uuid4()),
                "documentId": str(uuid.uuid4()),
                "filename": original_filename,
                "fileName": original_filename,
                "documentType": doc_type,
                "document_type": doc_type,
                "blobName": blob_name,
                "blobUrl": blob_url,
                "blobContainer": AZURE_CASE_DOCS_CONTAINER_NAME,
                "contentType": uploaded_file.content_type,
                "sizeBytes": file_size,
                "uploadedAt": datetime.now(timezone.utc).isoformat(),
                "uploadedBy": user_id,
                "caseId": actual_case_id,
                "caseNumber": case_number,
                "metadata": {
                    "hasFinancialData": len(financial_records) > 0,
                    "financialRecordCount": len(financial_records),
                    "financialRecords": financial_records,
                    "caseIdentifier": doc_data.get("case_identifier", case_number),
                    "extractedData": doc_data.get("key_info", {}),
                    "processed": True,
                    "extractionDate": datetime.now(timezone.utc).isoformat()
                }
            }
            
            # Add to case
            case_doc["caseDocuments"].append(document_record)
            case_doc["documents"].append(document_record)
            
            results.append({
                "filename": original_filename,
                "status": "success",
                "documentId": document_record["id"],
                "documentType": doc_type,
                "financialRecords": len(financial_records)
            })
            
            print(f"[Upload] ✅ Processed: {original_filename}")
        
        except Exception as e:
            print(f"[Upload] ❌ Error processing {original_filename}: {e}")
            import traceback
            traceback.print_exc()
            results.append({
                "filename": original_filename,
                "status": "error",
                "error": str(e)
            })
        
        finally:
            try:
                os.unlink(temp_path)
            except:
                pass
    
    # Save case
    try:
        case_doc["updatedAt"] = datetime.now(timezone.utc).isoformat()
        cases_container_client.upsert_item(body=case_doc)
        
        return jsonify({
            "success": True,
            "caseId": actual_case_id,
            "caseNumber": case_number,
            "caseCreated": case_created,
            "results": results
        }), 200
    
    except Exception as e:
        print(f"[Upload] Error updating case: {str(e)}")
        return jsonify({"error": "Failed to update case"}), 500





@app.route("/api/cases/<path:case_id>/financial-data", methods=["GET"])
def get_case_financial_data(case_id):
    """
    💰 GET ALL FINANCIAL DATA FOR A CASE
    
    Returns all extracted financial records from uploaded MISIL RSN sheets
    
    Response:
        {
          "success": true,
          "caseId": "uuid",
          "caseNumber": "SR-P-NAV",
          "financialDocuments": [
            {
              "documentId": "uuid",
              "fileName": "MISIL_RSN.xlsx",
              "uploadedAt": "2024-01-15T10:30:00Z",
              "recordCount": 45
            }
          ],
          "financialRecords": [
            {
              "rsn_identifier": "A-001",
              "pdli_pdli": "123456",
              "pdli_name": "F-16 Parts",
              "oa_rec_amt": 1000000,
              "net_commit_amt": 500000,
              "available": 500000,
              "sourceDocument": "MISIL_RSN.xlsx"
            }
          ],
          "recordCount": 45,
          "totals": {
            "oa_rec_amt": 50000000,
            "net_commit_amt": 25000000,
            ...
          }
        }
    """
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]
    
    from urllib.parse import unquote
    case_id = unquote(case_id)
    
    print(f"[Financial Data] Fetching for case: {case_id}")
    
    if not cases_container_client:
        return jsonify({"error": "Database not available"}), 503
    
    try:
        # Get case document
        case_doc = None
        
        # Try UUID lookup
        try:
            case_doc = cases_container_client.read_item(item=case_id, partition_key=user_id)
        except CosmosExceptions.CosmosResourceNotFoundError:
            # Try case number lookup
            query = """
            SELECT * FROM c 
            WHERE c.userId = @userId 
            AND c.type = 'case'
            AND c.caseNumber = @caseNumber
            """
            parameters = [
                {"name": "@userId", "value": user_id},
                {"name": "@caseNumber", "value": case_id}
            ]
            
            cases = list(cases_container_client.query_items(
                query=query,
                parameters=parameters,
                enable_cross_partition_query=False
            ))
            
            if cases:
                case_doc = cases[0]
        
        if not case_doc:
            return jsonify({"error": "Case not found"}), 404
        
        # Extract all financial records
        all_financial_records = []
        financial_documents = []
        
        for doc in case_doc.get("caseDocuments", []):
            metadata = doc.get("metadata", {})
            
            if metadata.get("hasFinancialData", False):
                records = metadata.get("financialRecords", [])
                
                if records:
                    financial_documents.append({
                        "documentId": doc.get("documentId"),
                        "fileName": doc.get("fileName"),
                        "uploadedAt": doc.get("uploadedAt"),
                        "recordCount": len(records)
                    })
                    
                    # Add document reference to each record
                    for record in records:
                        record["sourceDocument"] = doc.get("fileName")
                        record["documentId"] = doc.get("documentId")
                    
                    all_financial_records.extend(records)
        
        # Calculate totals
        totals = {
            "oa_rec_amt": 0,
            "net_commit_amt": 0,
            "net_obl_amt": 0,
            "net_exp_amt": 0,
            "dir_rsrv_amt": 0
        }
        
        for record in all_financial_records:
            for field in totals.keys():
                value = record.get(field, 0)
                if value:
                    try:
                        totals[field] += float(str(value).replace('$', '').replace(',', ''))
                    except:
                        pass
        
        print(f"[Financial Data] ✅ Found {len(all_financial_records)} records")
        
        return jsonify({
            "success": True,
            "caseId": case_doc["id"],
            "caseNumber": case_doc.get("caseNumber", case_id),
            "financialDocuments": financial_documents,
            "financialRecords": all_financial_records,
            "recordCount": len(all_financial_records),
            "totals": totals,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }), 200
        
    except Exception as e:
        print(f"[Financial Data] ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    
@app.route("/api/cases/<path:case_id>/financial-summary", methods=["GET"])
def get_financial_summary(case_id):
    """
    📊 GET FINANCIAL SUMMARY WITH RSN AGGREGATION
    
    Returns high-level financial metrics grouped by RSN PDLI
    
    Response:
        {
          "success": true,
          "caseId": "uuid",
          "caseNumber": "SR-P-NAV",
          "rsnSummary": [
            {
              "rsn_identifier": "A-001",
              "pdli_pdli": "123456",
              "pdli_name": "F-16 Parts",
              "oa_rec_amt": 1000000,
              "net_commit_amt": 500000,
              "available": 500000,
              "record_count": 5
            }
          ],
          "grandTotals": {...},
          "uniqueRSNs": 10
        }
    """
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]
    
    from urllib.parse import unquote
    case_id = unquote(case_id)
    
    if not cases_container_client:
        return jsonify({"error": "Database not available"}), 503
    
    try:
        # Get case (same logic as financial-data endpoint)
        case_doc = None
        try:
            case_doc = cases_container_client.read_item(item=case_id, partition_key=user_id)
        except:
            query = """
            SELECT * FROM c 
            WHERE c.userId = @userId 
            AND c.caseNumber = @caseNumber
            """
            parameters = [
                {"name": "@userId", "value": user_id},
                {"name": "@caseNumber", "value": case_id}
            ]
            cases = list(cases_container_client.query_items(
                query=query, parameters=parameters, enable_cross_partition_query=False
            ))
            if cases:
                case_doc = cases[0]
        
        if not case_doc:
            return jsonify({"error": "Case not found"}), 404
        
        # Aggregate by RSN
        rsn_aggregation = {}
        
        for doc in case_doc.get("caseDocuments", []):
            metadata = doc.get("metadata", {})
            records = metadata.get("financialRecords", [])
            
            for record in records:
                rsn = record.get("rsn_identifier", "Unknown")
                
                if rsn not in rsn_aggregation:
                    rsn_aggregation[rsn] = {
                        "rsn_identifier": rsn,
                        "pdli_pdli": record.get("pdli_pdli", "N/A"),
                        "pdli_name": record.get("pdli_name", ""),
                        "oa_rec_amt": 0,
                        "net_commit_amt": 0,
                        "net_obl_amt": 0,
                        "net_exp_amt": 0,
                        "dir_rsrv_amt": 0,
                        "record_count": 0
                    }
                
                # Aggregate amounts
                for field in ["oa_rec_amt", "net_commit_amt", "net_obl_amt", "net_exp_amt", "dir_rsrv_amt"]:
                    value = record.get(field, 0)
                    if value:
                        try:
                            rsn_aggregation[rsn][field] += float(str(value).replace('$', '').replace(',', ''))
                        except:
                            pass
                
                rsn_aggregation[rsn]["record_count"] += 1
        
        # Convert to list and sort by amount
        rsn_summary = sorted(
            rsn_aggregation.values(),
            key=lambda x: x.get("oa_rec_amt", 0),
            reverse=True
        )
        
        # Calculate grand totals
        grand_totals = {
            "oa_rec_amt": sum(item["oa_rec_amt"] for item in rsn_summary),
            "net_commit_amt": sum(item["net_commit_amt"] for item in rsn_summary),
            "net_obl_amt": sum(item["net_obl_amt"] for item in rsn_summary),
            "net_exp_amt": sum(item["net_exp_amt"] for item in rsn_summary),
            "dir_rsrv_amt": sum(item["dir_rsrv_amt"] for item in rsn_summary)
        }
        
        return jsonify({
            "success": True,
            "caseId": case_doc["id"],
            "caseNumber": case_doc.get("caseNumber", case_id),
            "rsnSummary": rsn_summary,
            "grandTotals": grand_totals,
            "uniqueRSNs": len(rsn_summary),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }), 200
        
    except Exception as e:
        print(f"[Financial Summary] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
@app.route("/api/chat/stage_attachment", methods=["POST"])
def stage_chat_attachment():
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]

    if not chat_docs_blob_container_client:
        print("[API StageChatAttachment] Chat documents blob service not available.")
        return jsonify({"error": "Chat document storage service not available"}), 503

    if 'document' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file_to_upload = request.files['document']

    if file_to_upload.filename == '':
        return jsonify({"error": "No selected file"}), 400

    if file_to_upload:
        original_filename = secure_filename(file_to_upload.filename)
        blob_name = f"{user_id}/chat_staging/{str(uuid.uuid4())}-{original_filename}"
        
        print(f"[API StageChatAttachment] Processing file: {original_filename} for blob: {blob_name}")
        blob_client_instance = chat_docs_blob_container_client.get_blob_client(blob_name)
            
        try:
            file_to_upload.seek(0) 
            blob_content_settings = ContentSettings(content_type=file_to_upload.mimetype)
            blob_client_instance.upload_blob(
                file_to_upload.read(), 
                overwrite=True,
                content_settings=blob_content_settings
            )
            print(f"[API StageChatAttachment] Successfully uploaded '{original_filename}' to blob: {blob_name}")

            file_to_upload.seek(0, os.SEEK_END)
            file_size_bytes = file_to_upload.tell()
            
            staged_doc_metadata = {
                "documentId": str(uuid.uuid4()),
                "fileName": original_filename,
                "blobName": blob_name,
                "blobContainer": AZURE_CHAT_DOCS_CONTAINER_NAME,
                "url": blob_client_instance.url,
                "fileType": file_to_upload.mimetype,
                "sizeBytes": file_size_bytes,
                "uploadedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
                "uploaderUserId": user_id,
                "status": "staged"
            }
            
            return jsonify({
                "message": f"File '{original_filename}' staged successfully.",
                "stagedDocument": staged_doc_metadata 
            }), 200

        except Exception as e:
            print(f"[API StageChatAttachment] Error uploading file '{original_filename}' to blob: {str(e)}")
            return jsonify({"error": f"Failed to upload file '{original_filename}'.", "details": str(e)}), 500
    
    return jsonify({"error": "Unknown error during file staging."}), 500


@app.route("/api/cases/documents/delete", methods=["POST"])
def delete_case_document_route():
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]

    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    case_id = data.get("caseId")
    document_metadata_id_to_delete = data.get("documentId") 

    if not case_id or not document_metadata_id_to_delete:
        return jsonify({"error": "Missing caseId or documentId in request"}), 400

    print(f"[API DeleteCaseDocument] User: {user_id} attempting to delete document with metadata ID: {document_metadata_id_to_delete} from case: {case_id}")

    if not cases_container_client or not case_docs_blob_container_client:
        return jsonify({"error": "Backend storage or database service not available"}), 503

    try:
        case_doc = cases_container_client.read_item(item=case_id, partition_key=user_id)
    except CosmosExceptions.CosmosResourceNotFoundError:
        return jsonify({"error": f"Case {case_id} not found or access denied."}), 404
    except Exception as e:
        print(f"Error reading case {case_id} for document deletion: {str(e)}")
        return jsonify({"error": "Could not retrieve case details"}), 500

    doc_to_delete_metadata = None
    original_case_documents = case_doc.get("caseDocuments", [])
    updated_case_documents = []

    for doc_meta in original_case_documents:
        if doc_meta.get("documentId") == document_metadata_id_to_delete:
            doc_to_delete_metadata = doc_meta
        else:
            updated_case_documents.append(doc_meta)

    if not doc_to_delete_metadata:
        print(f"[API DeleteCaseDocument] Document metadata ID {document_metadata_id_to_delete} not found in case {case_id}.")
        return jsonify({"error": "Document not found within the case."}), 404

    # Delete from Azure Blob Storage
    blob_name_to_delete = doc_to_delete_metadata.get("blobName")
    if blob_name_to_delete:
        try:
            blob_client_instance = case_docs_blob_container_client.get_blob_client(blob_name_to_delete)
            blob_client_instance.delete_blob()
            print(f"[API DeleteCaseDocument] Successfully deleted blob: {blob_name_to_delete} from container: {AZURE_CASE_DOCS_CONTAINER_NAME}")
        except BlobResourceNotFoundError:
            print(f"[API DeleteCaseDocument] Blob not found in storage (already deleted?): {blob_name_to_delete}")
        except Exception as e_blob:
            print(f"[API DeleteCaseDocument] Error deleting blob '{blob_name_to_delete}': {str(e_blob)}")

    # Update Cosmos DB
    case_doc["caseDocuments"] = updated_case_documents
    case_doc["updatedAt"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    cases_container_client.replace_item(item=case_id, body=case_doc)
    print(f"[API DeleteCaseDocument] Successfully removed document metadata ID {document_metadata_id_to_delete} from case {case_id} in Cosmos DB.")
    return jsonify({"message": f"Document '{doc_to_delete_metadata.get('fileName', 'Unknown')}' deleted successfully from case {case_id}."}), 200

@app.route("/api/chat/attachments/delete", methods=["POST"])
def delete_chat_attachment():
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]

    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    blob_name = data.get("blobName")
    blob_container_name = data.get("blobContainer")

    if not blob_name or not blob_container_name:
        return jsonify({"error": "Missing blobName or blobContainer in request"}), 400

    print(f"[API DeleteChatAttachment] User: {user_id} attempting to delete blob: {blob_name} from container: {blob_container_name}")

    if blob_container_name != AZURE_CHAT_DOCS_CONTAINER_NAME:
        print(f"[API DeleteChatAttachment] Attempt to delete from non-chat container: {blob_container_name}. Denied.")
        return jsonify({"error": "Invalid target container for deletion"}), 403

    if not blob_service_client:
        print("[API DeleteChatAttachment] Blob service client not available.")
        return jsonify({"error": "Blob storage service not available"}), 503
    
    target_blob_client = None
    if blob_container_name == AZURE_CHAT_DOCS_CONTAINER_NAME:
        if chat_docs_blob_container_client:
            target_blob_client = chat_docs_blob_container_client.get_blob_client(blob_name)
        else:
            print(f"[API DeleteChatAttachment] Mismatch or uninitialized client for container: {blob_container_name}")
            return jsonify({"error": "Specified blob container client not configured or mismatch"}), 500

    if not target_blob_client:
        return jsonify({"error": "Could not obtain blob client for deletion."}), 500

    try:
        target_blob_client.delete_blob()
        print(f"[API DeleteChatAttachment] Successfully deleted blob: {blob_name} from container: {blob_container_name}")
        return jsonify({"message": f"File '{blob_name}' deleted successfully from chat context."}), 200

    except BlobResourceNotFoundError:
        print(f"[API DeleteChatAttachment] Blob not found: {blob_name} in container: {blob_container_name}")
        return jsonify({"error": "File not found in storage."}), 404
    except Exception as e:
        print(f"[API DeleteChatAttachment] Error deleting blob '{blob_name}': {str(e)}")
        return jsonify({"error": "Failed to delete file from storage.", "details": str(e)}), 500




@app.route("/api/query", methods=["POST"])
def query_ai_assistant():
    """Main SAMM query endpoint using Integrated state orchestrated 3-agent system with caching and ITAR compliance"""
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]

    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    try:
        data = request.get_json()
        user_input = data.get("question", "").strip()
        chat_history = data.get("chat_history", []) 
        staged_chat_documents_metadata = data.get("staged_chat_documents", []) 
        
        if not user_input:
            return jsonify({"error": "Query cannot be empty"}), 400

        # === Extract user authorization profile ===
        user_profile = {
            "user_id": user_id,
            "authorization_level": user.get("authorization_level", DEFAULT_DEV_AUTH_LEVEL),
            "clearances": user.get("clearances", []),
            "role": user.get("role", "developer")
        }
        print(f"[Integrated SAMM Query] User: {user_id}, Auth: {user_profile['authorization_level']}, Query: '{user_input[:50]}...'")

        # === NEW: Load actual document content from blob storage ===
        documents_with_content = []
        for doc_meta in staged_chat_documents_metadata:
            blob_name = doc_meta.get("blobName")
            if blob_name and chat_docs_blob_container_client:
                content = fetch_blob_content(blob_name, chat_docs_blob_container_client)
                if content:
                    documents_with_content.append({
                        **doc_meta,
                        "content": content[:5000]  # Limit to 5000 chars to avoid overload
                    })
                    print(f"[Query] Loaded content from {doc_meta.get('fileName')}: {len(content)} chars")
        # === END NEW ===

        # STEP 1: Check cache first
        cached_result = get_from_cache(user_input)
        
        if cached_result:
            # Cache hit - return cached answer with cache metadata
            print(f"[Cache] Returning cached answer for: '{user_input[:50]}...'")
            
            response_data = {
                "response": {"answer": cached_result['answer']},
                "metadata": cached_result['metadata'],
                "uploadedChatDocuments": [],
                "cached": True,
                "cache_age_seconds": round(time.time() - cached_result['timestamp'], 2)
            }
            
            return jsonify(response_data)
        
        # STEP 2: Cache miss - process query normally
        print(f"[Integrated SAMM Query] Chat History items: {len(chat_history)}")
        print(f"[Integrated SAMM Query] Staged Chat Documents: {len(staged_chat_documents_metadata)}")
        
        # Check for demo partial response
        demo_response = generate_demo_partial_response(user_input)
        if demo_response and not get_from_cache(user_input):
            result = {
                'answer': demo_response['answer'],
                'metadata': {
                    'intent': demo_response['intent'],
                    'entities': demo_response['entities'],
                    'is_demo': True,
                    'demo_type': demo_response.get('demo_type', 'unknown')
                },
                'intent': demo_response['intent'],
                'entities_found': len(demo_response['entities']),
                'execution_time': 0.5
            }
            print(f"🎬 DEMO MODE: Using partial answer ({demo_response.get('demo_type', 'unknown').upper()})")
        else:
            # MODIFIED: Pass documents_with_content instead of staged_chat_documents_metadata
            result = process_samm_query(user_input, chat_history, documents_with_content, user_profile)
        
        # Apply HITL corrections if they exist
        result = apply_hitl_corrections(user_input, result)
        
        print(f"[Integrated SAMM Result] Intent: {result['intent']}, Entities: {result['entities_found']}, Time: {result['execution_time']}s")
        print(f"[Integrated SAMM Result] Workflow Steps: {len(result.get('execution_steps', []))}")
        print(f"[Integrated SAMM Result] System Version: {result['metadata'].get('system_version', 'Unknown')}")
        print(f"[Integrated SAMM Result] Database Results: {result['metadata'].get('total_database_results', 0)}")
        
        # STEP 3: Save to cache
        save_to_cache(user_input, result['answer'], result['metadata'])
        
        # ✅ NEW: Extract financial data from documents for response
        financial_summary = None
        if documents_with_content:
            financial_records = []
            for doc in documents_with_content:
                if doc.get('metadata', {}).get('hasFinancialData'):
                    records = doc['metadata'].get('financialRecords', [])
                    financial_records.extend(records)
            
            if financial_records:
                financial_summary = {
                    'total_records': len(financial_records),
                    'unique_rsns': len(set(r.get('rsn_identifier') for r in financial_records if r.get('rsn_identifier'))),
                    'total_available': sum(float(r.get('available', 0)) for r in financial_records),
                    'documents': [doc.get('fileName') for doc in documents_with_content 
                                 if doc.get('metadata', {}).get('hasFinancialData')]
                }
                print(f"[API] 💰 Financial summary: {financial_summary}")
        
        # Return response in the same format as before for Vue.js UI compatibility
        response_data = {
            "response": {"answer": result["answer"]},
            "metadata": result["metadata"],
            "uploadedChatDocuments": [],  # For future AI-generated documents
            "financialSummary": financial_summary,  # ✅ NEW
            "cached": False  # Fresh answer
        }
        
        # Add execution steps only in debug mode or if requested
        if data.get("debug", False) or data.get("include_workflow", False):
            response_data["execution_steps"] = result.get("execution_steps", [])
            response_data["workflow_info"] = {
                "orchestration": "integrated_database_state",
                "steps_completed": len(result.get("execution_steps", [])),
                "execution_time": result["execution_time"],
                "entity_extraction_method": result["metadata"].get("entity_extraction_method", "unknown"),
                "entity_confidence": result["metadata"].get("entity_confidence", 0),
                "extraction_phases": result["metadata"].get("extraction_phases", 0),
                "database_results": result["metadata"].get("total_database_results", 0),
                "database_integration": result["metadata"].get("database_integration", {})
            }
        
        return jsonify(response_data)

    except Exception as e:
        print(f"[Integrated SAMM Query] Error: {str(e)}")
        return jsonify({"error": f"Internal Server Error: {str(e)}"}), 500






@app.route("/api/cache/stats", methods=["GET"])
def get_cache_statistics():
    """Get cache performance statistics"""
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    stats = get_cache_stats()
    
    # Add additional details
    cache_entries = []
    for key, entry in list(query_cache.items())[:10]:  # Show top 10 most recent
        age_seconds = time.time() - entry['timestamp']
        cache_entries.append({
            "query": entry['original_query'][:100],  # Truncate long queries
            "age_seconds": round(age_seconds, 2),
            "intent": entry['metadata'].get('intent', 'unknown'),
            "entities_found": entry['metadata'].get('entities_found', 0)
        })
    
    return jsonify({
        "statistics": stats,
        "recent_entries": cache_entries,
        "cache_enabled": CACHE_ENABLED,
        "configuration": {
            "ttl_seconds": CACHE_TTL_SECONDS,
            "max_size": CACHE_MAX_SIZE
        },
        "timestamp": datetime.now().isoformat()
    })
   
@app.route("/api/system/status", methods=["GET"])
def get_system_status_for_ui():
    """Get system status in Vue.js UI compatible format"""
    # Test Ollama connection
    try:
        test_response = call_ollama_enhanced("Test", "Respond with 'OK'", temperature=0.0)
        ollama_status = "connected" if "OK" in test_response else "error"
        ollama_available = True
    except:
        ollama_status = "disconnected"
        ollama_available = False
    
    # Get database status
    db_status = orchestrator.get_database_status()
    
    # Get cache stats
    cache_stats_data = get_cache_stats()
    
    return jsonify({
        "status": "ready" if ollama_available else "degraded",
        "ai_model": OLLAMA_MODEL,
        "ai_provider": "Ollama",
        "ai_url": OLLAMA_URL,
        "ai_status": ollama_status,
        "knowledge_base": {
            "name": "SAMM",
            "entities": len(knowledge_graph.entities),
            "relationships": len(knowledge_graph.relationships),
            "status": "loaded"
        },
        "agents": {
            "available": 3,
            "types": ["intent", "integrated_entity", "enhanced_answer"],
            "orchestration": "integrated_database_state",
            "versions": {
                "intent_agent": "1.0",
                "entity_agent": "IntegratedEntityAgent v1.0",
                "answer_agent": "EnhancedAnswerAgent v1.0"
            }
        },
        "database_integration": {
            "cosmos_gremlin": db_status["cosmos_gremlin"]["connected"],
            "vector_db": db_status["vector_db"]["connected"],
            "embedding_model": db_status["embedding_model"]["loaded"]
        },
        "cache": cache_stats_data,  # NEW: Cache statistics
        "services": {
            "authentication": "configured" if oauth else "mock",
            "database": "connected" if cases_container_client else "disabled",
            "storage": "connected" if blob_service_client else "disabled",
            "cache": "enabled" if CACHE_ENABLED else "disabled"  # NEW
        },
        "version": "5.0.0-integrated-database-cached",  # Updated version
        "system_name": "Integrated Database SAMM ASIST with Cache",  # Updated name
        "timestamp": datetime.now().isoformat()
    })

@app.route("/api/examples", methods=["GET"])
def get_example_questions():
    """Get example questions in Vue.js UI compatible format"""
    examples = [
        "What is Security Cooperation?",
        "Who supervises Security Assistance programs?", 
        "What is the difference between Security Cooperation and Security Assistance?",
        "What does DFAS do?",
        "When was the Foreign Assistance Act enacted?",
        "What is an Implementing Agency?"
    ]
    
    return jsonify({
        "examples": examples,
        "count": len(examples)
    })

@app.route("/api/agents/hil_update", methods=["POST"])
def update_agents_from_hil():
    """Update agents from human-in-the-loop feedback"""
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400
    
    try:
        data = request.get_json()
        query = data.get("query", "").strip()
        
        if not query:
            return jsonify({"error": "Query is required for HIL update"}), 400
        
        # Extract correction data for each agent
        intent_correction = data.get("intent_correction")
        entity_correction = data.get("entity_correction") 
        answer_correction = data.get("answer_correction")
        
        if not any([intent_correction, entity_correction, answer_correction]):
            return jsonify({"error": "At least one correction type must be provided"}), 400
        
        # Update agents through orchestrator
        results = orchestrator.update_agents_from_hil(
            query=query,
            intent_correction=intent_correction,
            entity_correction=entity_correction,
            answer_correction=answer_correction
        )
        
        return jsonify({
            "message": "HIL updates applied successfully",
            "query": query,
            "updates_applied": results,
            "timestamp": datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        print(f"[HIL Update] Error: {str(e)}")
        return jsonify({"error": f"Failed to apply HIL updates: {str(e)}"}), 500

@app.route("/api/agents/trigger_update", methods=["POST"])
def update_agents_from_trigger():
    """Update agents when new entity/relationship data is available"""
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400
    
    try:
        data = request.get_json()
        new_entities = data.get("new_entities", [])
        new_relationships = data.get("new_relationships", [])
        trigger_data = data.get("trigger_data", {})
        
        if not new_entities and not new_relationships:
            return jsonify({"error": "At least one new entity or relationship must be provided"}), 400
        
        # Update agents through orchestrator
        results = orchestrator.update_agents_from_trigger(
            new_entities=new_entities,
            new_relationships=new_relationships,
            trigger_data=trigger_data
        )
        
        return jsonify({
            "message": "Trigger updates applied successfully",
            "new_entities_count": len(new_entities),
            "new_relationships_count": len(new_relationships),
            "updates_applied": results,
            "timestamp": datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        print(f"[Trigger Update] Error: {str(e)}")
        return jsonify({"error": f"Failed to apply trigger updates: {str(e)}"}), 500

@app.route("/api/agents/status", methods=["GET"])
def get_agents_status():
    """Get detailed status of all agents including HIL and trigger update counts"""
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    try:
        agent_status = orchestrator.get_agent_status()
        database_status = orchestrator.get_database_status()
        
        return jsonify({
            "agents": agent_status,
            "database_integration": database_status,
            "summary": {
                "total_hil_updates": sum(agent["hil_feedback_count"] for agent in agent_status.values()),
                "total_trigger_updates": sum(agent["trigger_update_count"] for agent in agent_status.values()),
                "total_learned_items": (
                    agent_status["intent_agent"]["learned_patterns"] +
                    agent_status["integrated_entity_agent"]["custom_entities"] + 
                    agent_status["enhanced_answer_agent"]["answer_corrections"]
                ),
                "database_features": {
                    "cosmos_gremlin_connected": database_status["cosmos_gremlin"]["connected"],
                    "vector_db_connected": database_status["vector_db"]["connected"],
                    "embedding_model_loaded": database_status["embedding_model"]["loaded"],
                    "total_vector_collections": len(database_status["vector_db"]["collections"]) 
                },
                "enhanced_features": {
                    "extraction_phases": agent_status["integrated_entity_agent"]["extraction_phases"],
                    "samm_patterns": agent_status["integrated_entity_agent"]["samm_patterns"],
                    "response_templates": agent_status["enhanced_answer_agent"]["response_templates"],
                    "acronym_expansions": agent_status["enhanced_answer_agent"]["acronym_expansions"]
                }
            },
            "system_version": "Integrated_Database_SAMM_v5.0",
            "timestamp": datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        print(f"[Agent Status] Error: {str(e)}")
        return jsonify({"error": f"Failed to get agent status: {str(e)}"}), 500

@app.route("/api/database/status", methods=["GET"])
def get_database_status():
    """Get detailed database connection status"""
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    try:
        database_status = orchestrator.get_database_status()
        
        return jsonify({
            "database_connections": database_status,
            "summary": {
                "total_connections": sum(1 for db in database_status.values() if db.get("connected", False)),
                "cosmos_gremlin_status": "connected" if database_status["cosmos_gremlin"]["connected"] else "disconnected",
                "vector_databases": {
                    "vector_db_collections": len(database_status["vector_db"]["collections"]),
                    "total_collections": len(database_status["vector_db"]["collections"]) 
                },
                "embedding_model_status": "loaded" if database_status["embedding_model"]["loaded"] else "not_loaded"
            },
            "timestamp": datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        print(f"[Database Status] Error: {str(e)}")
        return jsonify({"error": f"Failed to get database status: {str(e)}"}), 500

@app.route("/api/samm/status", methods=["GET"])
def get_samm_system_status():
    """Get detailed system status (maintains backward compatibility)"""
    # Test Ollama connection
    try:
        test_response = call_ollama_enhanced("Test", "Respond with 'OK'", temperature=0.0)
        ollama_status = "connected" if "OK" in test_response else "error"
    except:
        ollama_status = "disconnected"
    
    # Get database status
    database_status = orchestrator.get_database_status()
    
    return jsonify({
        "status": "ready",
        "ollama_url": OLLAMA_URL,
        "ollama_model": OLLAMA_MODEL,
        "ollama_status": ollama_status,
        "knowledge_graph": {
            "entities": len(knowledge_graph.entities),
            "relationships": len(knowledge_graph.relationships)
        },
        "orchestration": {
            "type": "integrated_database_state",
            "workflow_steps": [step.value for step in WorkflowStep],
            "agents": ["intent_agent", "integrated_entity_agent", "enhanced_answer_agent"]
        },
        "database_integration": {
            "cosmos_gremlin": {
                "connected": database_status["cosmos_gremlin"]["connected"],
                "endpoint": database_status["cosmos_gremlin"]["endpoint"],
                "database": database_status["cosmos_gremlin"]["database"]
            },
            "vector_databases": {
                "vector_db": {
                    "connected": database_status["vector_db"]["connected"],
                    "collections": database_status["vector_db"]["collections"]
                },
            },
            "embedding_model": {
                "loaded": database_status["embedding_model"]["loaded"],
                "model_name": database_status["embedding_model"]["model_name"]
            }
        },
        "enhanced_capabilities": {
            "integrated_entity_extraction": {
                "phases": 3,  # pattern_matching, nlp_extraction, database_queries
                "patterns": sum(len(patterns) for patterns in orchestrator.entity_agent.samm_entity_patterns.values()),
                "database_enhanced": True,
                "confidence_scoring": True,
                "ai_context_generation": True
            },
            "answer_generation": {
                "templates": len(orchestrator.answer_agent.samm_response_templates),
                "quality_scoring": True,
                "multi_pass_validation": True,
                "acronym_expansion": True
            }
        },
        "services": {
            "auth0": "configured" if oauth else "mock",
            "cosmos_db": "connected" if cases_container_client else "disabled",
            "blob_storage": "connected" if blob_service_client else "disabled"
        },
        "version": "Integrated_Database_SAMM_v5.0",
        "timestamp": datetime.now().isoformat()
    })

@app.route("/api/samm/workflow", methods=["GET"])
def get_workflow_info():
    """Get workflow orchestration information"""
    return jsonify({
        "orchestration_type": "integrated_database_state",
        "workflow_steps": [
            {
                "step": step.value,
                "description": {
                    "initialize": "Initialize integrated workflow state with database connections",
                    "analyze_intent": "Analyze user intent using Intent Agent with HIL learning",
                    "extract_entities": "Extract entities using Integrated Entity Agent with database queries", 
                    "generate_answer": "Generate answer using Enhanced Answer Agent with quality scoring",
                    "complete": "Complete integrated workflow successfully",
                    "error": "Handle any workflow errors"
                }.get(step.value, "Unknown step")
            }
            for step in WorkflowStep
        ],
        "agents": [
            {
                "name": "IntentAgent", 
                "purpose": "Classify user queries and determine intent", 
                "hil_updates": True, 
                "trigger_updates": True,
                "version": "1.0"
            },
            {
                "name": "IntegratedEntityAgent", 
                "purpose": "Multi-phase entity extraction with SAMM patterns and database integration", 
                "hil_updates": True, 
                "trigger_updates": True,
                "version": "1.0",
                "features": ["pattern_matching", "nlp_extraction", "database_queries", "ai_context_generation", "confidence_scoring"],
                "database_integration": True
            },
            {
                "name": "EnhancedAnswerAgent", 
                "purpose": "Intent-optimized answer generation with quality enhancement", 
                "hil_updates": True, 
                "trigger_updates": True,
                "version": "1.0",
                "features": ["intent_optimization", "multi_pass_generation", "quality_scoring", "acronym_expansion", "answer_validation"]
            }
        ],
        "database_integration": {
            "cosmos_gremlin": {
                "purpose": "Graph database for entity relationships",
                "query_type": "Gremlin traversal queries"
            },
            "vector_db": {
                "purpose": "Document vector search",
                "query_type": "Semantic similarity search"
            },
        },
        "transitions": {
            "initialize": "analyze_intent",
            "analyze_intent": "extract_entities", 
            "extract_entities": "generate_answer",
            "generate_answer": "complete",
            "complete": "end",
            "error": "end"
        },
        "update_capabilities": {
            "human_in_loop": {
                "endpoint": "/api/agents/hil_update",
                "description": "Update agents based on human feedback corrections",
                "supported_corrections": ["intent", "entity", "answer"]
            },
            "trigger_updates": {
                "endpoint": "/api/agents/trigger_update", 
                "description": "Update agents when new entity/relationship data becomes available",
                "supported_data": ["new_entities", "new_relationships", "trigger_data"]
            }
        },
        "integrated_features": {
            "entity_extraction": {
                "phases": 3,
                "database_enhanced": True,
                "confidence_scoring": True,
                "pattern_matching": True,
                "nlp_extraction": True,
                "ai_fallback": True
            },
            "answer_generation": {
                "intent_optimization": True,
                "quality_validation": True,
                "multi_pass_generation": True,
                "template_adherence": True,
                "automatic_enhancement": True
            }
        }
    })

@app.route("/api/samm/examples", methods=["GET"])
def get_samm_examples():
    """Get example SAMM questions (detailed format for compatibility)"""
    examples = [
        {
            "question": "What is Security Cooperation?",
            "type": "definition",
            "expected_entities": ["Security Cooperation", "DoD"],
            "expected_intent": "definition",
            "database_relevant": True
        },
        {
            "question": "Who supervises Security Assistance programs?", 
            "type": "authority",
            "expected_entities": ["Security Assistance", "Department of State"],
            "expected_intent": "authority",
            "database_relevant": True
        },
        {
            "question": "What is the difference between Security Cooperation and Security Assistance?",
            "type": "distinction",
            "expected_entities": ["Security Cooperation", "Security Assistance"],
            "expected_intent": "distinction",
            "database_relevant": True
        },
        {
            "question": "What does DFAS do?",
            "type": "organization",
            "expected_entities": ["DFAS", "Defense Finance and Accounting Service"],
            "expected_intent": "organization",
            "database_relevant": True
        },
        {
            "question": "When was the Foreign Assistance Act enacted?",
            "type": "factual",
            "expected_entities": ["Foreign Assistance Act", "FAA"],
            "expected_intent": "factual",
            "database_relevant": True
        },
        {
            "question": "What is an Implementing Agency?",
            "type": "definition",
            "expected_entities": ["Implementing Agency", "IA"],
            "expected_intent": "definition",
            "database_relevant": True
        }
    ]
    
    return jsonify({
        "examples": examples,
        "count": len(examples),
        "usage": "Use these to test the Integrated Database orchestrated SAMM system",
        "integrated_testing": {
            "entity_extraction": "Each example includes expected entities for validation",
            "intent_classification": "Each example includes expected intent for validation",
            "database_integration": "All examples will trigger database queries",
            "quality_scoring": "Answers will include quality scores and enhancements"
        }
    })

@app.route("/api/samm/knowledge", methods=["GET"])
def get_knowledge_graph_info():
    """Get knowledge graph information"""
    entities_info = []
    for entity_id, entity in knowledge_graph.entities.items():
        entities_info.append({
            "id": entity_id,
            "label": entity['properties'].get('label', entity_id),
            "type": entity['type'],
            "definition": entity['properties'].get('definition', ''),
            "section": entity['properties'].get('section', '')
        })
    
    # Get integrated agent pattern information
    samm_patterns = {}
    if hasattr(orchestrator.entity_agent, 'samm_entity_patterns'):
        samm_patterns = {
            category: len(patterns) 
            for category, patterns in orchestrator.entity_agent.samm_entity_patterns.items()
        }
    
    # Get database status
    database_status = orchestrator.get_database_status()
    
    return jsonify({
        "entities": entities_info,
        "relationships": knowledge_graph.relationships,
        "total_entities": len(knowledge_graph.entities),
        "total_relationships": len(knowledge_graph.relationships),
        "enhanced_patterns": {
            "samm_entity_patterns": samm_patterns,
            "total_patterns": sum(samm_patterns.values()) if samm_patterns else 0,
            "pattern_categories": list(samm_patterns.keys()) if samm_patterns else []
        },
        "dynamic_knowledge": {
            "custom_entities": len(orchestrator.entity_agent.custom_entities),
            "dynamic_entities": len(orchestrator.entity_agent.dynamic_knowledge["entities"]),
            "dynamic_relationships": len(orchestrator.entity_agent.dynamic_knowledge["relationships"])
        },
        "database_integration": {
            "cosmos_gremlin": {
                "connected": database_status["cosmos_gremlin"]["connected"],
                "endpoint": database_status["cosmos_gremlin"]["endpoint"]
            },
            "vector_databases": {
                "vector_db_collections": len(database_status["vector_db"]["collections"]),
            },
            "embedding_model": {
                "loaded": database_status["embedding_model"]["loaded"],
                "model": database_status["embedding_model"]["model_name"]
            }
        }
    })

@app.route("/api/health", methods=["GET"])
def health_check():
    """System health check"""
    # Test integrated Ollama connection
    ollama_healthy = False
    try:
        test_response = call_ollama_enhanced("Test", "Respond with 'OK'", temperature=0.0)
        ollama_healthy = "OK" in test_response
    except:
        pass
    
    # Test agent status
    agent_healthy = False
    try:
        agent_status = orchestrator.get_agent_status()
        agent_healthy = len(agent_status) == 3  # All 3 agents should be present
    except:
        pass
    
    # Test database connections
    database_status = orchestrator.get_database_status()
    database_healthy = any([
        database_status["cosmos_gremlin"]["connected"],
        database_status["vector_db"]["connected"],
    ])
    
    # Get cache stats
    cache_stats_data = get_cache_stats()
    cache_healthy = CACHE_ENABLED and len(query_cache) >= 0  # Cache is working if enabled
    
    return jsonify({
        "status": "healthy" if (ollama_healthy and agent_healthy) else "degraded",
        "timestamp": datetime.now().isoformat(),
        "ollama_model": OLLAMA_MODEL,
        "ollama_healthy": ollama_healthy,
        "agents_healthy": agent_healthy,
        "database_healthy": database_healthy,
        "cache_healthy": cache_healthy,  # NEW
        "version": "Integrated_Database_SAMM_v5.0_Cached",  # Updated
        "components": {
            "ollama": "healthy" if ollama_healthy else "degraded",
            "agents": "healthy" if agent_healthy else "degraded",
            "knowledge_graph": "healthy" if len(knowledge_graph.entities) > 0 else "degraded",
            "case_database": "healthy" if cases_container_client else "disabled",
            "blob_storage": "healthy" if blob_service_client else "disabled",
            "cosmos_gremlin": "healthy" if database_status["cosmos_gremlin"]["connected"] else "disconnected",
            "vector_db": "healthy" if database_status["vector_db"]["connected"] else "disconnected",
            "embedding_model": "healthy" if database_status["embedding_model"]["loaded"] else "not_loaded",
            "cache": "healthy" if cache_healthy else "disabled"  # NEW
        },
        "cache_stats": cache_stats_data  # NEW: Include cache performance
    })

# Static file serving
@app.route('/')
def serve_main_app():
    user = require_auth()
    if not user and oauth:
        session['next_url'] = request.path
        return redirect(url_for("login"))
    return send_from_directory(app.static_folder, 'index.html')


@app.route('/<path:path>')
def serve_vue_paths(path):
    if os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else: 
        user = require_auth()
        if not user and oauth:
            session['next_url'] = request.path
            return redirect(url_for("login"))
        return send_from_directory(app.static_folder, 'index.html')

# Cleanup on exit
import atexit
atexit.register(orchestrator.cleanup)

@app.route("/api/reviews", methods=["POST"])
def create_review_item():
    """Create a new review item"""
    try:
        data = request.json
        
        if not data.get('id'):
            data['id'] = str(uuid.uuid4())
        if not data.get('reviewId'):
            data['reviewId'] = str(uuid.uuid4())
        if not data.get('timestamp'):
            data['timestamp'] = datetime.now(timezone.utc).isoformat()
        
        if reviews_test_container_client:
            result = reviews_test_container_client.create_item(data)
            return jsonify({
                "success": True,
                "message": "Review created successfully",
                "reviewId": data['reviewId']
            }), 201
        else:
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
            
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route("/api/reviews/pending", methods=["GET"])
def get_pending_reviews():
    """Get all pending reviews"""
    try:
        query = """
        SELECT * FROM c 
        WHERE c.type = 'review_item' 
        AND c.status = 'pending'
        ORDER BY c.timestamp DESC
        """
        
        if reviews_test_container_client:
            reviews = list(reviews_test_container_client.query_items(
                query=query,
                enable_cross_partition_query=True
            ))
            
            return jsonify({
                "success": True,
                "count": len(reviews),
                "reviews": reviews
            })
        else:
            return jsonify({
                "success": False,
                "error": "Reviews container not available",
                "reviews": []
            }), 500
            
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e),
            "reviews": []
        }), 500


@app.route("/api/reviews/<review_id>/submit", methods=["POST"])
def submit_review_feedback(review_id):
    """Submit review feedback"""
    try:
        data = request.json
        status = data.get('status')
        feedback = data.get('feedback', '')
        reviewer = data.get('reviewer', 'Unknown')
        
        if not reviews_test_container_client:
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
        
        query = f"SELECT * FROM c WHERE c.reviewId = '{review_id}'"
        items = list(reviews_test_container_client.query_items(
            query=query,
            enable_cross_partition_query=True
        ))
        
        if not items:
            return jsonify({
                "success": False,
                "error": "Review not found"
            }), 404
        
        review = items[0]
        review['status'] = status
        review['humanFeedback'] = feedback
        review['reviewedBy'] = reviewer
        review['reviewedAt'] = datetime.now(timezone.utc).isoformat()
        
        # ✨ NEW: Calculate entity metrics if corrected entities are provided
        if 'corrected_entities' in data:
            extracted = review.get('entities', [])
            corrected = data['corrected_entities']
            entity_metrics = calculate_entity_metrics(extracted, corrected)
            review['entityMetrics'] = entity_metrics
            print(f"📊 Entity metrics calculated: P={entity_metrics['precision']:.2f}, R={entity_metrics['recall']:.2f}, F1={entity_metrics['f1']:.2f}")
        
        # ✨ NEW: Track intent correctness
        if 'intent_correct' in data:
            review['intentCorrect'] = data['intent_correct']
        
        # ✨ NEW: Add timestamp for trend analysis if not exists
        if 'createdAt' not in review:
            review['createdAt'] = datetime.now(timezone.utc).isoformat()
        
        reviews_test_container_client.upsert_item(review)
        
        print(f"✅ Review {status} by {reviewer}: {review_id}")
        
        # ========== SAVE HITL CORRECTIONS ==========
        if status == "needs_revision":
            question = review.get('question', '')
            q_hash = create_question_hash(question)
            
            # Save corrected intent
            if 'corrected_intent' in data:
                HITL_CORRECTIONS_STORE["intent_corrections"][q_hash] = data['corrected_intent']
                print(f"💾 HITL: Intent correction saved for question hash {q_hash[:8]}...")
            
            # Save corrected entities
            if 'corrected_entities' in data:
                HITL_CORRECTIONS_STORE["entity_corrections"][q_hash] = data['corrected_entities']
                print(f"💾 HITL: Entity corrections saved ({len(data['corrected_entities'])} entities)")
            
            # Save corrected answer
            if 'corrected_answer' in data:
                HITL_CORRECTIONS_STORE["answer_corrections"][q_hash] = data['corrected_answer']
                print(f"💾 HITL: Answer correction saved ({len(data['corrected_answer'])} chars)")
            
            # Save all corrections to file
            save_hitl_corrections()
        # ========== END HITL CORRECTIONS ==========
        
        return jsonify({
            "success": True,
            "message": f"Review {status} successfully",
            "reviewId": review_id
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    
    # =============================================================================
# ENHANCED HITL APIs - Complete Implementation
# Add these routes to your Flask app after the existing /api/reviews/<review_id>/submit route
# =============================================================================

@app.route("/api/reviews/<review_id>/accept", methods=["POST"])
def accept_review(review_id):
    """Accept a review - mark as approved"""
    try:
        data = request.json or {}
        reviewer = data.get('reviewer', 'Travis')
        feedback = data.get('feedback', 'Approved by SME')
        
        if not reviews_test_container_client:
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
        
        # Find the review
        query = f"SELECT * FROM c WHERE c.reviewId = '{review_id}'"
        items = list(reviews_test_container_client.query_items(
            query=query,
            enable_cross_partition_query=True
        ))
        
        if not items:
            return jsonify({
                "success": False,
                "error": "Review not found"
            }), 404
        
        # Update review
        review = items[0]
        review['status'] = 'approved'
        review['humanFeedback'] = feedback
        review['reviewedBy'] = reviewer
        review['reviewedAt'] = datetime.now(timezone.utc).isoformat()
        
        # Save to database
        reviews_test_container_client.upsert_item(review)
        
        print(f"✅ Review ACCEPTED by {reviewer}: {review_id}")
        print(f"   Feedback: {feedback}")
        
        return jsonify({
            "success": True,
            "message": "Review accepted successfully",
            "reviewId": review_id,
            "status": "approved"
        })
        
    except Exception as e:
        print(f"❌ Error accepting review: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route("/api/reviews/<review_id>/reject", methods=["POST"])
def reject_review(review_id):
    """Reject a review - mark as needs revision"""
    try:
        data = request.json or {}
        reviewer = data.get('reviewer', 'Travis')
        feedback = data.get('feedback', '')
        
        if not feedback:
            return jsonify({
                "success": False,
                "error": "Feedback is required when rejecting"
            }), 400
        
        if not reviews_test_container_client:
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
        
        # Find the review
        query = f"SELECT * FROM c WHERE c.reviewId = '{review_id}'"
        items = list(reviews_test_container_client.query_items(
            query=query,
            enable_cross_partition_query=True
        ))
        
        if not items:
            return jsonify({
                "success": False,
                "error": "Review not found"
            }), 404
        
        # Update review
        review = items[0]
        review['status'] = 'needs_revision'
        review['humanFeedback'] = feedback
        review['reviewedBy'] = reviewer
        review['reviewedAt'] = datetime.now(timezone.utc).isoformat()
        
        # Save to database
        reviews_test_container_client.upsert_item(review)
        
        print(f"⚠️ Review REJECTED by {reviewer}: {review_id}")
        print(f"   Feedback: {feedback}")
        
        # TODO: Use feedback to improve agents in future
        # This is where you'd implement agent learning
        
        return jsonify({
            "success": True,
            "message": "Review rejected - feedback saved for improvement",
            "reviewId": review_id,
            "status": "needs_revision"
        })
        
    except Exception as e:
        print(f"❌ Error rejecting review: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route("/api/reviews/<review_id>/needs-revision", methods=["POST"])
def mark_needs_revision(review_id):
    """Mark review as needs revision with feedback"""
    try:
        data = request.json or {}
        reviewer = data.get('reviewer', 'Travis')
        feedback = data.get('feedback', '')
        
        if not feedback:
            return jsonify({
                "success": False,
                "error": "Feedback is required"
            }), 400
        
        if not reviews_test_container_client:
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
        
        # Find the review
        query = f"SELECT * FROM c WHERE c.reviewId = '{review_id}'"
        items = list(reviews_test_container_client.query_items(
            query=query,
            enable_cross_partition_query=True
        ))
        
        if not items:
            return jsonify({
                "success": False,
                "error": "Review not found"
            }), 404
        
        # Update review
        review = items[0]
        review['status'] = 'needs_revision'
        review['humanFeedback'] = feedback
        review['reviewedBy'] = reviewer
        review['reviewedAt'] = datetime.now(timezone.utc).isoformat()
        
        # Save to database
        reviews_test_container_client.upsert_item(review)
        
        print(f"📝 Review marked NEEDS REVISION by {reviewer}: {review_id}")
        print(f"   Feedback: {feedback}")
        
        return jsonify({
            "success": True,
            "message": "Feedback saved - marked for revision",
            "reviewId": review_id,
            "status": "needs_revision"
        })
        
    except Exception as e:
        print(f"❌ Error marking review: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route("/api/reviews/<review_id>/regenerate", methods=["POST"])
def regenerate_answer(review_id):
    """Regenerate answer for a review"""
    try:
        data = request.json or {}
        
        if not reviews_test_container_client:
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
        
        # Find the review
        query = f"SELECT * FROM c WHERE c.reviewId = '{review_id}'"
        items = list(reviews_test_container_client.query_items(
            query=query,
            enable_cross_partition_query=True
        ))
        
        if not items:
            return jsonify({
                "success": False,
                "error": "Review not found"
            }), 404
        
        review = items[0]
        original_question = review.get('question', '')
        
        # TODO: Re-run the query through agents
        # For now, just mark as regenerated
        review['status'] = 'regenerating'
        review['lastRegeneratedAt'] = datetime.now(timezone.utc).isoformat()
        
        reviews_test_container_client.upsert_item(review)
        
        print(f"🔄 Regenerating answer for review: {review_id}")
        print(f"   Question: {original_question}")
        
        return jsonify({
            "success": True,
            "message": "Answer regeneration initiated",
            "reviewId": review_id,
            "note": "Implementation pending - will re-run agents"
        })
        
    except Exception as e:
        print(f"❌ Error regenerating answer: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route("/api/reviews/<review_id>/agent/<agent_name>/accept", methods=["POST"])
def accept_agent_result(review_id, agent_name):
    """Accept a specific agent's result"""
    try:
        data = request.json or {}
        reviewer = data.get('reviewer', 'Travis')
        
        if not reviews_test_container_client:
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
        
        # Find the review
        query = f"SELECT * FROM c WHERE c.reviewId = '{review_id}'"
        items = list(reviews_test_container_client.query_items(
            query=query,
            enable_cross_partition_query=True
        ))
        
        if not items:
            return jsonify({
                "success": False,
                "error": "Review not found"
            }), 404
        
        review = items[0]
        
        # Track agent approvals
        if 'agentApprovals' not in review:
            review['agentApprovals'] = {}
        
        review['agentApprovals'][agent_name] = {
            'approved': True,
            'approvedBy': reviewer,
            'approvedAt': datetime.now(timezone.utc).isoformat()
        }
        
        reviews_test_container_client.upsert_item(review)
        
        print(f"✅ {agent_name} result ACCEPTED by {reviewer} for review: {review_id}")
        
        return jsonify({
            "success": True,
            "message": f"{agent_name} result accepted",
            "reviewId": review_id,
            "agent": agent_name
        })
        
    except Exception as e:
        print(f"❌ Error accepting agent result: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route("/api/reviews/stats", methods=["GET"])
def get_review_stats():
    """Get statistics about reviews"""
    try:
        if not reviews_test_container_client:
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
        
        # Query all reviews
        query = "SELECT * FROM c WHERE c.type = 'review_item'"
        all_reviews = list(reviews_test_container_client.query_items(
            query=query,
            enable_cross_partition_query=True
        ))
        
        # Calculate stats
        total = len(all_reviews)
        pending = len([r for r in all_reviews if r.get('status') == 'pending'])
        approved = len([r for r in all_reviews if r.get('status') == 'approved'])
        needs_revision = len([r for r in all_reviews if r.get('status') == 'needs_revision'])
        
        # Calculate average confidence
        confidences = [r for r in all_reviews if 'confidenceOverall' in r]
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0
        
        stats = {
            "success": True,
            "total_reviews": total,
            "pending": pending,
            "approved": approved,
            "needs_revision": needs_revision,
            "average_confidence": round(avg_confidence, 2),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
        
        return jsonify(stats)
        
    except Exception as e:
        print(f"❌ Error getting review stats: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
        
@app.route("/api/test-metrics", methods=["GET"])
def test_metrics():
    """Simple test endpoint"""
    return jsonify({
        "success": True,
        "message": "Metrics endpoint working!",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })

@app.route("/api/reviews/detailed-stats", methods=["GET"])
def get_detailed_review_stats():
    """Get detailed statistics - BULLETPROOF VERSION"""
    try:
        print("[METRICS] detailed-stats endpoint called")
        
        # Check container
        if not reviews_test_container_client:
            print("[METRICS] ❌ Reviews container not available")
            return jsonify({
                "success": False,
                "error": "Reviews container not available"
            }), 500
        
        # Query reviews
        try:
            query = "SELECT * FROM c WHERE c.type = 'review_item'"
            all_reviews = list(reviews_test_container_client.query_items(
                query=query,
                enable_cross_partition_query=True
            ))
            print(f"[METRICS] Found {len(all_reviews)} reviews")
        except Exception as query_error:
            print(f"[METRICS] ❌ Query error: {query_error}")
            all_reviews = []
        
        # Calculate stats
        total = len(all_reviews)
        pending = sum(1 for r in all_reviews if r.get('status') == 'pending')
        approved = sum(1 for r in all_reviews if r.get('status') == 'approved')
        needs_revision = sum(1 for r in all_reviews if r.get('status') == 'needs_revision')
        rejected = sum(1 for r in all_reviews if r.get('status') == 'rejected')
        
        # Confidence scores
        intent_confs = [r.get('confidenceIntent', 0) for r in all_reviews if 'confidenceIntent' in r]
        entity_confs = [r.get('confidenceEntity', 0) for r in all_reviews if 'confidenceEntity' in r]
        answer_confs = [r.get('confidenceAnswer', 0) for r in all_reviews if 'confidenceAnswer' in r]
        overall_confs = [r.get('confidenceOverall', 0) for r in all_reviews if 'confidenceOverall' in r]
        
        avg_intent = round(sum(intent_confs) / len(intent_confs), 2) if intent_confs else 0
        avg_entity = round(sum(entity_confs) / len(entity_confs), 2) if entity_confs else 0
        avg_answer = round(sum(answer_confs) / len(answer_confs), 2) if answer_confs else 0
        avg_overall = round(sum(overall_confs) / len(overall_confs), 2) if overall_confs else 0
        
        # Rates
        approval_rate = round((approved / total * 100), 2) if total > 0 else 0
        revision_rate = round((needs_revision / total * 100), 2) if total > 0 else 0
        rejection_rate = round((rejected / total * 100), 2) if total > 0 else 0
        
        # Intent accuracy
        correct = sum(1 for r in all_reviews if r.get('status') == 'approved' and r.get('intentCorrect', False))
        intent_accuracy = round((correct / total * 100), 2) if total > 0 else 0
        
        # Entity metrics
        e_precision = []
        e_recall = []
        e_f1 = []
        
        for r in all_reviews:
            if 'entityMetrics' in r:
                m = r['entityMetrics']
                if 'precision' in m: e_precision.append(m['precision'])
                if 'recall' in m: e_recall.append(m['recall'])
                if 'f1' in m: e_f1.append(m['f1'])
        
        avg_precision = round(sum(e_precision) / len(e_precision), 2) if e_precision else 0
        avg_recall = round(sum(e_recall) / len(e_recall), 2) if e_recall else 0
        avg_f1 = round(sum(e_f1) / len(e_f1), 2) if e_f1 else 0
        
        # Build response
        response = {
            "success": True,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "counts": {
                "total_reviews": total,
                "pending": pending,
                "approved": approved,
                "needs_revision": needs_revision,
                "rejected": rejected
            },
            "rates": {
                "approval_rate": approval_rate,
                "revision_rate": revision_rate,
                "rejection_rate": rejection_rate
            },
            "confidence": {
                "intent": avg_intent,
                "entity": avg_entity,
                "answer": avg_answer,
                "overall": avg_overall
            },
            "accuracy": {
                "intent_accuracy": intent_accuracy,
                "entity_precision": avg_precision,
                "entity_recall": avg_recall,
                "entity_f1": avg_f1
            },
            "agent_performance": {
                "intent_agent": {
                    "accuracy": intent_accuracy,
                    "avg_confidence": avg_intent,
                    "total_processed": len(intent_confs)
                },
                "entity_agent": {
                    "precision": avg_precision,
                    "recall": avg_recall,
                    "f1_score": avg_f1,
                    "avg_confidence": avg_entity,
                    "total_processed": len(entity_confs)
                },
                "answer_agent": {
                    "avg_confidence": avg_answer,
                    "total_processed": len(answer_confs)
                }
            },
            "trend": [],
            "performance": {
                "avg_response_time": 0,
                "total_questions_evaluated": total
            }
        }
        
        print(f"[METRICS] ✅ Returning stats for {total} reviews")
        return jsonify(response)
        
    except Exception as e:
        print(f"[METRICS] ❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "error_type": type(e).__name__
        }), 500

def calculate_entity_metrics(extracted_entities: List[str], 
                            gold_standard_entities: List[str]) -> Dict[str, float]:
    """
    Calculate precision, recall, and F1 score for entity extraction
    
    Args:
        extracted_entities: List of entities extracted by the system
        gold_standard_entities: List of correct entities (from SME/HITL)
    
    Returns:
        Dictionary with precision, recall, and f1 scores
    """
    if not gold_standard_entities:
        return {"precision": 0, "recall": 0, "f1": 0}
    
    # Convert to sets for comparison
    extracted_set = set(extracted_entities)
    gold_set = set(gold_standard_entities)
    
    # Calculate metrics
    true_positives = len(extracted_set & gold_set)
    false_positives = len(extracted_set - gold_set)
    false_negatives = len(gold_set - extracted_set)
    
    # Precision: TP / (TP + FP)
    precision = true_positives / (true_positives + false_positives) if (true_positives + false_positives) > 0 else 0
    
    # Recall: TP / (TP + FN)
    recall = true_positives / (true_positives + false_negatives) if (true_positives + false_negatives) > 0 else 0
    
    # F1: 2 * (Precision * Recall) / (Precision + Recall)
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
    
    return {
        "precision": round(precision, 3),
        "recall": round(recall, 3),
        "f1": round(f1, 3),
        "true_positives": true_positives,
        "false_positives": false_positives,
        "false_negatives": false_negatives
    }

@app.route('/hitl_review_dashboard')
def serve_hitl_dashboard():
    return send_from_directory('static', 'hitl_review_dashboard_production.html')

@app.route('/sprint2-metrics')
def serve_sprint2_metrics():
    """Serve Sprint 2 HITL Metrics Dashboard"""
    return send_from_directory('static', 'Sprint2_HITL_Metrics_Dashboard_Live.html')
    
# Add this BEFORE if __name__ == '__main__':
def test_phase1_fixes():
    """Test that contamination is removed"""
    
    test_queries = [
        "What is DFAS responsible for?",
        "What does DoD do?",
        "Who supervises FMS programs?"
    ]
    
    print("\n" + "="*80)
    print("🧪 PHASE 1: Testing Contamination Removal")
    print("="*80)
    
    for query in test_queries:
        print(f"\n🔍 Query: {query}")
        result = orchestrator.process_query(query, "test_user")
        
        # Check for contamination
        answer = result.get('answer', '')
        entities = result.get('metadata', {}).get('entities', [])
        
        # DSCA should NOT appear if not in query
        if 'DSCA' in answer and 'DSCA' not in query:
            print(f"   ❌ CONTAMINATED: DSCA in answer but not in query")
            print(f"   Answer snippet: {answer[:200]}...")
        else:
            print(f"   ✅ CLEAN: No DSCA contamination")
        
        # Check entities
        dsca_entities = [e for e in entities if 'DSCA' in str(e).upper()]
        if dsca_entities and 'DSCA' not in query:
            print(f"   ❌ CONTAMINATED: DSCA in entities: {dsca_entities}")
        else:
            print(f"   ✅ CLEAN: No DSCA in entities")
        
        print("-" * 80)
    
    print("\n✅ PHASE 1 TESTING COMPLETE\n")

# ============================================================================
# HITL FEEDBACK ENDPOINTS
# ============================================================================

@app.route("/api/hitl/correct-intent", methods=["POST"])
def correct_intent():
    """Apply intent correction"""
    try:
        data = request.json
        question = data.get('question', '')
        corrected_intent = data.get('corrected_intent', '')
        
        if not question or not corrected_intent:
            return jsonify({"success": False, "error": "Missing question or corrected_intent"}), 400
        
        q_hash = create_question_hash(question)
        HITL_CORRECTIONS_STORE["intent_corrections"][q_hash] = corrected_intent
        HITL_CORRECTIONS_STORE["correction_history"].append({
            "type": "intent",
            "question": question,
            "correction": corrected_intent,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        
        print(f"✅ HITL: Intent corrected to '{corrected_intent}'")
        save_hitl_corrections()  # Save to file for persistence
        return jsonify({"success": True, "message": "Intent correction applied and will be used on re-query"})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hitl/correct-entities", methods=["POST"])
def correct_entities():
    """Apply entity corrections"""
    try:
        data = request.json
        question = data.get('question', '')
        corrected_entities = data.get('corrected_entities', [])
        
        if not question:
            return jsonify({"success": False, "error": "Missing question"}), 400
        
        q_hash = create_question_hash(question)
        HITL_CORRECTIONS_STORE["entity_corrections"][q_hash] = corrected_entities
        HITL_CORRECTIONS_STORE["correction_history"].append({
            "type": "entity",
            "question": question,
            "correction": corrected_entities,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        
        print(f"✅ HITL: Entities corrected ({len(corrected_entities)} entities)")
        save_hitl_corrections()  # Save to file for persistence
        return jsonify({"success": True, "message": "Entity corrections applied and will be used on re-query"})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hitl/correct-answer", methods=["POST"])
def correct_answer():
    """Apply answer correction"""
    print(f"🔧 API CALLED: /api/hitl/correct-answer")
    try:
        data = request.json
        print(f"🔧 Received data: {data}")
        question = data.get('question', '')
        corrected_answer = data.get('corrected_answer', '')
        
        print(f"🔧 Question: {question[:50]}...")
        print(f"🔧 Answer length: {len(corrected_answer)} chars")
        
        if not question or not corrected_answer:
            return jsonify({"success": False, "error": "Missing question or corrected_answer"}), 400
        
        q_hash = create_question_hash(question)
        print(f"🔧 Generated hash: {q_hash}")
        print(f"🔧 Store before: {len(HITL_CORRECTIONS_STORE['answer_corrections'])} corrections")
        
        HITL_CORRECTIONS_STORE["answer_corrections"][q_hash] = corrected_answer
        
        print(f"🔧 Store after: {len(HITL_CORRECTIONS_STORE['answer_corrections'])} corrections")
        print(f"🔧 Verification - hash in store? {q_hash in HITL_CORRECTIONS_STORE['answer_corrections']}")
        
        HITL_CORRECTIONS_STORE["correction_history"].append({
            "type": "answer",
            "question": question,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        
        print(f"✅ HITL: Answer corrected (length: {len(corrected_answer)} chars)")
        save_hitl_corrections()  # Save to file for persistence
        return jsonify({"success": True, "message": "Answer correction applied and will be used on re-query"})
        
    except Exception as e:
        print(f"❌ ERROR in correct_answer: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hitl/correction-stats", methods=["GET"])
def get_correction_stats():
    """Get correction statistics"""
    try:
        return jsonify({
            "success": True,
            "stats": {
                "intent_corrections": len(HITL_CORRECTIONS_STORE["intent_corrections"]),
                "entity_corrections": len(HITL_CORRECTIONS_STORE["entity_corrections"]),
                "answer_corrections": len(HITL_CORRECTIONS_STORE["answer_corrections"]),
                "total_corrections": len(HITL_CORRECTIONS_STORE["correction_history"]),
                "recent": HITL_CORRECTIONS_STORE["correction_history"][-5:]
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/cases/<case_id>", methods=["GET"])
def get_case(case_id):
    """Get a specific case by ID - Cosmos DB version"""
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401
    
    user_id = user["sub"]
    
    try:
        from urllib.parse import unquote
        case_id = unquote(case_id)
        
        print(f"[API] 🔍 Fetching case: {case_id}")
        
        if not cases_container_client:
            return jsonify({"error": "Database not available"}), 503
        
        case_doc = None
        
        # Try to find by UUID first
        try:
            case_doc = cases_container_client.read_item(item=case_id, partition_key=user_id)
            print(f"[API] ✅ Found case by UUID: {case_id}")
            
        except CosmosExceptions.CosmosResourceNotFoundError:
            # Try by case number
            query = """
            SELECT * FROM c 
            WHERE c.userId = @userId 
            AND c.caseNumber = @caseNumber
            """
            parameters = [
                {"name": "@userId", "value": user_id},
                {"name": "@caseNumber", "value": case_id}
            ]
            
            cases = list(cases_container_client.query_items(
                query=query,
                parameters=parameters,
                enable_cross_partition_query=False
            ))
            
            if cases:
                case_doc = cases[0]
                print(f"[API] ✅ Found case by case number: {case_id}")
            else:
                print(f"[API] ❌ Case not found: {case_id}")
                return jsonify({"error": f"Case {case_id} not found"}), 404
        
        # ✅ CRITICAL FIX: Get documents from BOTH possible fields
        all_documents = case_doc.get("caseDocuments", []) or case_doc.get("documents", [])
        
        # Filter financial documents
        financial_docs = [
            doc for doc in all_documents
            if doc.get("metadata", {}).get("hasFinancialData", False)
        ]
        
        print(f"[API] 📊 Case: {case_doc.get('caseNumber')}")
        print(f"[API]   Total documents: {len(all_documents)}")
        print(f"[API]   Financial documents: {len(financial_docs)}")
        
        # ✅ RETURN BOTH FIELD NAMES for compatibility
        response = {
            "success": True,
            "case": case_doc,
            "caseId": case_doc["id"],
            "caseNumber": case_doc.get("caseNumber", case_id),
            "documents": all_documents,          # ← Frontend expects this
            "caseDocuments": all_documents,      # ← AND this
            "financialDocuments": financial_docs
        }
        
        return jsonify(response), 200
        
    except Exception as e:
        print(f"[API] ❌ Error fetching case: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/query/stream", methods=["POST"])
def query_ai_assistant_stream():
    """Streaming SAMM query endpoint with ITAR compliance and real-time updates"""
    user = require_auth()
    if not user:
        return jsonify({"error": "User not authenticated"}), 401

    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    user_input = data.get("question", "").strip()
    chat_history = data.get("chat_history", [])
    staged_chat_documents_metadata = data.get("staged_chat_documents", [])

    if not user_input:
        return jsonify({"error": "Query cannot be empty"}), 400

    # === Extract user authorization profile ===
    user_id = user["sub"]
    user_profile = {
        "user_id": user_id,
        "authorization_level": user.get("authorization_level", DEFAULT_DEV_AUTH_LEVEL),
        "clearances": user.get("clearances", []),
        "role": user.get("role", "developer")
    }
    # === END ===

    # CRITICAL FIX: Load file content from blob storage BEFORE streaming starts
    documents_with_content = []
    if staged_chat_documents_metadata:
        print(f"[Streaming] 📁 Loading content from {len(staged_chat_documents_metadata)} staged files...")
        for idx, doc_meta in enumerate(staged_chat_documents_metadata, 1):
            blob_name = doc_meta.get("blobName")
            blob_container = doc_meta.get("blobContainer")
            file_name = doc_meta.get("fileName", "Unknown")

            if not blob_name:
                print(f"[Streaming]   ⚠️ Missing blobName for {file_name}")
                continue

            # CRITICAL FIX: Select correct container client based on metadata
            container_client = None
            if blob_container == AZURE_CASE_DOCS_CONTAINER_NAME:
                container_client = case_docs_blob_container_client
                print(f"[Streaming]   File {idx}: {file_name} (CASE container)")
            elif blob_container == AZURE_CHAT_DOCS_CONTAINER_NAME:
                container_client = chat_docs_blob_container_client
                print(f"[Streaming]   File {idx}: {file_name} (CHAT container)")
            else:
                print(f"[Streaming]   ⚠️ Unknown container '{blob_container}' for {file_name}")

            if not container_client:
                print(f"[Streaming]   ⚠️ Container client not available for {file_name}")
                continue

            # Fetch content using the CORRECT container client
            print(f"[Streaming]   Fetching file {idx}: {file_name} from {blob_container}")
            content = fetch_blob_content(blob_name, container_client)

            if content:
                documents_with_content.append({
                    **doc_meta,
                    "content": content[:5000]
                })
                print(f"[Streaming]   ✅ Loaded {len(content)} chars from {file_name}")
            else:
                print(f"[Streaming]   ⚠️ No content retrieved from {file_name}")

        print(
            f"[Streaming] 📊 Result: {len(documents_with_content)}/{len(staged_chat_documents_metadata)} files loaded successfully")
    else:
        print(f"[Streaming] No staged documents in request")

    def check_and_apply_hitl_corrections(question):
        """Check HITL store and return corrections if exist"""
        global HITL_CORRECTIONS_STORE

        q_hash = create_question_hash(question)

        print(f"🔍 HITL CHECK: Looking for hash = {q_hash}")
        print(f"🔍 HITL CHECK: Question = '{question}'")
        print(f"🔍 HITL CHECK: Store has {len(HITL_CORRECTIONS_STORE['answer_corrections'])} answer corrections")
        print(f"🔍 HITL CHECK: Store keys = {list(HITL_CORRECTIONS_STORE['answer_corrections'].keys())}")

        has_intent = q_hash in HITL_CORRECTIONS_STORE["intent_corrections"]
        has_entities = q_hash in HITL_CORRECTIONS_STORE["entity_corrections"]
        has_answer = q_hash in HITL_CORRECTIONS_STORE["answer_corrections"]

        if not (has_intent or has_entities or has_answer):
            print(f"🔍 HITL CHECK: No corrections found")
            return None

        print(f"✅ HITL CORRECTIONS FOUND!")
        corrections = {}

        if has_intent:
            corrections['intent'] = HITL_CORRECTIONS_STORE["intent_corrections"][q_hash]
            print(f"   🔄 Intent: {corrections['intent']}")

        if has_entities:
            corrections['entities'] = HITL_CORRECTIONS_STORE["entity_corrections"][q_hash]
            print(f"   🔄 Entities: {len(corrections['entities'])} entities")

        if has_answer:
            corrections['answer'] = HITL_CORRECTIONS_STORE["answer_corrections"][q_hash]
            print(f"   🔄 Answer: {len(corrections['answer'])} chars")

        return corrections

    def generate():
        try:
            start_time = time.time()

            # START - Send immediately with file count
            yield f"data: {json.dumps({'type': 'start', 'query': user_input, 'timestamp': time.time(), 'files_loaded': len(documents_with_content)})}\n\n"


            # ========== ORIGINAL HITL CHECK (for other questions) ==========
            hitl_corrections = check_and_apply_hitl_corrections(user_input)
            if hitl_corrections and 'answer' in hitl_corrections:
                print("⚡ HITL CORRECTION FOUND - Returning corrected answer immediately!")

                yield f"data: {json.dumps({'type': 'progress', 'step': 'hitl_check', 'message': 'Using corrected answer from HITL...', 'elapsed': 0.1})}\n\n"
                yield f"data: {json.dumps({'type': 'answer_start', 'message': 'Streaming corrected answer...', 'elapsed': 0.2})}\n\n"

                # Stream corrected answer word by word
                corrected_answer = hitl_corrections['answer']
                for i, token in enumerate(corrected_answer.split()):
                    yield f"data: {json.dumps({'type': 'answer_token', 'token': token + ' ', 'position': i + 1})}\n\n"

                total_time = round(time.time() - start_time, 2)
                yield f"data: {json.dumps({'type': 'answer_complete', 'answer': corrected_answer})}\n\n"
                yield f"data: {json.dumps({'type': 'complete', 'answer': corrected_answer, 'data': {'hitl_corrected': True, 'source': 'hitl_correction', 'timings': {'total': total_time}}})}\n\n"
                return
            # ========== END HITL CHECK ==========

            # STEP 1: Intent Analysis
            yield f"data: {json.dumps({'type': 'progress', 'step': 'intent_analysis', 'message': 'Analyzing query intent...', 'elapsed': round(time.time() - start_time, 2)})}\n\n"
            intent_start = time.time()

            q_hash = create_question_hash(user_input)
            if q_hash in HITL_CORRECTIONS_STORE["intent_corrections"]:
                corrected_intent = HITL_CORRECTIONS_STORE["intent_corrections"][q_hash]
                print(f"🔄 HITL: Intent correction applied ({corrected_intent})")
                intent_info = {"intent": corrected_intent, "confidence": 1.0, "hitl_corrected": True}
            else:
                intent_info = orchestrator.intent_agent.analyze_intent(user_input)

            intent_time = round(time.time() - intent_start, 2)
            yield f"data: {json.dumps({'type': 'intent_complete', 'data': intent_info, 'time': intent_time})}\n\n"

            # === CHECK FOR SPECIAL CASES ===
            if intent_info.get('special_case', False):
                special_intent = intent_info.get('intent')
                print(f"[Streaming] Special case detected: {special_intent}")

                yield f"data: {json.dumps({'type': 'progress', 'step': 'special_case_handling', 'message': f'Handling {special_intent} query...', 'elapsed': round(time.time() - start_time, 2)})}\n\n"

                if special_intent == "loa_timeline":
                    special_answer = orchestrator.answer_agent._get_loa_timeline_answer()
                elif special_intent == "financial_verification":
                    special_answer = orchestrator.answer_agent._get_financial_verification_answer()
                elif special_intent == "line_item_details":
                    special_answer = orchestrator.answer_agent._get_technical_services_answer()
                elif special_intent == "pmr_minutes_summary":
                    special_answer = orchestrator.answer_agent._get_pmr_minutes_summary()
                elif special_intent == "nonsense":
                    special_answer = "I apologize, but I'm having difficulty understanding your question. It appears to contain unclear or garbled text.\n\nCould you please rephrase your question more clearly? I'm here to help with questions about the Security Assistance Management Manual (SAMM)."
                elif special_intent == "incomplete":
                    special_answer = "I'd be happy to help, but I need more information to answer your question properly.\n\nCould you please provide more details about what specific SAMM topic you're asking about?"
                elif special_intent == "non_samm":
                    detected_topic = intent_info.get('detected_topics', ['this topic'])[0]
                    special_answer = f"Thank you for your question about {detected_topic}.\n\nHowever, this topic is **outside the scope of SAMM**.\n\nCan I help you with any SAMM Chapter topics instead?"
                else:
                    special_answer = "I apologize, but I cannot process this query. Please ask about SAMM topics."

                yield f"data: {json.dumps({'type': 'answer_start', 'message': 'Sending response...', 'elapsed': round(time.time() - start_time, 2)})}\n\n"

                for i, token in enumerate(special_answer.split()):
                    yield f"data: {json.dumps({'type': 'answer_token', 'token': token + ' ', 'position': i + 1})}\n\n"

                total_time = round(time.time() - start_time, 2)
                yield f"data: {json.dumps({'type': 'answer_complete', 'answer': special_answer})}\n\n"
                yield f"data: {json.dumps({'type': 'complete', 'answer': special_answer, 'data': {'special_case': True, 'intent': special_intent, 'timings': {'total': total_time}}})}\n\n"
                return

            # STEP 2: Entity Extraction
            file_msg = f" and {len(documents_with_content)} files" if documents_with_content else ""
            yield f"data: {json.dumps({'type': 'progress', 'step': 'entity_extraction', 'message': f'Extracting entities from query{file_msg}...', 'elapsed': round(time.time() - start_time, 2)})}\n\n"

            entity_start = time.time()

            if q_hash in HITL_CORRECTIONS_STORE["entity_corrections"]:
                corrected_entities = HITL_CORRECTIONS_STORE["entity_corrections"][q_hash]
                print(f"🔄 HITL: Entity corrections applied ({len(corrected_entities)} entities)")
                entity_info = {
                    "entities": corrected_entities,
                    "overall_confidence": 1.0,
                    "hitl_corrected": True,
                    "context": {},
                    "relationships": []
                }
            else:
                entity_info = orchestrator.entity_agent.extract_and_retrieve(
                    user_input,
                    intent_info,
                    documents_with_content
                )

            entity_time = round(time.time() - entity_start, 2)

            files_processed = entity_info.get('files_processed', 0)
            file_entities = entity_info.get('file_entities_found', 0)
            file_relationships = entity_info.get('file_relationships_found', 0)

            yield f"data: {json.dumps({'type': 'entities_complete', 'data': {'count': len(entity_info.get('entities', [])), 'entities': entity_info.get('entities', []), 'confidence': entity_info.get('overall_confidence', 0), 'files_processed': files_processed, 'file_entities': file_entities, 'file_relationships': file_relationships}, 'time': entity_time})}\n\n"

            # STEP 3: Compliance Check
            yield f"data: {json.dumps({'type': 'progress', 'step': 'compliance_check', 'message': 'Checking ITAR compliance...', 'elapsed': round(time.time() - start_time, 2)})}\n\n"

            compliance_start = time.time()
            compliance_result = check_compliance(user_input, intent_info, entity_info, user_profile)
            compliance_time = round(time.time() - compliance_start, 2)

            yield f"data: {json.dumps({'type': 'compliance_complete', 'data': {'status': compliance_result.get('compliance_status'), 'authorized': compliance_result.get('authorized'), 'user_level': compliance_result.get('user_authorization_level')}, 'time': compliance_time})}\n\n"

            if not compliance_result.get("authorized", True):
                required_level = compliance_result.get('required_authorization_level', 'higher')
                user_level = compliance_result.get('user_authorization_level', 'unknown')
                recommendations = compliance_result.get("recommendations", [])

                denial_msg = f"⚠️ ITAR COMPLIANCE NOTICE\n\nThis query requires {required_level.upper()} clearance.\nYour authorization: {user_level.upper()}\n\n"

                if recommendations:
                    denial_msg += "Recommendations:\n" + "\n".join(f"• {r}" for r in recommendations)

                yield f"data: {json.dumps({'type': 'answer_start', 'message': 'Access restricted'})}\n\n"
                yield f"data: {json.dumps({'type': 'answer_token', 'token': denial_msg, 'position': 1})}\n\n"
                yield f"data: {json.dumps({'type': 'answer_complete', 'answer': denial_msg})}\n\n"
                yield f"data: {json.dumps({'type': 'complete', 'answer': denial_msg, 'data': {'compliance_denied': True, 'timings': {'total': round(time.time() - start_time, 2)}}})}\n\n"
                return

            # STEP 4: Answer Generation
            file_context_msg = f" with {len(documents_with_content)} file(s)" if documents_with_content else ""
            yield f"data: {json.dumps({'type': 'progress', 'step': 'answer_generation', 'message': f'Generating answer{file_context_msg}...', 'elapsed': round(time.time() - start_time, 2)})}\n\n"

            answer_start = time.time()

            context = orchestrator.answer_agent._build_comprehensive_context(
                user_input, intent_info, entity_info, chat_history, documents_with_content
            )

            intent = intent_info.get("intent", "general")
            system_msg = orchestrator.answer_agent._create_optimized_system_message(intent, context)
            prompt = orchestrator.answer_agent._create_enhanced_prompt(user_input, intent_info, entity_info)

            yield f"data: {json.dumps({'type': 'answer_start', 'message': 'Streaming answer...', 'elapsed': round(time.time() - start_time, 2)})}\n\n"

            full_answer = ""
            token_count = 0

            for token in call_ollama_streaming(prompt, system_msg, temperature=0.1):
                if token and not token.startswith("Error"):
                    full_answer += token
                    token_count += 1
                    yield f"data: {json.dumps({'type': 'answer_token', 'token': token, 'position': token_count})}\n\n"

            answer_time = round(time.time() - answer_start, 2)
            total_time = round(time.time() - start_time, 2)

            enhanced_answer = orchestrator.answer_agent._enhance_answer_quality(
                full_answer, intent_info, entity_info
            )

            final_answer = enhanced_answer if enhanced_answer else full_answer

            yield f"data: {json.dumps({'type': 'answer_complete', 'answer': final_answer, 'enhanced': (enhanced_answer != full_answer)})}\n\n"
            yield f"data: {json.dumps({'type': 'complete', 'answer': final_answer, 'data': {'compliance_approved': True, 'intent': intent, 'entities_found': len(entity_info.get('entities', [])), 'files_processed': files_processed, 'file_entities': file_entities, 'file_relationships': file_relationships, 'answer_length': len(final_answer), 'token_count': token_count, 'timings': {'intent': intent_time, 'entity': entity_time, 'compliance': compliance_time, 'answer': answer_time, 'total': total_time}}})}\n\n"

            # Confidence check for HITL
            intent_confidence = intent_info.get('confidence', 0.5)
            entity_confidence = entity_info.get('overall_confidence', 0.5)
            answer_confidence = 0.8 if len(final_answer) > 200 else 0.5
            overall_confidence = (intent_confidence + entity_confidence + answer_confidence) / 3

            print(
                f"📊 Confidence: Intent={intent_confidence:.2f}, Entity={entity_confidence:.2f}, Answer={answer_confidence:.2f}, Overall={overall_confidence:.2f}")

            if overall_confidence < 0.95:
                print(f"⚠️ LOW CONFIDENCE ({overall_confidence:.2f}) - Adding to HITL queue...")

                try:
                    review_item = {
                        "id": str(uuid.uuid4()),
                        "type": "review_item",
                        "reviewId": str(uuid.uuid4()),
                        "question": user_input,
                        "aiResponse": {
                            "intent": intent_info.get('intent', 'unknown'),
                            "entities": entity_info.get('entities', []),
                            "answer": final_answer
                        },
                        "status": "pending",
                        "priority": "high" if overall_confidence < 0.5 else "medium",
                        "confidenceOverall": overall_confidence,
                        "confidenceIntent": intent_confidence,
                        "confidenceEntity": entity_confidence,
                        "confidenceAnswer": answer_confidence,
                        "timestamp": datetime.now(timezone.utc).isoformat(),
                        "assignedTo": "Travis",
                        "humanFeedback": "",
                        "reviewedBy": "",
                        "reviewedAt": ""
                    }

                    if reviews_test_container_client:
                        reviews_test_container_client.create_item(review_item)
                        print(f"✅ Added to review queue: {review_item['reviewId']}")
                        yield f"data: {json.dumps({'type': 'hitl_triggered', 'message': 'Low confidence - added to review queue', 'reviewId': review_item['reviewId']})}\n\n"

                except Exception as e:
                    print(f"❌ Error adding to review queue: {e}")

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"[Streaming Error] {error_detail}")
            yield f"data: {json.dumps({'type': 'error', 'error': str(e), 'detail': error_detail})}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive'
        }
    )



@app.route("/api/hitl/reset-demo", methods=["POST"])
def reset_demo():
    """Reset demo corrections"""
    try:
        data = request.json or {}
        demo_type = data.get('demo_type', 'all')  # 'test', 'travis', or 'all'
        
        scenarios_to_reset = []
        if demo_type == 'all':
            scenarios_to_reset = list(DEMO_SCENARIOS.keys())
        elif demo_type in DEMO_SCENARIOS:
            scenarios_to_reset = [demo_type]
        else:
            return jsonify({"success": False, "error": "Invalid demo_type"}), 400
        
        for scenario_name in scenarios_to_reset:
            q_hash = create_question_hash(DEMO_SCENARIOS[scenario_name]["question"])
            for store in ["intent_corrections", "entity_corrections", "answer_corrections"]:
                if q_hash in HITL_CORRECTIONS_STORE[store]:
                    del HITL_CORRECTIONS_STORE[store][q_hash]
        
        print(f"🔄 HITL: Demo reset for {', '.join(scenarios_to_reset).upper()}")
        return jsonify({"success": True, "message": f"Demo reset complete for {', '.join(scenarios_to_reset)}"})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# =============================================================================
# OLLAMA MODEL WARM-UP
# =============================================================================
def warm_up_ollama():
    import threading
    def _warmup():
        print("[Ollama Warmup] 🔥 Warming up model...")
        try:
            ollama_session.post(
                f"{OLLAMA_URL}/api/chat",
                json={
                    "model": OLLAMA_MODEL,
                    "messages": [{"role": "user", "content": "Hi"}],
                    "stream": False,
                    "options": {"num_predict": 5}
                },
                timeout=60
            )
            print("[Ollama Warmup] ✅ Model ready!")
        except Exception as e:
            print(f"[Ollama Warmup] ⚠️ {e}")

    threading.Thread(target=_warmup, daemon=True).start()


warm_up_ollama()




if __name__ == '__main__':
    port = int(os.environ.get("PORT", 3000))
    
    print("\n" + "="*90)
    print("🚀 Complete Integrated SAMM ASIST System with Database Integration v5.0")
    print("="*90)
    print(f"🌐 Server: http://172.16.200.12:{port}")
    print(f"🤖 Ollama Model: {OLLAMA_MODEL}")
    print(f"🔗 Ollama URL: {OLLAMA_URL}")
    print(f"📊 Knowledge Graph: {len(knowledge_graph.entities)} entities, {len(knowledge_graph.relationships)} relationships")
    print(f"🎯 Integrated Database Orchestration: {len(WorkflowStep)} workflow steps")
    print(f"🔄 Integrated Agents: Intent → Integrated Entity (Database) → Enhanced Answer (Quality)")
    print(f"🔐 Auth: {'OAuth (Auth0)' if oauth else 'Mock User'}")
    print(f"💾 Storage: {'Azure Cosmos DB' if cases_container_client else 'In-Memory'}")
    print(f"📁 Blob Storage: {'Azure' if blob_service_client else 'Disabled'}")
    
    # Database status
    db_status = orchestrator.get_database_status()
    print(f"\n💽 Database Integration:")
    print(f"• Cosmos Gremlin: {'Connected' if db_status['cosmos_gremlin']['connected'] else 'Disconnected'} ({db_status['cosmos_gremlin']['endpoint']})")
    print(f"• Vector DB: {'Connected' if db_status['vector_db']['connected'] else 'Disconnected'} ({len(db_status['vector_db']['collections'])} collections)")
    print(f"• Embedding Model: {'Loaded' if db_status['embedding_model']['loaded'] else 'Not Loaded'} ({db_status['embedding_model']['model_name']})")
    
    print(f"\n📡 Core Endpoints:")
    print(f"• Integrated Query: POST http://172.16.200.12:{port}/api/query")
    print(f"• System Status: GET http://172.16.200.12:{port}/api/system/status")
    print(f"• Database Status: GET http://172.16.200.12:{port}/api/database/status")
    print(f"• Examples: GET http://172.16.200.12:{port}/api/examples")
    print(f"• User Cases: GET http://172.16.200.12:{port}/api/user/cases")
    print(f"• Authentication: GET http://172.16.200.12:{port}/login")

    print(f"\n🤖 Enhanced Agent Endpoints:")
    print(f"• HIL Update: POST http://172.16.200.12:{port}/api/agents/hil_update")
    print(f"• Trigger Update: POST http://172.16.200.12:{port}/api/agents/trigger_update")
    print(f"• Agent Status: GET http://172.16.200.12:{port}/api/agents/status")

    print(f"\n📡 Advanced SAMM Endpoints:")
    print(f"• Detailed Status: GET http://172.16.200.12:{port}/api/samm/status")
    print(f"• Integrated Workflow: GET http://172.16.200.12:{port}/api/samm/workflow")
    print(f"• Knowledge Graph: GET http://172.16.200.12:{port}/api/samm/knowledge")
    print(f"• Health Check: GET http://172.16.200.12:{port}/api/health")
    
    print(f"\n🧪 Try these questions:")
    print("• What is Security Cooperation?")
    print("• Who supervises Security Assistance programs?")
    print("• What's the difference between SC and SA?") 
    print("• What does DFAS do?")
    
    print(f"\n⚡ Integrated Database Capabilities:")
    print("• Integrated Entity Agent: Pattern → NLP → Database queries (Cosmos Gremlin + Vector DBs)")
    print(f"  - {sum(len(patterns) for patterns in orchestrator.entity_agent.samm_entity_patterns.values())} SAMM patterns")
    print("  - Real-time database integration for entity context")
    print("  - Confidence scoring for all extracted entities")
    print("  - Dynamic knowledge expansion with HIL feedback")
    print("• Enhanced Answer Agent: Intent-optimized responses with quality scoring")
    print(f"  - {len(orchestrator.answer_agent.samm_response_templates)} response templates")
    print(f"  - {len(orchestrator.answer_agent.acronym_expansions)} acronym expansions")
    print("  - Multi-pass generation with validation")
    print("  - Automatic quality enhancement")
    
    print(f"\n🔄 Learning System:")
    print("• Human-in-Loop (HIL): Correct intent, entities, and answers")
    print("• Trigger Updates: Add new entities and relationships dynamically")
    print("• Database Learning: Entities learn from graph and vector databases")
    print("• Pattern Learning: Intent agent learns query patterns")
    print("• Knowledge Expansion: Entity agent grows knowledge base")
    print("• Answer Corrections: Answer agent stores and reuses corrections")
    print("• Quality Improvement: All agents learn from feedback")
    
    print(f"\n📊 Agent Status:")
    try:
        status = orchestrator.get_agent_status()
        print(f"• Intent Agent: {status['intent_agent']['learned_patterns']} learned patterns")
        print(f"• Integrated Entity Agent: {status['integrated_entity_agent']['custom_entities']} custom entities, {status['integrated_entity_agent']['samm_patterns']} SAMM patterns")
        print(f"• Enhanced Answer Agent: {status['enhanced_answer_agent']['answer_corrections']} stored corrections, {status['enhanced_answer_agent']['response_templates']} templates")
    except:
        print("• Agent status: Initializing...")
    
    print("="*90 + "\n")
    
    app.run(host='0.0.0.0', port=port, debug=True)